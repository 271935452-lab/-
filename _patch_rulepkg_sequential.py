# -*- coding: utf-8 -*-
"""Add 规则包类型说明 sheet; enrich S016 with sequential CS→sales→deduction demo numbers."""
import shutil

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

SRC = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并.xlsx"
BAK = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_before_rulepkg_patch.xlsx"
OUT = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_rulepkg序贯示例.xlsx"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def patch():
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=False)

    # --- New sheet: 规则包类型说明 (after index) ---
    idx_name = wb.sheetnames[0]
    if "规则包类型说明" in wb.sheetnames:
        del wb["规则包类型说明"]
    ws_new = wb.create_sheet("规则包类型说明", 1)
    ws_new.freeze_panes = "A4"
    rows = [
        ["规则包内编排类型", "含义（原型约定）", "典型示例"],
        [
            "顺序执行",
            "包内子规则按固定次序逐段结算；前一段完成后进入下一段。前置段常用于客服计提、区间门禁校验等，再计算业务员提成，最后统一经过万能池「提成扣」等发放层冲减。",
            "直客有客服：①客服比例规则 →②业务员入职/提成区间命中档位 →③业务员毛利阶梯或固定点 →④万能池提成扣。",
        ],
        [
            "互斥组",
            "同一互斥组编号内至多命中一条子规则（按优先级/特异性裁决）。",
            "多套毛利阶梯模板二选一；客服「有/无业务员」模板互斥。",
        ],
        [
            "相加组",
            "同组可同时命中多条子规则，各段计提结果代数相加后，再进入扣减池。",
            "毛利阶梯 + KA 专项加点并行叠加（示意）。",
        ],
        [],
        [
            "补充（直客业务员 · 顺序执行）",
            None,
            None,
        ],
        [
            None,
            "场景表 S016 用虚拟数据演示：同运单 WB-SZ-BT-202603001 上先跑客服段 CS_WITH_SALES_RATE（0.75%×归因应收），再跑业务员段 DIRECT_SALES_GP_TIER（净利基数×38.5%），业务员侧再减万能池提成扣 200 → 业务员应发 2880 CNY；客服段 396 CNY 计入客服员工台账，与业务员索引行分开汇总。",
            None,
        ],
    ]
    for r_i, row in enumerate(rows, start=1):
        for c_i, val in enumerate(row, start=1):
            c = ws_new.cell(r_i, c_i, val)
            c.border = thin_border()
            if r_i == 1:
                c.font = Font(bold=True)
            if c_i == 2 and val and isinstance(val, str) and len(val) > 40:
                c.alignment = Alignment(wrap_text=True, vertical="top")
    ws_new.merge_cells("A7:A8")
    ws_new["A7"].alignment = Alignment(vertical="top", wrap_text=True)
    ws_new.merge_cells("B7:B8")
    ws_new["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_new.column_dimensions["A"].width = 14
    ws_new.column_dimensions["B"].width = 62
    ws_new.column_dimensions["C"].width = 42

    # --- S016 ---
    ws = wb["S016"]
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    long_a3 = (
        "计薪月=2026-03；分公司=深圳；部门=深圳坂田；小组=—；提成岗位=直客业务员；映射生效=2024-03-01~2025-04-08；"
        "提成计算=2024-03-01~长期。【真实测算】账期 2026-03；运单底表：测试数据_运单底表_逻辑加工表人员映射.xlsx；"
        "万能调整池：测试数据_手工万能调整池_逻辑加工表人员映射.xlsx。步骤6/7为按列公式演示值。\n"
        "【规则包】编码 RP_SZ_BT_DIRECT_SEQ（示意），编排类型=顺序执行："
        "①先执行客服子规则（段1，计提记入客服账号）；②再执行业务员子规则并匹配入职/提成区间档位（段2）；"
        "③业务员侧毛利提成试算后，再扣万能池「提成扣」（段3·发放层）。互斥组/相加组在本包其他映射行另行示意。"
    )
    ws["A3"] = long_a3

    ws.insert_rows(6)

    # Row 6：客服段（A-I 基础列 + J-P 步骤列共 16 列）
    cs_row = [
        "坂田客服Demo",
        "深圳坂田客服（演示）",
        "WB-SZ-BT-202603001",
        "客服",
        52800,
        None,
        12800,
        44,
        None,
        "命中映射「坂田客服Demo+深圳坂田客服（演示）」（示意）",
        "映射生效∩2026-03：通过（示意）",
        "提成区间：覆盖（示意）；在职·发放正常（示意）",
        "【顺序执行·段1】规则包 RP_SZ_BT_DIRECT_SEQ · 命中 CS_WITH_SALES_RATE（客服计提先行）",
        "基数=本运单归因应收 52800.00（净应收汇总示意）",
        396,
        396,
    ]
    for col, val in enumerate(cs_row, start=1):
        ws.cell(6, col, val)

    # Row 7：业务员
    ws["A7"] = "坂田直客业务员1"
    ws["B7"] = "净利润固定提成点2024/3/1开始2025/4/8止"
    ws["C7"] = "WB-SZ-BT-202603001"
    ws["D7"] = "业务员"
    ws["E7"] = 52800
    ws["F7"] = 8600
    ws["G7"] = 12800
    ws["H7"] = 44
    ws["I7"] = None
    sales_steps = [
        "命中映射「坂田直客业务员1+净利润固定提成点…」",
        "映射生效∩2026-03：通过（示意）",
        "提成计算区间：覆盖（示意）；在职·发放正常（示意）",
        "【顺序执行·段2】段1客服计提已完成（396.00）→本段命中 DIRECT_SALES_GP_TIER；"
        "业务员入职/提成区间 [2024-03-01,2025-04-08] → 演示阶梯系数 38.5%",
        "基数=计算净利润 8600.00 − 表外毛利扣 600.00 = 8000.00",
        3080,
        2880,
    ]
    for i, txt in enumerate(sales_steps, start=10):
        ws.cell(7, i, txt)

    # Row 8：合计
    ws["A8"] = "合计（演示）"
    ws.merge_cells("A8:H8")
    ws["J8"] = (
        "Σ步骤7（业务员本人）= 2880.00 CNY。"
        "同运单顺序包段1客服应发 396.00 记入客服 Demo 台账，不参与本合计；索引表「坂田直客业务员1」与测算汇总取其本人 2880。"
    )
    ws.merge_cells("J8:P8")
    ws["J8"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("A1:P1")
    ws.merge_cells("A3:P3")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")

    # --- 索引 sheet row 坂田直客业务员1 / S016 -> 2880 ---
    ws_idx = wb[idx_name]
    for r in range(2, ws_idx.max_row + 1):
        sid = ws_idx.cell(r, 7).value
        if sid == "S016":
            ws_idx.cell(r, 8, 2880)
            break

    # --- 测算汇总 sheet ---
    sum_name = wb.sheetnames[2]  # after inserting sheet at 1, was 测算汇总 - verify
    # sheetnames: 0索引 1规则包类型说明 2测算汇总 -插入后原来是第二个变成第三个
    ws_sum = wb[wb.sheetnames[2]]
    if ws_sum.cell(3, 6).value in (None, ""):
        ws_sum.cell(3, 6, "备注（原型）")
    for r in range(4, ws_sum.max_row + 1):
        name = ws_sum.cell(r, 2).value
        rule = ws_sum.cell(r, 5).value
        if name and "坂田" in str(name) and "业务员1" in str(name):
            ws_sum.cell(r, 4, 2880)
            note = ws_sum.cell(r, 6).value or ""
            if isinstance(note, str) and note.strip():
                ws_sum.cell(r, 6, str(note) + "；顺序包示意")
            else:
                ws_sum.cell(r, 6, "顺序包：客服段→业务员段→提成扣")
            break

    try:
        wb.save(SRC)
        print("Saved:", SRC)
    except PermissionError:
        wb.save(OUT)
        print("Locked; saved:", OUT)
    wb.close()
    print("Backup:", BAK)


if __name__ == "__main__":
    patch()
