# -*- coding: utf-8 -*-
"""生成《各部门需求清单与汇报计划》Excel（提成管理二期汇报用）。

计划与优先级分两阶段维护：
  1）自拟：您在「分部门需求清单」「里程碑计划」中先填「自拟*」列与「计划阶段=自拟」；
  2）定稿：与领导、各部门确认后，再填写「确认后*」列，并将「计划阶段」改为「已定稿」。

同时从同目录「需求收集表.xlsx」读取第 1、2 个工作表，写入「需求收集汇总」「财务需求汇总」：
  · sheet1：在「功能名称」后插入「归纳部门」，按「功能名称」文本命中关键词（长词优先）归类；
  · sheet2：全部数据行「归纳部门」固定为「财务部」（列插在第一列）。
另生成「台账归部门规则」「台账按归纳部门统计」便于核对。
根据「分部门需求清单」模板数据自动生成「部门汇总（汇报用）」：按部门条数、P0/P1/P2 与要点合并，便于口头汇报。
"""
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIR = Path(__file__).resolve().parent
OUT = DIR / "各部门需求清单与汇报计划.xlsx"
SRC_REQUIREMENTS = DIR / "需求收集表.xlsx"

CELL_MAX_LEN = 32000  # Excel 单单元格上限约 32767，略留余量

HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 向领导口头汇报时建议的部门顺序（未出现在此列表的部门排在末尾，按名称排序）
PRESENTATION_DEPT_ORDER = [
    "管理层",
    "财务部",
    "产品/项目",
    "技术/数据",
    "销售/业务",
    "客服/运营",
    "人力/薪酬",
]

# sheet1「功能名称」→ 归纳部门：整格文本「包含」即命中；同一格多关键词时**长词优先**（先列长的再列短的）
FUNCTION_NAME_DEPT_RULES = [
    ("客户端小程序", "技术/数据"),
    ("业务员成本", "销售/业务"),
    ("运单模板", "技术/数据"),
    ("运单模块", "技术/数据"),
    ("财务模块", "财务部"),
    ("报表模块", "财务部"),
    ("员工管理", "人力/薪酬"),
    ("销售报价", "销售/业务"),
    ("客户管理", "销售/业务"),
    ("客服模块", "客服/运营"),
    ("产品模块", "产品/项目"),
    ("配置模块", "技术/数据"),
    ("操作模块", "客服/运营"),
    ("海外模块", "销售/业务"),
    ("待配仓", "客服/运营"),
    ("其他", "产品/项目"),
]
_LEDGER_RULES_SORTED = sorted(
    FUNCTION_NAME_DEPT_RULES, key=lambda kv: len(kv[0]), reverse=True
)
DEFAULT_LEDGER_DEPT_SHEET1 = "产品/项目"
LEDGER_DEPT_FINANCE_SHEET2 = "财务部"


def _cell_str(v):
    if v is None:
        return ""
    s = str(v).strip()
    if len(s) > CELL_MAX_LEN:
        return s[: CELL_MAX_LEN - 20] + "\n…(内容过长已截断)"
    return s


def autosize(ws, max_width=56, max_scan_rows=400):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        w = 10
        for row in range(1, min(ws.max_row + 1, max_scan_rows)):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v)
            w = max(
                w,
                min(
                    max_width,
                    len(s) // 2 + 12 if any(ord(c) > 127 for c in s) else len(s) + 2,
                ),
            )
        ws.column_dimensions[letter].width = w


def add_headers(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN


def read_source_sheet(path: Path, sheet_index: int):
    """返回 (表名, 二维列表，首行为表头)。"""
    if not path.is_file():
        return None, []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_index < 0 or sheet_index >= len(wb.sheetnames):
            return None, []
        name = wb.sheetnames[sheet_index]
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(tuple(_cell_str(c) for c in row))
        return name, rows
    finally:
        wb.close()


def _find_header_col(header_row, *candidates):
    hdr = [str(h).strip() if h is not None else "" for h in header_row]
    for cand in candidates:
        for i, h in enumerate(hdr):
            if h == cand:
                return i
    return None


def classify_dept_from_function_name(fn: str) -> str:
    s = (fn or "").strip()
    if not s:
        return "（功能名称为空）"
    for key, dept in _LEDGER_RULES_SORTED:
        if key in s:
            return dept
    return DEFAULT_LEDGER_DEPT_SHEET1


def _row_any_non_empty(r, skip_indices=None):
    skip = set(skip_indices or ())
    for i, x in enumerate(r):
        if i in skip:
            continue
        if str(x).strip():
            return True
    return False


def augment_sheet_requirement_collect(rows):
    """sheet1：在「功能名称」列后插入「归纳部门」。"""
    if not rows:
        return rows
    header = list(rows[0])
    idx_fn = _find_header_col(tuple(header), "功能名称")
    if idx_fn is None:
        idx_fn = 3
    new_header = header[: idx_fn + 1] + ["归纳部门"] + header[idx_fn + 1 :]
    out = [tuple(_cell_str(x) for x in new_header)]
    for r in rows[1:]:
        lst = [_cell_str(x) for x in r]
        while len(lst) <= idx_fn:
            lst.append("")
        fn = lst[idx_fn]
        dept = classify_dept_from_function_name(fn)
        new_r = lst[: idx_fn + 1] + [dept] + lst[idx_fn + 1 :]
        out.append(tuple(new_r))
    return out


def augment_sheet_finance_require(rows):
    """sheet2：首列「归纳部门」全部为财务部。"""
    if not rows:
        return rows
    header = [_cell_str(x) for x in rows[0]]
    new_header = ["归纳部门"] + header
    out = [tuple(new_header)]
    for r in rows[1:]:
        lst = [_cell_str(x) for x in r]
        while len(lst) < len(header):
            lst.append("")
        lst = lst[: len(header)]
        new_r = [LEDGER_DEPT_FINANCE_SHEET2] + lst
        out.append(tuple(new_r))
    return out


def build_ledger_rules_matrix():
    """说明 sheet1 关键词规则 + sheet2 全部归财务。"""
    m = [
        ("项", "说明"),
        ("sheet1（需求收集）", "在「功能名称」后插入列「归纳部门」；按下列关键词在「功能名称」整格中自长到短首次「包含」命中；均未命中则为「产品/项目」"),
        ("sheet2（财务需求）", "全部数据行第一列「归纳部门」固定为「财务部」"),
        ("", ""),
        ("关键词（包含即匹配）", "归纳部门"),
    ]
    for kw, dep in _LEDGER_RULES_SORTED:
        m.append((kw, dep))
    m.append(("（均未命中）", DEFAULT_LEDGER_DEPT_SHEET1))
    return m


def build_ledger_dept_tally(rows_req_aug, rows_fin_aug):
    """按「归纳部门」统计 sheet1 / sheet2 有效行数。"""
    counts = defaultdict(lambda: {"req": 0, "fin": 0})
    if rows_req_aug and len(rows_req_aug) > 1:
        hdr = list(rows_req_aug[0])
        try:
            idx_d = hdr.index("归纳部门")
        except ValueError:
            idx_d = None
        if idx_d is not None:
            for r in rows_req_aug[1:]:
                if len(r) <= idx_d:
                    continue
                if not _row_any_non_empty(r, skip_indices={idx_d}):
                    continue
                d = str(r[idx_d]).strip() or "（未填）"
                counts[d]["req"] += 1
    if rows_fin_aug and len(rows_fin_aug) > 1:
        hdr_f = list(rows_fin_aug[0])
        if hdr_f and hdr_f[0] == "归纳部门":
            for r in rows_fin_aug[1:]:
                if len(r) < 2:
                    continue
                if not _row_any_non_empty(r, skip_indices={0}):
                    continue
                d = str(r[0]).strip() or LEDGER_DEPT_FINANCE_SHEET2
                counts[d]["fin"] += 1
    out = [
        (
            "归纳部门",
            "需求收集条数（sheet1）",
            "财务需求条数（sheet2）",
            "合计",
        )
    ]
    def dept_sort_key(d):
        try:
            return (0, PRESENTATION_DEPT_ORDER.index(d), d)
        except ValueError:
            return (1, d)

    order = sorted(counts.keys(), key=dept_sort_key)
    tot_req = tot_fin = 0
    for d in order:
        a = counts[d]
        tot_req += a["req"]
        tot_fin += a["fin"]
        out.append((d, a["req"], a["fin"], a["req"] + a["fin"]))
    out.append(("合计", tot_req, tot_fin, tot_req + tot_fin))
    return out


def write_matrix_sheet(wb, title, rows, insert_index=None):
    """rows[0] 为表头（可为空则写「列1」…）；后续为数据行。"""
    if not rows:
        ws = wb.create_sheet(title) if insert_index is None else wb.create_sheet(title, insert_index)
        ws.cell(row=1, column=1, value="（源表无数据或文件不存在）").alignment = WRAP
        autosize(ws, max_width=48, max_scan_rows=20)
        return ws

    ncols = max(len(r) for r in rows)
    normalized = []
    for r in rows:
        lst = list(r)
        if len(lst) < ncols:
            lst.extend([""] * (ncols - len(lst)))
        else:
            lst = lst[:ncols]
        normalized.append(tuple(lst))

    ws = wb.create_sheet(title) if insert_index is None else wb.create_sheet(title, insert_index)
    for ri, row in enumerate(normalized, 1):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val if val != "" else None).alignment = WRAP
    autosize(ws, max_width=56, max_scan_rows=min(500, ws.max_row + 1))
    return ws


def summarize_by_department(detail_rows):
    """根据「分部门需求清单」数据行（元组：部门,编号,要点,自拟优先级,...）生成部门汇总行。

    返回 [(汇报顺序, 部门, 条数, P0, P1, P2, 其他优先级, 编号一览, 要点合并, 一句话汇报), ...]
    """
    groups = defaultdict(
        lambda: {
            "n0": 0,
            "n1": 0,
            "n2": 0,
            "n_other": 0,
            "ids": [],
            "lines": [],
        }
    )
    for row in detail_rows:
        dept = (row[0] or "").strip() or "（未分部门）"
        rid = _cell_str(row[1])
        point = _cell_str(row[2])
        pr = (row[3] or "").strip().upper()
        g = groups[dept]
        g["ids"].append(rid)
        g["lines"].append(f"【{rid}】{point}")
        if pr.startswith("P0"):
            g["n0"] += 1
        elif pr.startswith("P1"):
            g["n1"] += 1
        elif pr.startswith("P2"):
            g["n2"] += 1
        else:
            g["n_other"] += 1

    def dept_sort_key(d):
        try:
            return (0, PRESENTATION_DEPT_ORDER.index(d), d)
        except ValueError:
            return (1, d)

    ordered = sorted(groups.keys(), key=dept_sort_key)
    out = []
    for seq, dept in enumerate(ordered, 1):
        g = groups[dept]
        n = len(g["ids"])
        ids_join = "、".join(g["ids"])
        merged = "\n".join(g["lines"])
        if len(merged) > 12000:
            merged = merged[:11980] + "\n…(以下截断，详见「分部门需求清单」)"
        preview = "；".join([ln.replace("\n", " ")[:48] for ln in g["lines"][:2]])
        if len(g["lines"]) > 2:
            preview += f"…等共{n}条"
        pr_tail = ""
        if g["n_other"]:
            pr_tail = f"；未标 P0-P2 共{g['n_other']}条"
        one_liner = (
            f"{dept}共{n}条（P0 {g['n0']} / P1 {g['n1']} / P2 {g['n2']}{pr_tail}），"
            f"编号：{ids_join}。摘要：{preview}"
        )
        if len(one_liner) > 3000:
            one_liner = one_liner[:2980] + "…"
        out.append(
            (
                seq,
                dept,
                n,
                g["n0"],
                g["n1"],
                g["n2"],
                g["n_other"],
                ids_join,
                merged,
                one_liner,
            )
        )
    return out


def main():
    wb = Workbook()

    # --- Sheet: 使用说明 ---
    ws0 = wb.active
    ws0.title = "使用说明"
    lines = [
        "《各部门需求清单与汇报计划》使用说明",
        "",
        "一、计划与优先级（两阶段，请先自拟、再定稿）",
        "1. 第一阶段（自拟）：由您先在「分部门需求清单」填写「自拟优先级」「自拟计划完成」「对接人」等；在「里程碑计划」填写「自拟计划起止」「负责人」。"
        "「计划阶段」列默认为「自拟」，表示尚未经领导与各部门确认。",
        "2. 第二阶段（定稿）：与领导、各部门对齐后，再填写「确认后优先级」「确认后计划完成」（及里程碑「确认后计划起止」），将「计划阶段」改为「已定稿」，并在「确认纪要」中摘录结论或纪要编号。",
        "3. 优先级口径（供自拟与定稿共用）：P0=不上线会阻塞月结或发放；P1=效率与争议风险；P2=体验优化或后续迭代。定稿时可调整级别，以确认后列为准。",
        "",
        "二、其他说明",
        "4. 现金流量、工单、询价等若独立子项目，建议在「风险与依赖」中标注系统边界，避免主项目范围蔓延。",
        "5. 「需求收集汇总」「财务需求汇总」由脚本从同目录《需求收集表.xlsx》第 1、2 个工作表自动带入，并写入「归纳部门」：sheet1 按「功能名称」关键词归类（见「台账归部门规则」）；sheet2 全部归为「财务部」。另生成「台账按归纳部门统计」可快速看各部门台账条数。源文件须与本脚本同目录，并保持前两表顺序。",
        "6. 详细步骤见工作表「计划定稿流程」。",
        "7. 「部门汇总（汇报用）」由脚本根据「分部门需求清单」自动生成：向领导汇报时可先按「汇报顺序」逐行过一遍「一句话汇报」列；需要展开时再到「分部门需求清单」。台账两页已带「归纳部门」，可与本表对照；会上可不展开原始人名列，会后点对点核对。",
        "",
        "生成脚本：生成_各部门需求清单与汇报计划_excel.py",
    ]
    for i, t in enumerate(lines, 1):
        ws0.cell(row=i, column=1, value=t).alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 92

    # --- 从需求收集表.xlsx 读取前两表（先读入内存，稍后插入到固定位置之后） ---
    name0, rows0 = read_source_sheet(SRC_REQUIREMENTS, 0)
    name1, rows1 = read_source_sheet(SRC_REQUIREMENTS, 1)

    # --- Sheet: 计划定稿流程（自拟 → 确认 → 定稿） ---
    ws_flow = wb.create_sheet("计划定稿流程", 1)
    flow_lines = [
        "计划定稿流程（建议）",
        "",
        "【自拟】您维护「分部门需求清单」「里程碑计划」中的自拟列，计划阶段填「自拟」。",
        "【汇报】口头对齐可先打开「部门汇总（汇报用）」再下钻「分部门需求清单」。",
        "【征求意见】将本表（或 PDF）发给各部门负责人，收集书面/会议反馈，可记在「确认纪要」或单独附件。",
        "【领导确认】召开短会或与领导对齐范围、资源、优先级；更新「确认后」列。",
        "【定稿】将「计划阶段」统一改为「已定稿」，作为对开发与测试的基线；后续变更走变更记录。",
        "",
        "「计划阶段」建议取值：自拟 | 征求意见中 | 已定稿",
    ]
    for i, t in enumerate(flow_lines, 1):
        ws_flow.cell(row=i, column=1, value=t).alignment = Alignment(wrap_text=True, vertical="top")
    ws_flow.column_dimensions["A"].width = 88

    draft = "自拟"
    rows = [
        ("财务部", "F01", "运单底表增加服务商字段；内陆费按服务商配置匹配", "P0", "", "底表明细与财务核对一致", "", "", draft, "", ""),
        ("财务部", "F02", "提成明细展示内陆费、操作费单号级明细；体现个税", "P0", "", "与凭证/工资计提可对账", "", "", draft, "", ""),
        ("财务部", "F03", "回款率、未回款明细、杂费创建时间等，支持调账与说明", "P0", "", "可追溯调整记录", "", "", draft, "", ""),
        ("财务部", "F04", "律师费、仓储费等从财务凭证按费用类型+付款时间进基数口径", "P0", "", "口径文档+样例数据签字", "", "", draft, "", ""),
        ("财务部", "F05", "应发/实发提成、手动放行或不计提原因、发放日志", "P0", "", "审计字段齐全", "", "", draft, "", ""),
        ("财务部", "F06", "现金流量表、银行流水认领与对账规则", "P1", "", "认领规则与对账报表", "", "", draft, "", ""),
        ("财务部", "F07", "追溯月份统一提成调整、调整原因留痕", "P1", "", "调整单或等效日志", "", "", draft, "", ""),
        ("人力/薪酬", "H01", "人员映射增加结算提成时间等字段，默认入职可改", "P0", "", "与薪酬政策一致", "", "", draft, "", ""),
        ("人力/薪酬", "H02", "离职状态在提成明细体现；离职发放判断规则与权限", "P0", "", "流程说明+权限配置", "", "", draft, "", ""),
        ("人力/薪酬", "H03", "绩效系数导入与规则包绩效挂钩一致", "P0", "", "导入模板+校验通过", "", "", draft, "", ""),
        ("人力/薪酬", "H04", "部分扣社保毛利、部分不扣等场景在规则/基数可表达", "P1", "", "用例+试算共签", "", "", draft, "", ""),
        ("销售/业务", "S01", "规则包命中：分公司、角色、账期、入职时段、员工状态等", "P0", "", "命中说明可给员工查阅", "", "", draft, "", ""),
        ("销售/业务", "S02", "回款截止时间（如15/20号或自定义）在规则包层配置", "P0", "", "与签约/回款政策一致", "", "", draft, "", ""),
        ("销售/业务", "S03", "客户类型等维度在规则包配置；收敛易混回款配置项", "P1", "", "PRD与原型一致", "", "", draft, "", ""),
        ("销售/业务", "S04", "个别业务员刷新、接近回款率标记等运营诉求", "P1", "", "操作说明", "", "", draft, "", ""),
        ("销售/业务", "S05", "多角色、多规则包、客服与业务员并存等复杂场景", "P0", "", "场景清单+试算用例通过", "", "", draft, "", ""),
        ("客服/运营", "C01", "客服提成先于业务员等执行顺序在规则中心可配置/可说明", "P0", "", "与现行口头规则书面一致", "", "", draft, "", ""),
        ("客服/运营", "C02", "工单附加费、拦截费等是否进表外或运单（系统边界）", "P1", "", "书面边界：提成系统 vs 工单", "", "", draft, "", ""),
        ("客服/运营", "C03", "港前港后、索赔 SLA 等业务流程与数据回写", "P2", "", "若本期不做则列入不做清单", "", "", draft, "", ""),
        ("产品/项目", "P01", "规则中心各页面 PRD（按一期模板），含联动逻辑", "P0", "", "每页 PRD 评审通过", "", "", draft, "", ""),
        ("产品/项目", "P02", "试算顺序、公式、测试数据文档（开发/测试/财务共用）", "P0", "", "每场景可复算", "", "", draft, "", ""),
        ("产品/项目", "P03", "原型与 PRD 同步（字段增减、回款与客户类型等）", "P0", "", "变更记录可追溯", "", "", draft, "", ""),
        ("技术/数据", "T01", "映射、规则包、规则按账期独立生效；试算顺序：映射→规则包→规则", "P0", "", "自动化用例通过", "", "", draft, "", ""),
        ("技术/数据", "T02", "同角色同规则包互斥；同组织同角色同时仅一条生效规则包", "P0", "", "冲突检测与提示", "", "", draft, "", ""),
        ("技术/数据", "T03", "计费日志格式统一、可导出", "P1", "", "样例日志验收", "", "", draft, "", ""),
        ("技术/数据", "T04", "月度归档、运单底表增加内陆费/操作费及说明", "P0", "", "财务联调通过", "", "", draft, "", ""),
        ("技术/数据", "T05", "与询价、工单等外围系统接口清单（若本期对接）", "P1", "", "接口表+Mock/联调", "", "", draft, "", ""),
        ("管理层", "M01", "一期/二期范围与「不做清单」", "P0", "", "范围公告或纪要", "", "", draft, "", ""),
        ("管理层", "M02", "上线后月结 RACI、争议仲裁机制", "P0", "", "制度一页纸", "", "", draft, "", ""),
        ("管理层", "M03", "资源投入与关键里程碑", "P0", "", "甘特或里程碑表定稿", "", "", draft, "", ""),
    ]

    # --- Sheet: 部门汇总（汇报用）：先部门后明细，减轻「台账太乱」的阅读负担 ---
    ws_dept = wb.create_sheet("部门汇总（汇报用）", 2)
    hd = [
        "汇报顺序",
        "部门",
        "需求条数",
        "P0",
        "P1",
        "P2",
        "其他优先级",
        "覆盖编号",
        "汇报口径摘要（编号+要点，可按部门宣读）",
        "一句话汇报（口播稿）",
    ]
    add_headers(ws_dept, hd)
    dept_summary = summarize_by_department(rows)
    for r, row in enumerate(dept_summary, 2):
        for c, val in enumerate(row, 1):
            ws_dept.cell(row=r, column=c, value=val).alignment = WRAP
    note_row = 3 + len(dept_summary)
    ws_dept.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=len(hd))
    ws_dept.cell(
        row=note_row,
        column=1,
        value=(
            "使用建议（可删）：\n"
            "1）对领导：按「汇报顺序」从上到下，每部门先念「一句话汇报」，领导追问再展开「汇报口径摘要」或切到「分部门需求清单」。\n"
            "2）对部门：把本部门行单独复制到会议纪要或邮件正文即可。\n"
            "3）「需求收集汇总」「财务需求汇总」已根据规则写入「归纳部门」；详见「台账归部门规则」「台账按归纳部门统计」。"
        ),
    ).alignment = WRAP
    autosize(ws_dept, max_width=52, max_scan_rows=min(80, note_row + 3))

    # --- Sheet: 分部门需求清单 ---
    ws1 = wb.create_sheet("分部门需求清单", 3)
    h1 = [
        "部门",
        "编号",
        "需求要点",
        "自拟优先级",
        "确认后优先级",
        "产出/验收标准",
        "自拟计划完成",
        "确认后计划完成",
        "计划阶段",
        "对接人",
        "确认纪要",
    ]
    add_headers(ws1, h1)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws1.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws1)

    # --- Sheet: 需求收集汇总（源表第 1 张） ---
    title_req = "需求收集汇总" + (f"（{name0}）" if name0 else "")
    if rows0:
        rows0_out = augment_sheet_requirement_collect(list(rows0))
    else:
        rows0_out = [("（无数据）",)]

    title_fin = "财务需求汇总" + (f"（{name1}）" if name1 else "")
    if rows1:
        rows1_out = augment_sheet_finance_require(list(rows1))
    else:
        rows1_out = [("（无数据）",)]

    write_matrix_sheet(wb, title_req, rows0_out)
    write_matrix_sheet(wb, title_fin, rows1_out)

    fin_idx = wb.sheetnames.index(title_fin)
    write_matrix_sheet(
        wb, "台账归部门规则", build_ledger_rules_matrix(), insert_index=fin_idx + 1
    )
    rules_idx = wb.sheetnames.index("台账归部门规则")
    tally_m = build_ledger_dept_tally(
        rows0_out if rows0 else [],
        rows1_out if rows1 else [],
    )
    write_matrix_sheet(
        wb, "台账按归纳部门统计", tally_m, insert_index=rules_idx + 1
    )

    # --- Sheet: 里程碑计划 ---
    ws2 = wb.create_sheet("里程碑计划")
    h2 = [
        "阶段",
        "建议周期",
        "交付物",
        "参与部门",
        "自拟计划起止",
        "确认后计划起止",
        "负责人",
        "计划阶段",
        "备注",
    ]
    add_headers(ws2, h2)
    mrows = [
        ("D1 口径冻结会", "第1周", "基数与扣减口径、映射与账期规则（签字版）", "财务、人力、业务、产品", "", "", "", draft, ""),
        ("D2 PRD/原型封版", "第2-3周", "各页 PRD + 原型走查纪要", "产品、业务、财务", "", "", "", draft, ""),
        ("D3 开发迭代1", "第4-6周", "人员映射、规则包核心、试算主链路", "技术、测试", "", "", "", draft, ""),
        ("D4 联调试算", "第7周", "试算场景+公式+测试数据全绿", "财务、测试、业务", "", "", "", draft, ""),
        ("D5 UAT/月结演练", "第8周", "模拟月结：底表到归档", "全部门", "", "", "", draft, ""),
        ("D6 上线与稳态", "第9周起", "监控、问题清单 SLA", "技术+业务值班", "", "", "", draft, ""),
    ]
    for r, row in enumerate(mrows, 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws2)

    # --- Sheet: 风险与依赖 ---
    ws3 = wb.create_sheet("风险与依赖")
    h3 = ["类型", "描述", "影响", "缓解措施", "责任人", "状态"]
    add_headers(ws3, h3)
    rrows = [
        ("风险", "口径未冻结即开发", "返工、延期", "D1 前书面冻结基数与映射规则", "", ""),
        ("风险", "外围系统（工单/询价）边界不清", "集成扯皮、范围蔓延", "本期是否对接与字段责任书面确认", "", ""),
        ("风险", "历史数据清洗不足", "试算不准、争议", "单独数据治理子任务与验收", "", ""),
        ("风险", "多角色互斥与执行顺序复杂", "客服/业务投诉", "P0 场景用例+财务共签", "", ""),
        ("依赖", "财务提供凭证科目映射与样例账期", "F04/F02 等", "指定对接人与时间表", "", ""),
        ("依赖", "人力提供人事字段与异动生效规则", "H01/H02", "字段字典与流程确认", "", ""),
        ("依赖", "业务确认角色清单与回款政策例外", "S02/S05", "例外清单书面化", "", ""),
    ]
    for r, row in enumerate(rrows, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws3)

    # --- Sheet: 汇报节奏 ---
    ws4 = wb.create_sheet("汇报节奏")
    h4 = ["会议", "建议时长", "目的", "主要参与人", "频率/时机"]
    add_headers(ws4, h4)
    prows = [
        ("启动会", "1h", "目标、范围、RACI、里程碑、风险Top5", "领导+各部门负责人", "项目启动时"),
        ("计划确认会", "30–45min", "对齐自拟与部门反馈，形成可发布的确认优先级与日期", "领导+各部门负责人", "自拟完成后、定稿前"),
        ("双周例会", "30min", "需求表状态、阻塞项、变更评审", "项目组+部门对接人", "每双周"),
        ("口径专题会", "按需", "争议口径决策，输出决策记录", "财务+业务+产品", "有开放问题时"),
        ("上线前评审会", "1h", "试算样例签字、映射规则、角色命中确认", "财务+人力+业务+技术", "上线前一周"),
    ]
    for r, row in enumerate(prows, 2):
        for c, val in enumerate(row, 1):
            ws4.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws4)

    # --- Sheet: 汇报材料清单 ---
    ws5 = wb.create_sheet("汇报材料清单")
    h5 = ["材料名称", "用途", "负责部门", "是否必备"]
    add_headers(ws5, h5)
    mat = [
        ("部门汇总（汇报用）", "按部门条数、P0/P1/P2、口播稿；会上主看本表", "项目经理", "是"),
        ("分部门需求清单（自拟列→确认后列）", "先自拟再定稿；对齐优先级与验收", "产品牵头", "是"),
        ("里程碑计划（自拟/确认后起止）", "先自拟再定稿；进度与投入可见", "项目经理", "是"),
        ("计划定稿流程（本表工作表）", "对内说明两阶段用法", "项目经理", "建议"),
        ("风险与依赖表", "领导决策与协调", "项目经理", "是"),
        ("未决问题与不做清单", "范围控制", "产品", "是"),
        ("试算公式与测试数据册", "开发测试财务共用", "产品+财务", "是"),
        ("需求收集表.xlsx（全员台账）", "与「需求收集汇总」页交叉核对；已带「归纳部门」", "产品/PMO", "建议"),
        ("财务需求台账", "与「财务需求汇总」页交叉核对；已全部归财务部", "财务", "建议"),
    ]
    for r, row in enumerate(mat, 2):
        for c, val in enumerate(row, 1):
            ws5.cell(row=r, column=c, value=val).alignment = WRAP
    autosize(ws5)

    try:
        wb.save(OUT)
        out_path = OUT
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "_生成副本.xlsx")
        wb.save(alt)
        out_path = alt
        print("提示：原文件可能被 Excel 占用，已写入副本:", alt)
    print("已生成:", out_path)
    if name0:
        print(
            "  已并入源表:",
            SRC_REQUIREMENTS.name,
            "/",
            name0,
            "→「需求收集汇总」（已按「功能名称」写入「归纳部门」）",
        )
    else:
        print("  未读取到源表第 1 张（请确认路径:", SRC_REQUIREMENTS, "）")
    if name1:
        print(
            "  已并入源表:",
            SRC_REQUIREMENTS.name,
            "/",
            name1,
            "→「财务需求汇总」（全部行「归纳部门」= 财务部）",
        )


if __name__ == "__main__":
    main()
![1778831375372](image/生成_各部门需求清单与汇报计划_excel/1778831375372.png)![1778831379451](image/生成_各部门需求清单与汇报计划_excel/1778831379451.png)![1778831397992](image/生成_各部门需求清单与汇报计划_excel/1778831397992.png)![1778831405052](image/生成_各部门需求清单与汇报计划_excel/1778831405052.png)