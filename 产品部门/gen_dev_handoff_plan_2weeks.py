# -*- coding: utf-8 -*-
"""生成《产品→开发 · 原型与设计计划 · 两周版》HTML + Excel。

周期：2026-08-10（周一）～ 08-21（周五）
口径：从四周滚动计划压成两周可执行版；**前两周只梳理主链路**；不影响流程的放第三周。
节奏：周一～周三出完流程图；周四起做主链路原型/PRD；第三周做报表/预警/邮箱等旁支。
"""
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DIR = Path(__file__).resolve().parent
OUT_XLSX = DIR / "产品部门-对开发-原型与设计计划表_两周版.xlsx"
OUT_HTML = DIR / "产品部门-对开发-原型与设计计划表_两周版.html"

thin = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 原型链接（相对「产品部门/」目录，HTML/Excel 可点）
def _p(*pairs):
    return [{"label": a, "href": b} for a, b in pairs]


# —— 日排程 ——
# items: group, action, features, protos, ids, pri
DAYS = [
    {
        "week": "第1周 · 计划与流程图",
        "date": "08/10 周一",
        "focus": "做计划 + 画流程图（私卡询价报价优先）",
        "items": [
            {
                "group": "产品+报价",
                "action": "两周主链路对齐：Must=流程链；非流程项进第三周",
                "features": "①主链路顺序：询价报价→合约价确认→提单→报关→查货/查验→进港→预估→尾单→工单 ②旁支进第三周（BJ002 开发中、GW021 已完成不排）③周五/下周五交付节点",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("四周滚动版", "产品部门-对开发-原型与设计计划表.html"),
                    ("需求清单评审版", "产品部门-各组需求清单-评审版.html"),
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                ),
                "ids": "计划",
                "pri": "P0",
            },
            {
                "group": "报价",
                "action": "画「私卡询价→报价→确认→绑运单」主流程图（As-Is / To-Be）",
                "features": "流程图节点：①询价进入报价员列表 ②AI报价生成 ③人工确认 ④尾端派送费 ⑤散货多绑运单 ⑥有效期 ⑦旧AI下线边界；标注系统缺口与页面落点",
                "protos": _p(
                    ("0731 AI报价方案（流程底稿）", "各组sop/私卡报价-0731-AI报价有效期运单绑定-大致方案.html"),
                    ("报价员列表·MVP", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("AI报价规则配置", "各组sop/私卡报价-AI报价规则配置-MVP.html"),
                ),
                "ids": "BJ001",
                "pri": "P0",
            },
            {
                "group": "报价+船务",
                "action": "画「合约价确认」与运价列表写回流程图（不含点价订舱）",
                "features": "①合约价展示 ②确认开关 ③写回字段与状态；决策点标红待拍（点价订舱已从主链路拿掉）",
                "protos": _p(
                    ("报价员列表·AI确认", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("运价管理·我司/代理订舱", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                ),
                "ids": "BJ001 / CW008",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "主链路 vs 第三周旁支：在全流程评审版上标色",
                "features": "①主链路节点标绿（两周必梳）②旁支（时效/POD/报表/邮箱/HOLD/比对）标灰进第三周 ③发群对齐口径",
                "protos": _p(
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                ),
                "ids": "计划",
                "pri": "P0",
            },
        ],
    },
    {
        "week": "第1周 · 计划与流程图",
        "date": "08/11 周二",
        "focus": "做计划 + 画流程图（各组泳道/切片）",
        "items": [
            {
                "group": "风控+关务",
                "action": "画查货合规 / 查验备注流程图（复审与落点对齐）",
                "features": "①查货七类标记→成本/通知 ②免查保函同页 ③查验备注谁登记 ④国内/国外分开展示；标出待复审点",
                "protos": _p(
                    ("查货操作·AI识图", "各组sop/风控组-查货操作-AI识图品名识别-原型.html"),
                    ("查验报表", "各组sop/关务组-查验报表-提单运单维度-MVP.html"),
                    ("8/6 查货评审纪要", "风控组调研纪要_2026-08-06_查货原型评审.html"),
                ),
                "ids": "FK002–006 / 查验",
                "pri": "P0",
            },
            {
                "group": "船务",
                "action": "画提单管理（按周同船）+ 运价确认开关流程图",
                "features": "①按周→同船→装完色标 ②费用含水单 ③取消IT ④运价确认写回；标 UI 回改点（不含点价订舱）",
                "protos": _p(
                    ("提单管理", "各组sop/船务组-提单管理-按周分组与分配-MVP.html"),
                    ("运价管理", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                    ("0730 费用节点方案", "各组sop/船务组-0730-费用节点提单配仓-大致方案.html"),
                ),
                "ids": "提单 / CW008",
                "pri": "P0",
            },
            {
                "group": "关务+海外",
                "action": "画报关四切片序 + 海外预估/尾单/工单边界流程图",
                "features": "①分单号→双通道→改配→买单 ②主单预估↔尾程预估 ③尾单转工单边界；标注本周可交 vs 下周补齐",
                "protos": _p(
                    ("资料核准双通道", "各组sop/关务组-资料核准双通道-流程与页面设计.html"),
                    ("提单主单预估", "各组sop/海外对接组-提单主单预估费用表-MVP.html"),
                    ("尾单跟踪", "各组sop/海外对接组-港后-尾单单票跟踪-原型.html"),
                    ("工单入口", "../工单/工单提交-入口-MVP.html"),
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                ),
                "ids": "关务切片 / 海外",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "汇总流程图清单 + 确认周四起进入原型/PRD细节（发群/对齐开发）",
                "features": "①流程图文件清单（私卡优先）②Thu–Fri 及第2周原型/PRD排期 ③开发可先看的流程图链接包",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("复杂泳道（参考）", "各组sop/产品部门-复杂泳道全图-drawio.mmd"),
                ),
                "ids": "计划",
                "pri": "—",
            },
        ],
    },
    {
        "week": "第1周 · 计划与流程图",
        "date": "08/12 周三",
        "focus": "流程图出完日 · 收口评审",
        "items": [
            {
                "group": "产品+各组",
                "action": "【节点】各组流程图出完：走查缺口、统一图例与落点标注",
                "features": "①私卡询价报价主流程定稿图 ②查货/查验流程定稿图 ③提单+运价流程定稿图 ④报关切片+海外预估/尾单/工单边界图；缺页标红",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                    ("0731 AI报价方案", "各组sop/私卡报价-0731-AI报价有效期运单绑定-大致方案.html"),
                ),
                "ids": "流程图收口",
                "pri": "P0",
            },
            {
                "group": "报价",
                "action": "私卡流程图终校：确认按钮/有效期/多绑节点写清规则与默认",
                "features": "①确认开关默认建议 ②未确认能否订舱 ③有效期与绑运单规则 ④旧AI下线边界；供周四原型/PRD直接引用",
                "protos": _p(
                    ("报价员列表·MVP", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("AI报价规则配置", "各组sop/私卡报价-AI报价规则配置-MVP.html"),
                    ("运价管理", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                ),
                "ids": "BJ001",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "发布《流程图包 v1》给开发预读 + 周四起原型/PRD细节任务拆解",
                "features": "①流程图链接清单 ②对应已有MVP/方案索引 ③待拍默认策略 ④周四～下周五交付切片（私卡优先）",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("需求清单", "产品部门-各组需求清单-评审版.html"),
                ),
                "ids": "计划",
                "pri": "P0",
            },
            {
                "group": "各组",
                "action": "按流程图核对原型缺口：哪些页要回改/补PRD（列清单）",
                "features": "①页面级缺口表（有原型/缺PRD/缺交互）②周四优先改哪些页 ③可延后项显式写入第2周",
                "protos": _p(
                    ("四周滚动版（对照）", "产品部门-对开发-原型与设计计划表.html"),
                ),
                "ids": "缺口表",
                "pri": "P1",
            },
        ],
    },
    {
        "week": "第1周 · 原型与PRD细节",
        "date": "08/13 周四",
        "focus": "【优先】私卡询价报价 · 原型回改 + PRD细节",
        "items": [
            {
                "group": "报价",
                "action": "【优先】私卡询价报价：按流程图回改原型 + 补 PRD 细节",
                "features": "①询价列表→报价员作业台（待办/超时/状态）②新AI：本州/外州规则（713）③报价确认 ④尾端派送费 ⑤散货多绑运单 ⑥有效期与运单绑定 ⑦旧AI下线；字段/状态/权限写入PRD",
                "protos": _p(
                    ("报价员列表·AI确认/多绑", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("AI报价规则配置", "各组sop/私卡报价-AI报价规则配置-MVP.html"),
                    ("0731 AI报价/有效期/绑运单方案", "各组sop/私卡报价-0731-AI报价有效期运单绑定-大致方案.html"),
                    ("本州地址报价测算", "各组sop/海外仓至本州地址报价系统.html"),
                    ("外州地址报价测算", "各组sop/海外仓至外州地址报价系统.html"),
                ),
                "ids": "BJ001",
                "pri": "P0",
            },
            {
                "group": "报价+船务",
                "action": "合约价旁「确认」定稿写入 PRD（询价报价×运价联动）",
                "features": "①确认按钮可配置开/关 ②写回字段与状态；与流程图周三结论一致（不含点价订舱）",
                "protos": _p(
                    ("报价员列表·AI确认", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("运价管理·我司/代理订舱", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                ),
                "ids": "BJ001 / CW008",
                "pri": "P0",
            },
            {
                "group": "报价",
                "action": "私卡主流程验收样例 + 状态机一页（交接最低包）",
                "features": "①正常/超时/多绑/有效期过期 4 类样例 ②状态机：询价→AI报价→确认→绑运单 ③权限：报价员/业务可见范围",
                "protos": _p(
                    ("报价员列表·MVP", "../二期/ESS询价报价-报价员列表-MVP.html"),
                    ("0731 方案", "各组sop/私卡报价-0731-AI报价有效期运单绑定-大致方案.html"),
                ),
                "ids": "BJ001",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "主链路缺口复核：周四后只补流程断点，旁支不插队",
                "features": "①对照全流程绿标节点 ②断点写入周五周交付 ③BJ002/BJ004 等旁支确认进第三周",
                "protos": _p(
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                ),
                "ids": "计划",
                "pri": "P1",
            },
        ],
    },
    {
        "week": "第1周 · 原型与PRD细节",
        "date": "08/14 周五",
        "focus": "风控查货 + 周交付会（查验报表 → 第三周）",
        "items": [
            {
                "group": "风控",
                "action": "查货原型按 8/6 回改 + PRD 对接清单（约郭佳玉复审）",
                "features": "①七类合规标记 ②免查保函同页 ③备注标准化可选 ④直客只加成本 ⑤AI识图辅助 ⑥附加费/超品名费 ⑦扣件罚款500+通知 ⑧扣货不赔；输出字段/状态/权限 ⑨操作查货率报表→第三周（FK001旁支）",
                "protos": _p(
                    ("查货操作·AI识图/品名", "各组sop/风控组-查货操作-AI识图品名识别-原型.html"),
                    ("0729 查货方案", "各组sop/风控组-0729-查货管理-大致方案.html"),
                    ("8/6 查货评审纪要", "风控组调研纪要_2026-08-06_查货原型评审.html"),
                ),
                "ids": "FK002–006",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "【周交付】主链路进度：流程图包 v1 + 私卡 + 查货",
                "features": "①主链路流程图出完 ②私卡/查货 Done·Slip ③船务→报关→海外滚入第2周 ④旁支清单确认进第三周（含查验报表/查货率报表）",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("需求清单", "产品部门-各组需求清单-评审版.html"),
                ),
                "ids": "—",
                "pri": "—",
            },
        ],
    },
    {
        "week": "第2周 · 原型与PRD细节",
        "date": "08/17 周一",
        "focus": "船务提单/运价 + 关务切片启动",
        "items": [
            {
                "group": "船务",
                "action": "提单管理：按流程图做 UI 回改 + PRD 要点",
                "features": "①按周·同船 ②装完两色 ③AMS/ISF一体 ④费用含水单 ⑤取消IT（船务侧）⑥查验/改单标记 ⑦截单/开船预警 ⑧一水绑柜；分配发群/仓位播报/绑进口商",
                "protos": _p(
                    ("提单管理·按周/费用水单/IT", "各组sop/船务组-提单管理-按周分组与分配-MVP.html"),
                    ("提单创建·AI识别", "各组sop/船务组-提单创建批量-AI识别-MVP.html"),
                    ("进口商管理", "../进口商管理-MVP.html"),
                ),
                "ids": "提单管理 / CW002",
                "pri": "P0",
            },
            {
                "group": "船务",
                "action": "运价管理：确认开关化 + PRD（随周四私卡结论；不含点价订舱）",
                "features": "①我司价 vs 代理价（含Local）②合约运价上传/沉淀 ③确认按钮开关 ④批量创建+AI识别",
                "protos": _p(
                    ("运价管理·我司/代理订舱", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                    ("提单创建·AI识别", "各组sop/船务组-提单创建批量-AI识别-MVP.html"),
                ),
                "ids": "CW008/010",
                "pri": "P0",
            },
            {
                "group": "关务",
                "action": "报关核心切片清单 + 进口商/清关行绑定（原型/PRD启动）",
                "features": "切片①分单号 ②资料核准双通道 ③改配轨迹 ④买单留底；前序：进口商/清关行/柜子绑定",
                "protos": _p(
                    ("分单号规则配置", "../报关行对接/分单号规则配置-MVP.html"),
                    ("资料核准双通道", "各组sop/关务组-资料核准双通道-流程与页面设计.html"),
                    ("改配三类场景", "各组sop/关务组-改配三类场景-轨迹与同步.html"),
                    ("进口商管理", "../进口商管理-MVP.html"),
                    ("清关行·港口配置", "../进口商-清关行港口配置-MVP.html"),
                ),
                "ids": "GW001–007等",
                "pri": "P0",
            },
            {
                "group": "船务+海外",
                "action": "进港口径一页纸启动 + 海外工单流程梳理启动",
                "features": "①进港：抓取源/落点/异常（写入0730附录）②工单：角色/触发/状态/与尾单边界（本周只梳到可对稿）",
                "protos": _p(
                    ("0730 费用节点·进港/落箱", "各组sop/船务组-0730-费用节点提单配仓-大致方案.html"),
                    ("工单提交入口", "../工单/工单提交-入口-MVP.html"),
                    ("尾单·转工单", "各组sop/海外对接组-港后-尾单单票跟踪-原型.html"),
                ),
                "ids": "CW001 / 工单",
                "pri": "P0",
            },
        ],
    },
    {
        "week": "第2周 · 原型与PRD细节",
        "date": "08/18 周二",
        "focus": "海外预估 + 尾单（主链路港后）",
        "items": [
            {
                "group": "海外",
                "action": "提单主单预估 PRD 对齐 8/7",
                "features": "①头程四维价：代理×船司×始发港×柜型 ②固定费标准+差异 ③拖车（堆场/省市区）④报关合票差 ⑤EDI≈¥10/舱单 ⑥装卸柜进预估 ⑦取消IT操作费 ⑧查验费：开关（默认关，预留字段）",
                "protos": _p(
                    ("提单主单预估·MVP", "各组sop/海外对接组-提单主单预估费用表-MVP.html"),
                    ("提单主单预估·PRD", "各组sop/海外对接组-提单主单预估费用表-PRD.html"),
                ),
                "ids": "预估主线",
                "pri": "P0",
            },
            {
                "group": "海外",
                "action": "尾程运单预估 + 港后尾单/卡派；工单跳转写入 PRD",
                "features": "①尾程/海外费按运单预估（含海外仓费）②尾单四桶跟踪（私卡派/卡车派→约仓分桶）③卡派跟进(新)：港后客服、双超时维 ④异常/改派送→「转工单」",
                "protos": _p(
                    ("尾程预估·MVP", "各组sop/海外对接组-尾程预估费用表-MVP.html"),
                    ("尾程预估·PRD", "各组sop/海外对接组-尾程预估-运单维-PRD.html"),
                    ("尾单单票跟踪", "各组sop/海外对接组-港后-尾单单票跟踪-原型.html"),
                    ("卡派跟进(新)", "各组sop/海外对接组-卡派跟进新-MVP.html"),
                    ("工单提交入口", "../工单/工单提交-入口-MVP.html"),
                ),
                "ids": "HW005",
                "pri": "P0",
            },
            {
                "group": "海外",
                "action": "主单/尾程预估字段对齐（保证主链路费用一致）",
                "features": "对齐：费用类型编码、币种、柜号/运单号、预估来源版本号；柜维差异比对（HW001）进第三周",
                "protos": _p(
                    ("提单主单预估·PRD", "各组sop/海外对接组-提单主单预估费用表-PRD.html"),
                    ("尾程预估·PRD", "各组sop/海外对接组-尾程预估-运单维-PRD.html"),
                ),
                "ids": "预估主线",
                "pri": "P0",
            },
            {
                "group": "关务",
                "action": "报关切片①②深化：分单号 + 资料核准双通道 PRD",
                "features": "①分单号规则/自动生成 ②双通道核准状态机 ③与进口商绑定衔接；改配/买单可周三继续",
                "protos": _p(
                    ("分单号规则配置", "../报关行对接/分单号规则配置-MVP.html"),
                    ("资料核准双通道", "各组sop/关务组-资料核准双通道-流程与页面设计.html"),
                    ("做资料分票核准", "../报关行对接/报关管理-做资料分票核准-MVP.html"),
                ),
                "ids": "GW001–007",
                "pri": "P0",
            },
        ],
    },
    {
        "week": "第2周 · 原型与PRD细节",
        "date": "08/19 周三",
        "focus": "主链路端到端走查 + 报关切片收口",
        "items": [
            {
                "group": "产品+各组",
                "action": "【硬门禁】主链路端到端走查：断点清单清零或显式 Slip",
                "features": "按序过一遍：询价报价→合约价确认→提单→报关四切片→查货/查验→进港→主单/尾程预估→尾单→转工单；每环：有图/有原型或PRD/有对接要点",
                "protos": _p(
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                ),
                "ids": "主链路",
                "pri": "P0",
            },
            {
                "group": "关务",
                "action": "报关切片③④：改配轨迹 + 买单留底 PRD/原型要点",
                "features": "①改配三类场景轨迹同步 ②买单清单留底+报关编号/报关单存档 ③开发顺序①→②→③→④冻结",
                "protos": _p(
                    ("改配三类场景", "各组sop/关务组-改配三类场景-轨迹与同步.html"),
                    ("报关管理列表", "../报关行对接/报关管理-列表-MVP.html"),
                ),
                "ids": "GW切片",
                "pri": "P0",
            },
            {
                "group": "船务",
                "action": "提单/运价走查遗留回改（主链路断点补齐）",
                "features": "①按周同船/装完色标 ②费用含水单 ③取消IT ④确认开关与私卡结论一致 ⑤进港字段预留",
                "protos": _p(
                    ("提单管理", "各组sop/船务组-提单管理-按周分组与分配-MVP.html"),
                    ("运价管理", "各组sop/船务组-运价管理-我司与代理订舱-MVP.html"),
                ),
                "ids": "提单 / CW008",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "第三周旁支清单冻结（本周不插队）",
                "features": "旁支进第三周：POD、查货率报表、LTL、信号旗、HW001、邮箱、HOLD、时效例外等；BJ002 开发中 / GW021 已完成（不排）",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                ),
                "ids": "第三周",
                "pri": "—",
            },
        ],
    },
    {
        "week": "第2周 · 原型与PRD细节",
        "date": "08/20 周四",
        "focus": "工单对稿 + 进港口径定稿（主链路收口）",
        "items": [
            {
                "group": "海外+工单",
                "action": "海外工单流程评审对稿（尾单转工单边界）",
                "features": "通过则立项：工单类型/表单/工作台/轨迹回写/权限；未过则输出缺口清单（无开发排期）",
                "protos": _p(
                    ("工单提交入口", "../工单/工单提交-入口-MVP.html"),
                    ("工单工作台", "../工单/工单工作台-MVP.html"),
                    ("尾单·转工单", "各组sop/海外对接组-港后-尾单单票跟踪-原型.html"),
                ),
                "ids": "工单",
                "pri": "P0",
            },
            {
                "group": "船务",
                "action": "进港口径一页纸定稿（触发落箱等主链路节点）",
                "features": "定稿后可开发：①进港时间字段CRUD/抓取任务 ②落箱费节点触发 ③未进港/异常列表与提示；写入0730方案附录",
                "protos": _p(
                    ("0730 费用节点·进港/落箱", "各组sop/船务组-0730-费用节点提单配仓-大致方案.html"),
                ),
                "ids": "CW001",
                "pri": "P0",
            },
            {
                "group": "风控",
                "action": "复审遗留清零；查货包标「可开发」",
                "features": "FK002–006 功能点清单冻结版本号；开发按冻结清单排期",
                "protos": _p(
                    ("查货操作·AI识图/品名", "各组sop/风控组-查货操作-AI识图品名识别-原型.html"),
                    ("0729 查货方案", "各组sop/风控组-0729-查货管理-大致方案.html"),
                ),
                "ids": "FK002–006",
                "pri": "P0",
            },
            {
                "group": "产品",
                "action": "主链路交接包预装订：链接+PRD+样例索引",
                "features": "按主链路顺序整理可交包目录；旁支仅列第三周标题不展开细节",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("需求清单", "产品部门-各组需求清单-评审版.html"),
                ),
                "ids": "—",
                "pri": "P0",
            },
        ],
    },
    {
        "week": "第2周 · 收口交付",
        "date": "08/21 周五",
        "focus": "两周收口交付会",
        "items": [
            {
                "group": "产品",
                "action": "【两周交付】主链路可交开发包 v2",
                "features": "按主链路合并：私卡询价报价→合约价确认→提单/运价→报关四切片→查货/查验→进港→预估/尾单→工单边界；旁支一律进第三周（不含点价订舱）",
                "protos": _p(
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                    ("四周滚动版", "产品部门-对开发-原型与设计计划表.html"),
                    ("需求清单", "产品部门-各组需求清单-评审版.html"),
                    ("全流程评审版", "各组sop/产品部门-标准化全流程-评审版.html"),
                    ("8/6–8/7 收口纪要", "调研纪要梳理_2026-08-06至08-07_卢慧恒_原型评审.html"),
                ),
                "ids": "—",
                "pri": "—",
            },
            {
                "group": "产品",
                "action": "【待拍清单】带默认建议（不阻塞主链路）",
                "features": "①查验进预估默认关 ②信号旗默认人工（第三周再评技术）③截单→关务复用待办 ④进港遗留项清单",
                "protos": _p(
                    ("8/6–8/7 收口纪要", "调研纪要梳理_2026-08-06至08-07_卢慧恒_原型评审.html"),
                    ("提单主单预估·PRD", "各组sop/海外对接组-提单主单预估费用表-PRD.html"),
                ),
                "ids": "—",
                "pri": "—",
            },
            {
                "group": "产品",
                "action": "【第三周启动包】非主流程旁支清单定稿",
                "features": "第三周：POD验收、查货率报表、LTL、信号旗、HW001、邮箱、HOLD/LFD、时效例外、滞箱全页等；不含 BJ002（开发中）、GW021（已完成）",
                "protos": _p(
                    ("四周滚动版（第三周）", "产品部门-对开发-原型与设计计划表.html"),
                    ("0804 方案（邮箱/HOLD）", "各组sop/海外对接组-0804-港前港后整柜价格-大致方案.html"),
                ),
                "ids": "第三周",
                "pri": "—",
            },
            {
                "group": "各组",
                "action": "业务确认下周开发对接 Top5（仅主链路）",
                "features": "从已冻结主链路功能点中选出开发优先5项（会议输出）",
                "protos": _p(
                    ("需求清单·按P0筛", "产品部门-各组需求清单-评审版.html"),
                    ("本两周版计划", "产品部门-对开发-原型与设计计划表_两周版.html"),
                ),
                "ids": "—",
                "pri": "—",
            },
        ],
    },
]

# —— 各组两周目标（卡片）—— 主链路优先
GROUPS = [
    {
        "name": "报价组",
        "color": "#389e0d",
        "bg": "#f6ffed",
        "w1": "✅ 08/18 原型已完成（BJ001）",
        "w2": "维护已交包；响应船务联调",
        "must": ["BJ001 已完成"],
        "stretch": ["BJ002 🔧开发中", "BJ004/003 → 第三周"],
        "handoff": "报价员列表 + 我的询价 + AI规则 + 需求PRD",
    },
    {
        "name": "风控组",
        "color": "#389e0d",
        "bg": "#f6ffed",
        "w1": "✅ 08/18 原型已完成（FK002–006）",
        "w2": "不占关务/船务档",
        "must": ["查货合规 PRD+原型 已完成"],
        "stretch": ["FK001 查货率报表 → 第三周"],
        "handoff": "查货操作 + 品名库 + 查货合规 PRD",
    },
    {
        "name": "关务组",
        "color": "#1890ff",
        "bg": "#e6f7ff",
        "w1": "报关流程图；四切片 PRD",
        "w2": "报关四切片+进口商（主链路）；查验报表/AN → 第三周",
        "must": ["分单号→双通道→改配→买单"],
        "stretch": ["查验报表备注 → 第三周", "GW021 清关放行轨迹 → ✅已完成", "GW020 全原型 → 第三周"],
        "handoff": "报关系列MVP + 双通道/改配 + 进口商配置",
    },
    {
        "name": "船务组",
        "color": "#13c2c2",
        "bg": "#e6fffb",
        "w1": "提单/运价流程图",
        "w2": "提单+运价+进港口径（主链路）；滞箱全页 → 第三周",
        "must": ["提单管理可交接", "运价确认开关可交接", "进港口径书面化"],
        "stretch": ["滞箱滞港全页 → 第三周"],
        "handoff": "提单管理MVP + 运价MVP + 提单创建AI + 0730方案",
    },
    {
        "name": "海外对接组",
        "color": "#cf1322",
        "bg": "#fff1f0",
        "w1": "预估/尾单/工单边界流程图",
        "w2": "预估+尾单+工单边界（主链路）；HW001/邮箱/HOLD/时效 → 第三周",
        "must": ["预估主线可开发", "HW005 尾单可开发", "工单流程有评审结论"],
        "stretch": ["HW001/003/004、时效例外 → 第三周"],
        "handoff": "预估MVP/PRD + 尾单原型 + 卡派新",
    },
]

# 旁支 / 进度：事项, 需求ID, 优先级, 状态, 说明
BACKLOG = [
    ("私卡时效报表 / 报价时效看板", "BJ002", "P0", "🔧 开发中", "2026-08-11 标记：开发已开工；产品侧不重复排第三周交付"),
    ("清关放行轨迹（文案/节点）", "GW021", "P1", "✅ 已完成", "2026-08-11 标记：开发已完成；产品侧关闭，不进第三周"),
    ("点价订舱", "CW008", "P0", "暂缓", "已从主链路拿掉；两周不做"),
    ("出库/签收/POD预警与验收", "BJ004", "P0", "第三周", "流程后监控 · 不影响主流程"),
    ("LTL轨迹自动抓取", "BJ003", "P0", "第三周", ""),
    ("查验报表备注落点（提单/运单维）", "查验报表", "P1", "第三周", "GW021已开发完成；产品PRD后置"),
    ("操作查货率报表字段口径", "FK001", "P1", "第三周", "风控报表旁支；原型已有"),
    ("信号旗抓IT/保税号技术结论", "FK关联", "P1", "第三周", "默认先人工"),
    ("柜维预估差异比对首版", "HW001", "P0", "第三周", "对账旁支"),
    ("目的港邮箱工作台", "HW003", "P0", "第三周", ""),
    ("HOLD/LFD/提柜预警", "HW004/012", "P0", "第三周", ""),
    ("时效表例外规则", "时效MVP", "P1", "第三周", "统计旁支"),
    ("AN/费用单上传要点", "GW025", "P1", "第三周", ""),
    ("滞箱/滞港/免用箱全页", "CW007/005", "P0", "第三周", ""),
    ("查验前后端同步全原型", "GW020", "P0", "第三周+", ""),
    ("签入无成本卡控设计", "BJ005", "P1", "第三周+", "待设计"),
    ("沃尔玛仓费用流转", "BJ006", "P2", "第三周+", "待设计"),
    ("询价数据导出", "BJ007", "P2", "第三周+", "待设计"),
    ("整柜还柜轨迹对客服", "HW015", "P1", "第三周", "可插缝"),
    ("DO 归档 / 快递查询慢", "—", "P3", "暂缓", "体验项"),
]

MUST_DELIVERABLES = [
    "【已完成】报价 BJ001 + 风控 FK002–006 原型/PRD（截止 08/18）",
    "主链路通：询价报价(✅)→查货(✅)→提单/运价→报关四切片→进港→预估→尾单→工单",
    "提单管理 + 运价管理（确认开关）+ 进港口径一页纸",
    "关务：报关四切片序可交（查验报表 → 第三周）",
    "海外：主单预估 + 尾程预估 + 尾单/卡派 + 工单评审结论",
]

def esc(s):
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def write_xlsx():
    wb = Workbook()

    # 总览
    ov = wb.active
    ov.title = "总览"
    ov["A1"] = "产品 → 开发 · 原型与设计计划 · 两周版"
    ov["A1"].font = Font(bold=True, size=14, color="001529")
    rows = [
        [],
        ["周期", "2026-08-10（周一）～ 08-21（周五）"],
        ["定位", "前两周只梳主链路（询价→合约价确认→提单→报关→查货/查验→进港→预估→尾单→工单）；不含点价订舱；旁支进第三周"],
        ["来源", "各组需求清单102条 + 0727–0807纪要（含8/6–8/7收口）+ 四周版计划"],
        ["关联", "完整滚动版见：产品部门-对开发-原型与设计计划表.html"],
        [],
        ["开发进度标记（2026-08-11）"],
        ["私卡时效报表 BJ002", "🔧 开发中 · 产品侧不重复排期"],
        ["清关放行轨迹 GW021", "✅ 已完成 · 产品侧关闭"],
        [],
        ["两周 Must（必须交付）"],
        *[[f"{i+1}", x] for i, x in enumerate(MUST_DELIVERABLES)],
        [],
        ["交接最低包"],
        ["①", "可点击原型（HTML）"],
        ["②", "一页 PRD：范围/非范围/字段/状态/权限"],
        ["③", "验收样例"],
        ["④", "待拍板清单（带默认建议）"],
        [],
        ["并行待拍（带默认，不阻塞）"],
        ["查验费进预估", "默认：开关关闭，预留字段"],
        ["信号旗抓 IT/保税号", "默认：先不做自动抓，人工录入+标记"],
        ["截单预警→关务", "默认：复用现有待办通道，产品核实是否已触达"],
        ["合约价「确认」", "周三流程图定规则，周四写入PRD并开关化"],
    ]
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ov.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row and c == 1 and len(row) == 1:
                cell.font = Font(bold=True, color="003A8C")
    ov.column_dimensions["A"].width = 22
    ov.column_dimensions["B"].width = 72

    # 日排程
    ws = wb.create_sheet("日排程")
    headers = ["周次", "日期", "当日焦点", "组别", "产品动作", "功能点清单（给开发）", "原型链接", "需求ID", "优先级"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="001529")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = thin
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    r = 2
    for day in DAYS:
        for it in day["items"]:
            proto_txt = "\n".join(
                f"{p['label']}: {p['href']}" for p in it.get("protos") or []
            ) or "—"
            vals = [
                day["week"],
                day["date"],
                day["focus"],
                it["group"],
                it["action"],
                it["features"],
                proto_txt,
                it["ids"],
                it["pri"],
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(r, c, v)
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if "第1周" in day["week"] and c == 1:
                    cell.font = Font(bold=True, color="CF1322")
                elif "第2周" in day["week"] and c == 1:
                    cell.font = Font(bold=True, color="D46B08")
            r += 1
    for i, w in enumerate([14, 12, 14, 10, 26, 46, 42, 12, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{r-1}"

    # 按组
    gws = wb.create_sheet("按组两周目标")
    gh = ["组别", "第1周目标", "第2周目标", "Must", "Stretch", "交接包"]
    for c, h in enumerate(gh, 1):
        cell = gws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="001529")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = thin
    for i, g in enumerate(GROUPS, 2):
        vals = [
            g["name"],
            g["w1"],
            g["w2"],
            "；".join(g["must"]),
            "；".join(g["stretch"]),
            g["handoff"],
        ]
        for c, v in enumerate(vals, 1):
            cell = gws.cell(i, c, v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, w in enumerate([10, 36, 36, 36, 28, 40], 1):
        gws.column_dimensions[get_column_letter(i)].width = w

    # 延后
    bl = wb.create_sheet("第三周旁支")
    bh = ["事项", "需求ID", "优先级", "状态", "说明"]
    for c, h in enumerate(bh, 1):
        cell = bl.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="595959")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = thin
    for i, row in enumerate(BACKLOG, 2):
        for c, v in enumerate(row, 1):
            cell = bl.cell(i, c, v)
            cell.border = thin
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # 开发中 / 已完成 高亮
            st = row[3] if len(row) > 3 else ""
            if c == 4 and "开发中" in str(st):
                cell.fill = PatternFill("solid", fgColor="FFF7E6")
                cell.font = Font(color="D46B08", bold=True)
            elif c == 4 and "已完成" in str(st):
                cell.fill = PatternFill("solid", fgColor="F6FFED")
                cell.font = Font(color="389E0D", bold=True)
    for i, w in enumerate([32, 12, 8, 12, 48], 1):
        bl.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)
    print("wrote", OUT_XLSX)


def write_html():
    # 日排程表
    day_rows = []
    cur_week = None
    for day in DAYS:
        if day["week"] != cur_week:
            cur_week = day["week"]
            cls = "w1" if "第1周" in cur_week else "w2"
            day_rows.append(
                f'<tr class="wave-hd {cls}"><td colspan="6"><span class="wtag">{esc(cur_week)}</span></td></tr>'
            )
        first = True
        for it in day["items"]:
            date_cell = (
                f'<td class="date" rowspan="{len(day["items"])}"><b>{esc(day["date"])}</b>'
                f'<div class="sub">{esc(day["focus"])}</div></td>'
                if first
                else ""
            )
            first = False
            pri = it["pri"]
            pri_cls = "p0" if pri == "P0" else ("p1" if pri == "P1" else "px")
            feats = it["features"]
            if "①" in feats:
                chunks = [c.strip() for c in re.split(r"(?=①|②|③|④|⑤|⑥|⑦|⑧|⑨|⑩)", feats) if c.strip()]
                if len(chunks) > 1:
                    feat_html = "<ul class='feat'>" + "".join(f"<li>{esc(c)}</li>" for c in chunks) + "</ul>"
                else:
                    feat_html = esc(feats)
            else:
                feat_html = esc(feats)
            protos = it.get("protos") or []
            if protos:
                links = "".join(
                    f'<a class="proto" href="{esc(p["href"])}" target="_blank" rel="noopener">{esc(p["label"])}</a>'
                    for p in protos
                )
                feat_html += f'<div class="proto-bar"><span class="proto-lab">原型：</span>{links}</div>'
            else:
                feat_html += '<div class="proto-bar muted">原型：待产出</div>'
            day_rows.append(
                f"""<tr>
  {date_cell}
  <td>{esc(it['group'])}<div class="sub">{esc(it['ids'])}</div></td>
  <td class="act">{esc(it['action'])}</td>
  <td class="feat-cell">{feat_html}</td>
  <td><span class="pri {pri_cls}">{esc(pri)}</span></td>
</tr>"""
            )

    cards = []
    for g in GROUPS:
        must = "".join(f"<li>{esc(x)}</li>" for x in g["must"])
        stretch = "".join(f"<li>{esc(x)}</li>" for x in g["stretch"])
        cards.append(
            f"""<section class="card" style="--accent:{g['color']};--bg:{g['bg']}">
  <header><h2>{esc(g['name'])}</h2></header>
  <div class="row"><b>第1周</b><span>{esc(g['w1'])}</span></div>
  <div class="row"><b>第2周</b><span>{esc(g['w2'])}</span></div>
  <div class="two">
    <div><b class="must">Must</b><ul>{must}</ul></div>
    <div><b class="stretch">Stretch</b><ul>{stretch}</ul></div>
  </div>
  <div class="hand"><b>交接包</b> {esc(g['handoff'])}</div>
</section>"""
        )

    def _st_cls(st):
        s = str(st)
        if "开发中" in s:
            return "st-dev"
        if "已完成" in s:
            return "st-done"
        if "暂缓" in s:
            return "st-hold"
        return "st-w3"

    backlog_rows = "".join(
        f'<tr><td>{esc(a)}</td><td>{esc(b)}</td><td>{esc(c)}</td>'
        f'<td><span class="st-pill {_st_cls(d)}">{esc(d)}</span></td><td>{esc(e)}</td></tr>'
        for a, b, c, d, e in BACKLOG
    )
    must_ol = "".join(f"<li>{esc(x)}</li>" for x in MUST_DELIVERABLES)

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>产品→开发 · 原型与设计计划 · 两周版</title>
  <style>
    :root {{
      --bg:#f0f2f5; --card:#fff; --line:#e8e8e8; --text:rgba(0,0,0,.85); --sub:rgba(0,0,0,.45);
      --w1:#cf1322; --w2:#d46b08;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Microsoft YaHei","PingFang SC",Arial,sans-serif; background:var(--bg); color:var(--text); font-size:13px; }}
    a {{ color:#1890ff; text-decoration:none; }}
    a:hover {{ text-decoration:underline; }}
    .top {{
      position:sticky; top:0; z-index:10; background:#001529; color:#fff;
      padding:12px 16px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;
    }}
    .top h1 {{ margin:0; font-size:16px; font-weight:600; }}
    .top .meta {{ color:rgba(255,255,255,.7); font-size:12px; }}
    .top .acts {{ margin-left:auto; display:flex; gap:8px; flex-wrap:wrap; }}
    .top .acts a {{
      color:#fff; background:rgba(255,255,255,.12); padding:4px 10px; border-radius:2px; font-size:12px;
    }}
    .wrap {{ max-width:1280px; margin:0 auto; padding:14px 16px 48px; }}
    .banner {{
      background:linear-gradient(135deg,#1f4e79,#1890ff); color:#fff;
      border-radius:8px; padding:14px 16px; margin-bottom:12px; line-height:1.65;
    }}
    .banner h2 {{ margin:0 0 6px; font-size:15px; }}
    .banner p {{ margin:0; font-size:12px; opacity:.95; }}
    .banner .deadline {{
      margin-top:10px; background:rgba(255,255,255,.15); border-radius:6px; padding:8px 12px; font-weight:600; font-size:12px;
    }}
    .split {{ display:grid; grid-template-columns:1fr 1fr; gap:10px; margin-bottom:12px; }}
    @media (max-width:800px) {{ .split {{ grid-template-columns:1fr; }} }}
    .box {{
      background:#fff; border:1px solid var(--line); border-radius:8px; padding:12px 14px;
    }}
    .box h3 {{ margin:0 0 8px; font-size:13px; }}
    .box.must-box {{ border-top:3px solid var(--w1); }}
    .box.rule-box {{ border-top:3px solid #1890ff; }}
    .box ol, .box ul {{ margin:0; padding-left:18px; line-height:1.7; font-size:12px; }}
    .cards {{ display:grid; grid-template-columns:1fr; gap:10px; margin-bottom:14px; }}
    .card {{
      background:#fff; border:1px solid var(--line); border-radius:8px; overflow:hidden;
      border-top:4px solid var(--accent);
    }}
    .card header {{
      padding:10px 14px; background:var(--bg); border-bottom:1px solid #f0f0f0;
    }}
    .card h2 {{ margin:0; font-size:15px; color:var(--accent); }}
    .card .row {{
      display:grid; grid-template-columns:52px 1fr; gap:8px; padding:8px 14px;
      font-size:12px; line-height:1.55; border-bottom:1px dashed #f0f0f0;
    }}
    .card .row b {{ color:var(--accent); }}
    .card .two {{
      display:grid; grid-template-columns:1fr 1fr; gap:8px; padding:10px 14px; font-size:12px;
    }}
    @media (max-width:700px) {{ .card .two {{ grid-template-columns:1fr; }} }}
    .card .two ul {{ margin:4px 0 0; padding-left:16px; line-height:1.6; }}
    .must {{ color:#cf1322; }}
    .stretch {{ color:#d46b08; }}
    .hand {{
      padding:8px 14px 12px; font-size:11px; color:var(--sub); background:#fafafa;
      border-top:1px solid #f0f0f0;
    }}
    table {{
      width:100%; border-collapse:collapse; background:#fff; border:1px solid var(--line);
      border-radius:8px; overflow:hidden; margin-bottom:14px;
    }}
    th {{
      background:#fafafa; text-align:left; padding:8px 10px; font-size:12px; color:var(--sub);
      border-bottom:1px solid var(--line); position:sticky; top:48px; z-index:5;
    }}
    td {{ padding:7px 10px; border-bottom:1px solid #f0f0f0; vertical-align:top; line-height:1.5; font-size:12px; }}
    td.date {{ background:#fafafa; white-space:nowrap; width:100px; }}
    td.act {{ width:18%; color:#334155; }}
    td.feat-cell {{ width:42%; background:#fafcff; }}
    ul.feat {{ margin:0; padding-left:16px; line-height:1.55; font-size:11px; color:#1f2937; }}
    ul.feat li {{ margin-bottom:2px; }}
    .proto-bar {{
      margin-top:8px; padding-top:6px; border-top:1px dashed #d6e4ff;
      display:flex; flex-wrap:wrap; gap:6px; align-items:center;
    }}
    .proto-lab {{ font-size:11px; color:#8c8c8c; font-weight:600; }}
    a.proto {{
      display:inline-block; font-size:11px; padding:1px 8px; border-radius:2px;
      background:#e6f7ff; color:#0958d9; border:1px solid #91caff; text-decoration:none;
    }}
    a.proto:hover {{ background:#bae0ff; text-decoration:underline; }}
    .proto-bar.muted {{ color:#bfbfbf; font-size:11px; border-top-color:#f0f0f0; }}
    tr.wave-hd td {{ color:#fff; font-weight:600; padding:8px 10px; border:none; }}
    tr.wave-hd.w1 td {{ background:#cf1322; }}
    tr.wave-hd.w2 td {{ background:#d46b08; }}
    .wtag {{ display:inline-block; padding:1px 8px; border-radius:2px; background:rgba(255,255,255,.18); }}
    .sub {{ color:var(--sub); font-size:11px; margin-top:2px; }}
    .sm {{ font-size:11px; color:#334155; }}
    .pri {{
      display:inline-block; padding:0 6px; border-radius:2px; font-size:11px; font-weight:700; border:1px solid;
    }}
    .pri.p0 {{ color:#cf1322; background:#fff1f0; border-color:#ffa39e; }}
    .pri.p1 {{ color:#d46b08; background:#fff7e6; border-color:#ffd591; }}
    .pri.px {{ color:#8c8c8c; background:#fafafa; border-color:#d9d9d9; }}
    h2.sec {{ margin:18px 0 8px; font-size:15px; }}
    .foot {{ text-align:center; font-size:11px; color:var(--sub); margin-top:8px; }}
    .st-pill {{
      display:inline-block; padding:2px 8px; border-radius:4px; font-size:12px; font-weight:600;
      white-space:nowrap;
    }}
    .st-pill.st-dev {{ background:#fff7e6; color:#d46b08; border:1px solid #ffd591; }}
    .st-pill.st-done {{ background:#f6ffed; color:#389e0d; border:1px solid #b7eb8f; }}
    .st-pill.st-w3 {{ background:#f5f5f5; color:#595959; border:1px solid #d9d9d9; }}
    .st-pill.st-hold {{ background:#fff1f0; color:#8c8c8c; border:1px solid #ffccc7; }}
  </style>
</head>
<body>
  <div class="top">
    <h1>产品 → 开发 · 原型与设计计划 · 两周版</h1>
    <span class="meta">08/10–08/21 · 前两周主链路梳理 · 旁支进第三周</span>
    <div class="acts">
      <a href="产品部门-对开发-原型与设计计划表_两周版.xlsx">下载 Excel</a>
      <a href="产品部门-对开发-原型与设计计划表.html">四周滚动版</a>
      <a href="产品部门-各组需求清单-评审版.html">需求清单</a>
      <a href="产品部门-导航.html">返回导航</a>
    </div>
  </div>
  <div class="wrap">
    <div class="banner">
      <h2>前两周只梳主链路；不影响流程的进第三周</h2>
      <p>主链路：询价报价 → 合约价确认 → 提单/运价 → 报关四切片 → 查货/查验 → 进港 → 预估 → 尾单 → 工单边界。旁支进第三周；不含点价订舱。</p>
      <div class="deadline">⏰ 08/10–12 流程图 · 08/13–14 私卡+查货/查验 · 08/17–20 船务/关务/海外主链路 · 08/19 端到端走查 · 08/21 主链路交付包</div>
      <div class="deadline" style="background:rgba(255,255,255,.14);margin-top:8px">📌 08/11 进度：私卡时效报表 BJ002 · 🔧开发中 ｜ 清关放行轨迹 GW021 · ✅已完成（产品侧不重复排）</div>
    </div>

    <div class="split">
      <div class="box must-box">
        <h3>两周 Must（必须交付）</h3>
        <ol>{must_ol}</ol>
      </div>
      <div class="box rule-box">
        <h3>交接最低包 / 待拍默认</h3>
        <ul>
          <li>①可点原型 ②一页PRD ③验收样例 ④待拍清单（带默认）</li>
          <li>查验进预估 → 默认开关关，预留字段</li>
          <li>信号旗 → 默认先不做自动抓</li>
          <li>截单→关务 → 复用待办，产品核实触达</li>
          <li>合约价「确认」→ 周三流程图定规则，周四写入 PRD</li>
        </ul>
      </div>
    </div>

    <h2 class="sec">各组两周目标</h2>
    <div class="cards">
      {''.join(cards)}
    </div>

    <h2 class="sec">日排程（产品动作 + 功能点清单）</h2>
    <table>
      <thead>
        <tr>
          <th style="width:100px">日期</th>
          <th style="width:88px">组别 / ID</th>
          <th style="width:18%">产品动作</th>
          <th>功能点清单（给开发）· 含原型链接</th>
          <th style="width:48px">优先级</th>
        </tr>
      </thead>
      <tbody>
        {''.join(day_rows)}
      </tbody>
    </table>

    <h2 class="sec">旁支与进度（第三周 / 开发中 / 已完成）</h2>
    <table>
      <thead>
        <tr><th>事项</th><th style="width:100px">需求ID</th><th style="width:60px">优先级</th><th style="width:90px">状态</th><th>说明</th></tr>
      </thead>
      <tbody>
        {backlog_rows}
      </tbody>
    </table>

    <p class="foot">前两周 = 主链路梳理 · 第三周 = 旁支 · 四周滚动版作完整 backlog</p>
  </div>
</body>
</html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")
    print("wrote", OUT_HTML)


if __name__ == "__main__":
    # fix accidental bad import line if present — rewrite clean
    write_xlsx()
    write_html()
