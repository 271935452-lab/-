# -*- coding: utf-8 -*-
"""按小组维度生成调研行动清单（发大群用）。"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 复用两周版数据
from gen_research_plan_2weeks import GROUP_EST, PREWORK, REQ_LIST, SCHEDULE

OUT = Path(r"d:\Cursor\头程项目\产品部门\产品部门调研计划_7.27起_各组行动清单.xlsx")

GROUPS = [
    {"key": "关务组", "sheet": "关务组", "prework_key": "关务组"},
    {"key": "风控组", "sheet": "风控组", "prework_key": "风控组"},
    {"key": "船务组", "sheet": "船务组", "prework_key": "船务组"},
    {"key": "报价组", "sheet": "报价组", "prework_key": "报价组"},
    {"key": "海外对接组", "sheet": "海外对接组", "prework_key": "海外组"},
]

INTRO = [
    {"说明项": "文件用途", "内容": "产品部门业务调研 · 各组行动清单（基于两周版计划整理，发群用）"},
    {"说明项": "调研周期", "内容": "2026-07-27（周一下午）正式开场 ~ 2026-08-07（周五）收敛；会前材料 7/26 截止"},
    {"说明项": "各组怎么看", "内容": "请打开本组对应 Sheet：关务组 / 风控组 / 船务组 / 报价组 / 海外对接组"},
    {"说明项": "会前必做（所有业务组）", "内容": "7/26 前提交样例材料 + 对需求清单做 P0/P1/P2 初判 + 指定接口人"},
    {"说明项": "跨组事项", "内容": "见 Sheet「跨组与全员事项」：跨组专题派代表、8/6 补访、8/7 主管优先级工作坊"},
    {"说明项": "完整版计划", "内容": "详见：产品部门调研计划_7.27起_两周版.xlsx"},
]

CROSS_ALL = [
    {
        "日期": "2026-07-26",
        "星期": "周六",
        "时段": "截止",
        "事项": "【会前】提交样例材料",
        "各组动作": "按本组 Sheet 顶部「会前材料」清单提交；指定1名接口人",
        "参与": "各组主管指定接口人",
    },
    {
        "日期": "2026-08-05",
        "星期": "周三",
        "时段": "14:00-17:30",
        "事项": "跨组① · 业务数据统计字段",
        "各组动作": "各组派 1 名代表 + 准备本组统计字段缺口",
        "参与": "各组代表 + 产品",
    },
    {
        "日期": "2026-08-06",
        "星期": "周四",
        "时段": "09:30-12:30",
        "事项": "跨组② · 7.18轨迹10条 + 公开通知页",
        "各组动作": "各组派 1 名代表；船务/海外/报价重点参与轨迹规则",
        "参与": "各组 + 客服 + 研发",
    },
    {
        "日期": "2026-08-06",
        "星期": "周四",
        "时段": "14:00-17:30",
        "事项": "补访 + 研发方案初评 + 原型走查",
        "各组动作": "按开放问题清单补访；P0 需求负责人到场或线上 15min",
        "参与": "各组 + 产品 + 研发",
    },
    {
        "日期": "2026-08-07",
        "星期": "周五",
        "时段": "14:00-16:00",
        "事项": "优先级确认工作坊",
        "各组动作": "各组主管必到；确认本组需求 P0/P1/P2",
        "参与": "各组主管 + 产品 + 研发负责人",
    },
]

CALENDAR = [
    {"日期": "7/24-7/26", "阶段": "会前", "关务组": "提交材料", "风控组": "提交材料", "船务组": "提交材料", "报价组": "提交材料", "海外对接组": "提交材料"},
    {"日期": "7/27 下午", "阶段": "W1", "关务组": "专场①②", "风控组": "-", "船务组": "-", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/28", "阶段": "W1", "关务组": "专场③④", "风控组": "-", "船务组": "-", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/29 上午", "阶段": "W1", "关务组": "联合⑤", "风控组": "联合⑤", "船务组": "-", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/29 下午", "阶段": "W1", "关务组": "-", "风控组": "专场⑥", "船务组": "-", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/30 上午", "阶段": "W1", "关务组": "-", "风控组": "专场⑦", "船务组": "-", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/30 下午", "阶段": "W1", "关务组": "-", "风控组": "-", "船务组": "专场⑧", "报价组": "-", "海外对接组": "-"},
    {"日期": "7/31", "阶段": "W1", "关务组": "-", "风控组": "-", "船务组": "专场⑨⑩", "报价组": "派员", "海外对接组": "-"},
    {"日期": "8/3", "阶段": "W2", "关务组": "-", "风控组": "-", "船务组": "派员", "报价组": "专场⑪⑫", "海外对接组": "-"},
    {"日期": "8/4 上午", "阶段": "W2", "关务组": "-", "风控组": "-", "船务组": "-", "报价组": "联合⑬", "海外对接组": "联合⑬"},
    {"日期": "8/4 下午", "阶段": "W2", "关务组": "-", "风控组": "-", "船务组": "-", "报价组": "-", "海外对接组": "专场⑭"},
    {"日期": "8/5 上午", "阶段": "W2", "关务组": "-", "风控组": "-", "船务组": "-", "报价组": "-", "海外对接组": "专场⑮"},
    {"日期": "8/5 下午", "阶段": "W2", "关务组": "派员", "风控组": "派员", "船务组": "派员", "报价组": "派员", "海外对接组": "派员"},
    {"日期": "8/6", "阶段": "W2", "关务组": "派员+补访", "风控组": "派员+补访", "船务组": "派员+补访", "报价组": "派员+补访", "海外对接组": "派员+补访"},
    {"日期": "8/7 下午", "阶段": "收敛", "关务组": "主管到场", "风控组": "主管到场", "船务组": "主管到场", "报价组": "主管到场", "海外对接组": "主管到场"},
]


def _match_group(obj: str, group_key: str) -> bool:
    if group_key == "关务组":
        return "关务" in obj
    if group_key == "风控组":
        return "风控" in obj
    if group_key == "船务组":
        return "船务" in obj
    if group_key == "报价组":
        return "报价" in obj
    if group_key == "海外对接组":
        return "海外" in obj
    return group_key in obj


def _session_type(obj: str, group_key: str) -> str:
    if "【会前】" in obj:
        return "会前准备"
    if "+" in obj and _match_group(obj, group_key):
        others = obj.replace(group_key.replace("对接组", "组"), "").replace(group_key, "")
        if "+" in obj:
            return "联合调研"
    if obj in ("跨组专题",):
        return "跨组（派代表）"
    if obj in ("补访+方案评审", "各组主管"):
        return "全员配合"
    if _match_group(obj, group_key) and obj.count("+") == 0:
        return "本组专场"
    return "配合"


def sessions_for(group_key: str) -> list[dict]:
    rows = []
    for s in SCHEDULE:
        obj = s["调研对象"]
        if obj.startswith("【会前】各组") or obj == "【会前】各组":
            rows.append({**s, "场次类型": "会前准备", "您组需做什么": "提交样例材料 + 优先级初判 + 确认接口人"})
            continue
        if _match_group(obj, group_key):
            stype = _session_type(obj, group_key)
            action = "准时到场，按「准备材料」带样例；现场演示痛点 + 确认规则"
            if "+" in obj:
                action = "联合场次：带本组相关案例，协同确认规则边界"
            rows.append({**s, "场次类型": stype, "您组需做什么": action})
    # 跨组与收敛
    for c in CROSS_ALL:
        if c["事项"].startswith("【会前】"):
            continue
        rows.append({
            "日期": c["日期"],
            "星期": c["星期"],
            "时段": c["时段"],
            "周次": "W2",
            "调研对象": "跨组/收敛",
            "调研重点": c["事项"],
            "主题": c["事项"],
            "时长(h)": "-",
            "地点/形式": "会议室",
            "参与人": c["参与"],
            "准备材料": "见跨组 Sheet",
            "当日目标": c["各组动作"],
            "场次类型": "跨组/收敛" if "主管" not in c["事项"] else "主管必到",
            "您组需做什么": c["各组动作"],
        })
    return rows


def reqs_for(group_key: str) -> list[dict]:
    g = group_key if group_key != "海外对接组" else "海外组"
    if g == "关务组":
        g = "关务组"
    rows = [r for r in REQ_LIST if r["组别"] == g or (group_key == "海外对接组" and r["组别"] == "海外组")]
    return rows


def group_meta(group_key: str) -> dict:
    for g in GROUP_EST:
        if g["组别"] == group_key:
            return g
    return {}


def prework_for(prework_key: str) -> dict:
    for p in PREWORK:
        if p["组别"] == prework_key:
            return p
    return {}


def build_group_overview(group_key: str, prework_key: str) -> pd.DataFrame:
    meta = group_meta(group_key)
    pw = prework_for(prework_key)
    return pd.DataFrame([
        {"项目": "组别", "内容": group_key},
        {"项目": "调研时间", "内容": meta.get("场次安排", "")},
        {"项目": "调研重点", "内容": meta.get("调研重点", "").replace("★ ", "")},
        {"项目": "建议到场人员", "内容": meta.get("建议参与人", "")},
        {"项目": "会前截止", "内容": f"{pw.get('截止', '7/26')} · 接口人：{pw.get('接口人', '主管指定')}"},
        {"项目": "会前需提交", "内容": pw.get("需提交材料", "")},
        {"项目": "调研后产出", "内容": meta.get("预期产出", "")},
        {"项目": "本组需求条数", "内容": str(meta.get("需求条目数", len(reqs_for(group_key))))},
    ])


def build_group_sessions(group_key: str) -> pd.DataFrame:
    rows = []
    for i, s in enumerate(sessions_for(group_key), 1):
        rows.append({
            "序号": i,
            "日期": s.get("日期", ""),
            "星期": s.get("星期", ""),
            "时段": s.get("时段", ""),
            "场次类型": s.get("场次类型", ""),
            "主题": s.get("主题", ""),
            "调研重点": s.get("调研重点", ""),
            "您组需做什么": s.get("您组需做什么", s.get("准备材料", "")),
            "需携带材料": s.get("准备材料", ""),
            "参与人": s.get("参与人", ""),
            "当日目标": s.get("当日目标", ""),
        })
    return pd.DataFrame(rows)


def build_group_reqs(group_key: str) -> pd.DataFrame:
    rows = []
    for i, r in enumerate(reqs_for(group_key), 1):
        rows.append({
            "序号": i,
            "需求/问题": r["需求/问题"],
            "类型": r["类型"],
            "初判优先级": r["初步优先级"],
            "对应场次": r.get("调研场次(W2)", r.get("调研场次", "")),
            "会前请您标注": "确认/调整优先级，补充遗漏项",
        })
    return pd.DataFrame(rows)


def style_header_row(ws, row=1, fill="1F4E79"):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hfill = PatternFill("solid", fgColor=fill)
    hfont = Font(color="FFFFFF", bold=True)
    for col in range(1, ws.max_column + 1):
        c = ws.cell(row=row, column=col)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border


def style_data_area(ws, start_row=2, wrap_height=52):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in range(start_row, ws.max_row + 1):
        ws.row_dimensions[row].height = wrap_height
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap
            cell.border = border


def write_group_sheet(writer, group_key: str, prework_key: str, sheet_name: str):
    overview = build_group_overview(group_key, prework_key)
    sessions = build_group_sessions(group_key)
    reqs = build_group_reqs(group_key)

    # 块起始行（1-based）：标题1 · 概况2 · 场次标题12 · 需求标题动态
    overview.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    sess_title_row = 1 + 1 + len(overview) + 1
    sess_hdr_row = sess_title_row + 1
    sessions.to_excel(writer, sheet_name=sheet_name, index=False, startrow=sess_hdr_row)
    req_title_row = sess_hdr_row + 1 + len(sessions) + 1
    req_hdr_row = req_title_row + 1
    reqs.to_excel(writer, sheet_name=sheet_name, index=False, startrow=req_hdr_row)

    ws = writer.sheets[sheet_name]
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=f"【{group_key}】调研行动清单 — 请本组对照执行").font = Font(
        bold=True, size=14, color="0958D9"
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    sess_title_row += 1
    sess_hdr_row += 1
    req_title_row += 1
    req_hdr_row += 1

    sec_font = Font(bold=True, size=12, color="0958D9")
    sec_fill = PatternFill("solid", fgColor="E6F7FF")
    ws.insert_rows(sess_title_row)
    ws.cell(row=sess_title_row, column=1, value="▎ 您的调研场次（请相关人员预留时间）").font = sec_font
    ws.cell(row=sess_title_row, column=1).fill = sec_fill
    sess_hdr_row += 1
    req_title_row += 1
    req_hdr_row += 1

    ws.insert_rows(req_title_row)
    ws.cell(row=req_title_row, column=1, value="▎ 本组需求清单（会前请标注 P0/P1/P2）").font = sec_font
    ws.cell(row=req_title_row, column=1).fill = sec_fill
    req_hdr_row += 1

    widths = {"A": 8, "B": 14, "C": 8, "D": 14, "E": 12, "F": 34, "G": 30, "H": 26, "I": 22, "J": 18, "K": 26}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    style_header_row(ws, row=2)
    style_header_row(ws, row=sess_hdr_row)
    style_header_row(ws, row=req_hdr_row)
    style_data_area(ws, 3, 36)
    style_data_area(ws, sess_hdr_row + 1, 52)
    style_data_area(ws, req_hdr_row + 1, 48)

    action_fill = PatternFill("solid", fgColor="FFF7E6")
    for row in range(sess_hdr_row + 1, sess_hdr_row + 1 + len(sessions)):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=sess_hdr_row, column=col).value == "您组需做什么":
                ws.cell(row=row, column=col).fill = action_fill
                ws.cell(row=row, column=col).font = Font(bold=True, color="AD4E00")


def main():
    with pd.ExcelWriter(OUT, engine="openpyxl") as w:
        pd.DataFrame(INTRO).to_excel(w, sheet_name="发群说明", index=False)
        pd.DataFrame(CALENDAR).to_excel(w, sheet_name="总日程一览", index=False)
        pd.DataFrame(CROSS_ALL).to_excel(w, sheet_name="跨组与全员事项", index=False)
        for g in GROUPS:
            write_group_sheet(w, g["key"], g["prework_key"], g["sheet"])

    wb = load_workbook(OUT)
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == "发群说明":
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 72
        elif sn == "总日程一览":
            for c in "ABCDEFGH":
                ws.column_dimensions[c].width = 14
        elif sn == "跨组与全员事项":
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["D"].width = 28
            ws.column_dimensions["E"].width = 42
            ws.column_dimensions["F"].width = 28
        style_header_row(ws, row=1)
        style_data_area(ws, 2)
    wb.save(OUT)
    print(f"已生成: {OUT}")


if __name__ == "__main__":
    main()
