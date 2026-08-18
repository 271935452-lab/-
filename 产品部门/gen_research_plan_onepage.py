# -*- coding: utf-8 -*-
"""生成「各组一屏总览」— 单 Sheet Excel + 可截图 HTML（发群用）。"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from gen_research_plan_2weeks import GROUP_EST, PREWORK, SCHEDULE

OUT_XLSX = Path(r"d:\Cursor\头程项目\产品部门\产品部门调研计划_7.27起_发群一屏.xlsx")
OUT_HTML = Path(r"d:\Cursor\头程项目\产品部门\产品部门调研计划_7.27起_发群一屏.html")

GROUPS = [
    {"key": "关务组", "color": "#1890ff", "bg": "#e6f7ff", "prework": "关务组"},
    {"key": "风控组", "color": "#722ed1", "bg": "#f9f0ff", "prework": "风控组"},
    {"key": "船务组", "color": "#13c2c2", "bg": "#e6fffb", "prework": "船务组"},
    {"key": "报价组", "color": "#fa8c16", "bg": "#fff7e6", "prework": "报价组"},
    {"key": "海外对接组", "color": "#52c41a", "bg": "#f6ffed", "prework": "海外组"},
]

ALL_HANDS = [
    ("7/26 前", "【所有组】提交会前材料 + 指定接口人 + 需求优先级初判"),
    ("7/29 上午", "【关务+风控】合并报关 & 国内查验联合场（各组带案例）"),
    ("7/31 下午", "【船务+报价】运价/仓位边界（报价组派代表）"),
    ("8/4 上午", "【报价+海外】私卡轨迹/POD 联合场"),
    ("8/5 下午", "【跨组】业务数据统计字段 — 各组派 1 代表"),
    ("8/6 上午", "【跨组】7.18 轨迹 10 条 — 各组派 1 代表"),
    ("8/6 下午", "【补访】按开放问题清单线上/现场 15min"),
    ("8/7 下午", "【各组主管必到】优先级确认工作坊 14:00-16:00"),
]


def _match(obj: str, key: str) -> bool:
    m = {"关务组": "关务", "风控组": "风控", "船务组": "船务", "报价组": "报价", "海外对接组": "海外"}
    return m.get(key, key) in obj


def _short_date(d: str) -> str:
    if not d or d.startswith("2026-") is False:
        return d
    parts = d.split("-")
    return f"{int(parts[1])}/{int(parts[2])}"


def _short_theme(theme: str) -> str:
    if "：" in theme:
        theme = theme.split("：", 1)[-1]
    for prefix in ("① ", "② ", "③ ", "④ ", "⑤ ", "⑥ ", "⑦ ", "⑧ ", "⑨ ", "⑩ ",
                   "⑪ ", "⑫ ", "⑬ ", "⑭ ", "⑮ ", "⑯ ", "⑰ ", "⑱ ", "⑲ ", "⑳ ", "㉑ "):
        if theme.startswith(prefix):
            theme = theme[len(prefix):]
    return theme[:42] + ("…" if len(theme) > 42 else "")


def group_sessions(key: str) -> list[dict]:
    rows = []
    for s in SCHEDULE:
        obj = s["调研对象"]
        if obj.startswith("【会前】各组"):
            continue
        if _match(obj, key):
            rows.append({
                "日期": _short_date(s["日期"]),
                "星期": s.get("星期", ""),
                "时段": s["时段"],
                "内容": _short_theme(s["主题"]),
                "准备": s.get("准备材料", ""),
                "类型": "联合" if "+" in obj and key not in obj else "专场",
            })
    return rows


def group_meta(key: str) -> dict:
    for g in GROUP_EST:
        if g["组别"] == key:
            return g
    return {}


def group_prework(pk: str) -> str:
    for p in PREWORK:
        if p["组别"] == pk:
            return p["需提交材料"]
    return ""


def build_flat_rows() -> list[dict]:
    """单 Sheet 平铺：每组一块，含会前+各场次。"""
    rows = []
    rows.append({
        "组别": "📌 全员",
        "类型": "会前",
        "时间": "7/26 前截止",
        "安排/场次": "提交样例材料 · 指定接口人 · 需求 P0/P1/P2 初判",
        "需准备": "见本表各组「会前准备」行",
        "到场人员": "各组主管指定接口人",
    })
    for ah in ALL_HANDS:
        rows.append({
            "组别": "📌 全员",
            "类型": "跨组/收敛",
            "时间": ah[0],
            "安排/场次": ah[1],
            "需准备": "按事项说明",
            "到场人员": "见说明",
        })

    for g in GROUPS:
        key = g["key"]
        meta = group_meta(key)
        focus = meta.get("调研重点", "").replace("★ ", "")
        rows.append({
            "组别": key,
            "类型": "▶ 概况",
            "时间": meta.get("场次安排", ""),
            "安排/场次": f"重点：{focus}",
            "需准备": "—",
            "到场人员": meta.get("建议参与人", ""),
        })
        pw = next((p for p in PREWORK if p["组别"] == g["prework"]), {})
        rows.append({
            "组别": key,
            "类型": "⚠ 会前",
            "时间": "7/26 前",
            "安排/场次": "提交会前材料（必做）",
            "需准备": group_prework(g["prework"]),
            "到场人员": pw.get("接口人", "主管指定"),
        })
        for s in group_sessions(key):
            rows.append({
                "组别": key,
                "类型": s["类型"],
                "时间": f"{s['日期']} {s['星期']} {s['时段']}",
                "安排/场次": s["内容"],
                "需准备": s["准备"],
                "到场人员": meta.get("建议参与人", "").split("+")[0][:20],
            })
        rows.append({"组别": key, "类型": "", "时间": "", "安排/场次": "", "需准备": "", "到场人员": ""})
    return rows


def build_html() -> str:
    cards = []
    for g in GROUPS:
        key = g["key"]
        meta = group_meta(key)
        sessions = group_sessions(key)
        sess_rows = "".join(
            f"""<tr>
              <td class="t">{s['日期']}<br><span class="sub">{s['时段']}</span></td>
              <td>{s['内容']}</td>
              <td class="prep">{s['准备']}</td>
            </tr>"""
            for s in sessions
        )
        cards.append(f"""
        <section class="card" style="--accent:{g['color']};--bg:{g['bg']}">
          <header>
            <h2>{key}</h2>
            <span class="badge">{meta.get('场次安排','')}</span>
          </header>
          <div class="focus"><b>调研重点</b> {meta.get('调研重点','').replace('★ ','')}</div>
          <div class="people"><b>建议到场</b> {meta.get('建议参与人','')}</div>
          <div class="prework">
            <b>⚠ 7/26 前必交</b>
            <p>{group_prework(g['prework'])}</p>
          </div>
          <table>
            <thead><tr><th width="88">时间</th><th>做什么</th><th width="120">带什么</th></tr></thead>
            <tbody>{sess_rows}</tbody>
          </table>
        </section>""")

    all_rows = "".join(f"<li><b>{t}</b> {d}</li>" for t, d in ALL_HANDS)

    return f"""<!doctype html>
<html lang="zh-CN"><head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=750"/>
<title>产品部门调研 · 各组行动一览</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: "Microsoft YaHei","PingFang SC",sans-serif;
  font-size: 13px; color: #262626; background: #f5f5f5;
  width: 750px; margin: 0 auto; padding: 16px 14px 24px;
}}
.banner {{
  background: linear-gradient(135deg,#1f4e79,#1890ff);
  color: #fff; border-radius: 10px; padding: 16px 18px; margin-bottom: 14px;
}}
.banner h1 {{ font-size: 18px; margin-bottom: 8px; }}
.banner p {{ font-size: 12px; line-height: 1.65; opacity: .95; }}
.banner .deadline {{
  margin-top: 10px; background: rgba(255,255,255,.15);
  border-radius: 6px; padding: 8px 12px; font-weight: 600;
}}
.all {{
  background: #fff; border: 1px solid #ffd591; border-radius: 10px;
  padding: 12px 14px; margin-bottom: 14px;
}}
.all h3 {{ color: #d46b08; font-size: 14px; margin-bottom: 8px; }}
.all ul {{ padding-left: 18px; line-height: 1.75; font-size: 12px; }}
.card {{
  background: #fff; border-radius: 10px; margin-bottom: 12px;
  border: 1px solid #e8e8e8; overflow: hidden;
  border-top: 4px solid var(--accent);
}}
.card header {{
  display: flex; justify-content: space-between; align-items: center;
  padding: 10px 14px; background: var(--bg);
  border-bottom: 1px solid #f0f0f0;
}}
.card h2 {{ font-size: 16px; color: var(--accent); }}
.badge {{ font-size: 11px; color: #595959; background: #fff;
  padding: 2px 8px; border-radius: 4px; border: 1px solid #d9d9d9; }}
.focus, .people {{ padding: 8px 14px; font-size: 12px; line-height: 1.55; border-bottom: 1px dashed #f0f0f0; }}
.prework {{
  margin: 10px 14px; padding: 10px 12px; background: #fff7e6;
  border: 1px solid #ffd591; border-radius: 6px; font-size: 12px; line-height: 1.6;
}}
.prework b {{ color: #d46b08; display: block; margin-bottom: 4px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
th {{ background: #fafafa; color: #8c8c8c; font-weight: 500;
  padding: 8px 10px; text-align: left; border-bottom: 1px solid #f0f0f0; }}
td {{ padding: 8px 10px; border-bottom: 1px solid #f5f5f5; vertical-align: top; line-height: 1.5; }}
td.t {{ white-space: nowrap; font-weight: 600; color: var(--accent); }}
td.t .sub {{ font-weight: 400; color: #8c8c8c; font-size: 11px; }}
td.prep {{ color: #595959; font-size: 11px; }}
.footer {{ text-align: center; font-size: 11px; color: #8c8c8c; margin-top: 8px; }}
@media print {{ body {{ width: 750px; }} }}
</style></head><body>
<div class="banner">
  <h1>产品部门业务调研 · 各组行动一览</h1>
  <p>周期：7/27（周一）下午开场 → 8/7（周五）收敛 · 共两周<br/>
  请各组对照下方卡片，预留专场时间、按时提交会前材料。</p>
  <div class="deadline">⏰ 所有组：7/26（周六）前提交会前材料 + 指定 1 名接口人</div>
</div>
<div class="all">
  <h3>📌 全员还需配合（除本组专场外）</h3>
  <ul>{all_rows}</ul>
</div>
{"".join(cards)}
<p class="footer">截图发群即可 · 详细版见「产品部门调研计划_7.27起_各组行动清单.xlsx」</p>
</body></html>"""


def style_sheet(ws):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(color="FFFFFF", bold=True)

    for col, w in zip("ABCDEF", [10, 8, 22, 38, 36, 22]):
        ws.column_dimensions[col].width = w

    for c in range(1, 7):
        cell = ws.cell(row=1, column=c)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    group_fills = {
        "关务组": "E6F7FF", "风控组": "F9F0FF", "船务组": "E6FFFB",
        "报价组": "FFF7E6", "海外对接组": "F6FFED", "📌 全员": "FFFBE6",
    }
    prework_fill = PatternFill("solid", fgColor="FFF7E6")

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 48
        grp = ws.cell(row=row, column=1).value or ""
        typ = ws.cell(row=row, column=2).value or ""
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap
            cell.border = border
            if grp in group_fills and col == 1:
                cell.fill = PatternFill("solid", fgColor=group_fills[grp])
                cell.font = Font(bold=True)
            if typ == "⚠ 会前":
                for c in range(1, 7):
                    ws.cell(row=row, column=c).fill = prework_fill


def main():
    df = pd.DataFrame(build_flat_rows())
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="各组一屏总览", index=False)

    wb = load_workbook(OUT_XLSX)
    style_sheet(wb["各组一屏总览"])
    wb.save(OUT_XLSX)

    OUT_HTML.write_text(build_html(), encoding="utf-8")
    print(f"已生成 Excel: {OUT_XLSX}")
    print(f"已生成 HTML:  {OUT_HTML}")
    print("发群建议：浏览器打开 HTML → 截图（或整页长截图）→ 发企微/微信群")


if __name__ == "__main__":
    main()
