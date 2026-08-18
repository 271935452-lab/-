# -*- coding: utf-8 -*-
"""生成《产品→开发 · 原型与设计交付计划表》HTML + Excel。

口径：
  · 面向「产品输出给开发」的原型 / PRD / 交互设计交付节奏（非研发开发排期）
  · 来源：各组需求清单评审版 + 0727–0807 调研/原型评审纪要
  · 今日：2026-08-10（周一）起按周滚动
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DIR = Path(__file__).resolve().parent
OUT_XLSX = DIR / "产品部门-对开发-原型与设计计划表.xlsx"
OUT_HTML = DIR / "产品部门-对开发-原型与设计计划表.html"

# 波次 × 交付包
PACKAGES = [
    # —— W0：本周可交接 / 收口 ——
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "风控组",
        "pkg": "查货操作合规包",
        "ids": "FK002–FK006",
        "pri": "P0",
        "deliver": "原型回改定稿 + 交互说明 + 开发对接清单",
        "status": "有原型·待复审",
        "gate": "按 8/6 回改（七类标记/免查同页/备注可选/直客只加成本）→ 郭佳玉复审通过",
        "artifacts": "风控组-查货操作-AI识图品名识别-原型.html；风控组-0729-查货管理-大致方案.html；风控组调研纪要_2026-08-06",
        "dev_ready": "复审通过后即可排期",
        "owner": "产品·风控",
        "note": "豆包辅助不作唯一卡控；罚款500/扣货告知与授权联动",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "风控组",
        "pkg": "操作查货率报表",
        "ids": "FK001",
        "pri": "P1",
        "deliver": "MVP 定稿 + 字段口径表（总票/空白/保函/漏查）",
        "status": "有原型",
        "gate": "与查货提成统计口径对齐（查出算提成；月初半小时痛点产品化可后置）",
        "artifacts": "风控组-操作查货率报表-MVP.html",
        "dev_ready": "可先对接字段与统计规则",
        "owner": "产品·风控",
        "note": "8/7：导出量级约 5 万条，导出能力见 PT002",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "船务组",
        "pkg": "提单管理（按周/分配/费用水单/IT）",
        "ids": "CW002,CW012,CW024 + 费用登记",
        "pri": "P0",
        "deliver": "MVP 定稿 + PRD 抽屉要点 + UI 批注回改",
        "status": "有原型",
        "gate": "17:25 UI：去多余分组、按指定截图样式；装完色标仅未装完/装完；发送标记与 AMS/ISF 一体",
        "artifacts": "船务组-提单管理-按周分组与分配-MVP.html；船务组-0730-费用节点提单配仓-大致方案.html",
        "dev_ready": "UI 回改后可交接核心能力清单",
        "owner": "产品·船务",
        "note": "取消 IT 仅在船务提单侧操作；清关只读展示",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "船务组",
        "pkg": "运价管理（我司/代理订舱）+ 提单创建 AI",
        "ids": "CW008,CW010",
        "pri": "P0",
        "deliver": "MVP + 价差对比规则说明",
        "status": "有原型",
        "gate": "合约价旁「确认」按钮：周一对照私卡页再定（现先留参考）",
        "artifacts": "船务组-运价管理-我司与代理订舱-MVP.html；船务组-提单创建批量-AI识别-MVP.html",
        "dev_ready": "点价订舱主流程可先对接；确认按钮可开关配置",
        "owner": "产品·船务",
        "note": "活动价后置；代理价 vs 我司合约价对比报表可 W2",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "海外对接组",
        "pkg": "提单主单预估 + 尾程运单预估",
        "ids": "HW001 相关·预估主线",
        "pri": "P0",
        "deliver": "PRD 定稿 + MVP 对齐 8/7 四维价表口径",
        "status": "有原型/有PRD",
        "gate": "查验费是否进提单预估仍待拍；取消 IT 操作费已确认",
        "artifacts": "海外对接组-提单主单预估费用表-MVP/PRD；尾程预估费用表-MVP/PRD",
        "dev_ready": "头程四维+拖车/报关/EDI 可对接；查验费做开关",
        "owner": "产品·海外",
        "note": "头程：代理×船司×始发港×柜型；尾程按运单；海外仓费进预估",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "海外对接组",
        "pkg": "港后尾单跟踪 + 卡派跟进(新)",
        "ids": "HW005",
        "pri": "P0",
        "deliver": "原型定稿 + 工单跳转说明",
        "status": "有原型",
        "gate": "异常/改派送已迁工单；尾单保留「转工单」",
        "artifacts": "海外对接组-港后-尾单单票跟踪-原型.html；卡派跟进新-MVP.html；工单提交入口",
        "dev_ready": "可交接港后四桶与超时维",
        "owner": "产品·海外",
        "note": "约仓分桶仅私卡派/卡车派送",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "关务组",
        "pkg": "查验报表（提单/运单维）备注落点",
        "ids": "GW020 部分 + 查验报表",
        "pri": "P0/P1",
        "deliver": "原型回改：功能栏备注；登记人=跟进人",
        "status": "有原型·口径已拍",
        "gate": "8/7 张珊珊确认：查验页面功能栏备注即可",
        "artifacts": "关务组-查验报表-提单运单维度-MVP.html",
        "dev_ready": "备注字段可进本周开发对接",
        "owner": "产品·关务",
        "note": "国内/国外分开、去重复时间轴（UI 批注）",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "关务组",
        "pkg": "报关/资料核准核心链",
        "ids": "GW001–GW007,GW010,GW011,GW016,GW026",
        "pri": "P0",
        "deliver": "已有原型包梳理「开发切片」+ 依赖顺序",
        "status": "有原型",
        "gate": "分单号规则表 + 双通道 + 改配轨迹 作为第一切片",
        "artifacts": "报关管理系列 MVP；资料核准双通道；改配三类场景；进口商/清关行配置；分单号规则",
        "dev_ready": "建议切片：①分单号 ②资料核准双通道 ③改配轨迹 ④买单留底",
        "owner": "产品·关务",
        "note": "案例型条目(GW027–031)沉淀为查验处置模板，非独立页面",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "报价组",
        "pkg": "新 AI 报价 + 报价时效看板",
        "ids": "BJ001,BJ002",
        "pri": "P0",
        "deliver": "MVP/规则配置定稿 + 有效数据 PRD",
        "status": "有原型",
        "gate": "周一对照私卡：合约价旁「确认」去留",
        "artifacts": "ESS询价报价-报价员列表-MVP；AI报价规则配置-MVP；报价时效统计看板；0731方案",
        "dev_ready": "确认按钮定稿后可完整交接；看板可并行",
        "owner": "产品·报价",
        "note": "旧 AI 取消；散货多绑运单；713 六仓距离测价",
    },
    {
        "wave": "W0 本周可交接",
        "week": "08/10–08/14",
        "group": "报价组",
        "pkg": "出库/签收/POD 预警",
        "ids": "BJ004",
        "pri": "P0",
        "deliver": "验收问题清单闭环 → 设计变更纪要（若有）",
        "status": "验收中",
        "gate": "验收问题清零或明确延期项",
        "artifacts": "B21-出库签收预警-MVP；订单管理-POD下载；客户API-POD对接说明",
        "dev_ready": "以验收结论为准；变更走小版本",
        "owner": "产品·报价",
        "note": "与客户 API POD 对接说明一并交付",
    },
    # —— W1：补原型/口径 ——
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "跨组",
        "pkg": "海外工单端到端",
        "ids": "（新建·流程梳完后挂号）",
        "pri": "P0",
        "deliver": "流程梳理稿 → 场景表 → 交互原型 → 开发对接",
        "status": "待梳流程",
        "gate": "指定人梳完流程后再约评审对稿（8/7 收口）",
        "artifacts": "工单/工单提交-入口-MVP；工单工作台-MVP（承接尾单异常/改派送）",
        "dev_ready": "流程评审通过后进入原型周，再交接开发",
        "owner": "产品·海外+工单",
        "note": "先流程后人效字段，避免边梳边做",
    },
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "船务组",
        "pkg": "进港时间口径与落点",
        "ids": "CW001 部分",
        "pri": "P0",
        "deliver": "口径一页纸 + 字段落点（抓取源/用途/异常）→ 补进提单/运价方案",
        "status": "待确认",
        "gate": "谁抓、用在哪、异常怎么用（多次点名仍待确认）",
        "artifacts": "船务组-0730-费用节点提单配仓-大致方案（F1 落箱费）",
        "dev_ready": "口径确认后补字段设计再开发",
        "owner": "产品·船务",
        "note": "倾向国内出口码头进港；散货费用异常/海运费×进港提示",
    },
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "海外对接组",
        "pkg": "私卡尾端费用预估自动比对（柜维）",
        "ids": "HW001",
        "pri": "P0",
        "deliver": "交互原型（柜维比对差异清单）+ 规则说明",
        "status": "待原型",
        "gate": "与主单/尾程预估字段对齐后再画",
        "artifacts": "依赖提单主单预估/尾程预估 PRD",
        "dev_ready": "原型评审通过后 W2 交接",
        "owner": "产品·海外",
        "note": "清单唯一「待原型」P0",
    },
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "风控/关务",
        "pkg": "信号旗：清关 IT / 保税号抓取可行性",
        "ids": "FK 关联 + 清关",
        "pri": "P1",
        "deliver": "技术可行性结论 → 决定是否做原型",
        "status": "待问技术",
        "gate": "问技术能否抓；不能则智能历史判断方案",
        "artifacts": "清关管理-MVP；风控 8/6 纪要",
        "dev_ready": "结论出来再排",
        "owner": "产品·风控+技术",
        "note": "授权分国内国外+导出模版可并行设计",
    },
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "关务组",
        "pkg": "清关放行轨迹文案 + 选项",
        "ids": "GW021",
        "pri": "P1",
        "deliver": "原型改文案 + 选项交互",
        "status": "方案中",
        "gate": "「清关已放行，等待提柜」+ 选项「等待安排上火车」",
        "artifacts": "清关管理-MVP.html",
        "dev_ready": "改动小，可随清关切片交接",
        "owner": "产品·关务",
        "note": "取消 IT 不在清关侧",
    },
    {
        "wave": "W1 补原型与口径",
        "week": "08/17–08/21",
        "group": "报价组",
        "pkg": "LTL 轨迹自动抓取方案收口",
        "ids": "BJ003",
        "pri": "P0",
        "deliver": "方案→原型（抓取源/失败重试/展示位）",
        "status": "方案中",
        "gate": "与私卡轨迹/POD 链路对齐",
        "artifacts": "0731 方案；客户API-POD对接说明",
        "dev_ready": "原型出后交接",
        "owner": "产品·报价",
        "note": "快递查询慢为 P3 体验项",
    },
    # —— W2：方案转原型 ——
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "海外对接组",
        "pkg": "目的港邮箱工作台",
        "ids": "HW003",
        "pri": "P0",
        "deliver": "工作台原型（邮箱洪峰/费用类型/水单回邮→业务员成本）",
        "status": "有方案",
        "gate": "Foxmail 回复能力、还柜抓取待确认项列出",
        "artifacts": "0804 海外对接组大致方案；0804 全天整合纪要",
        "dev_ready": "原型评审后交接",
        "owner": "产品·海外",
        "note": "~103 邮箱；滞港/滞箱/AN/到达通知",
    },
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "海外对接组",
        "pkg": "港前 HOLD/提柜预警 + LFD",
        "ids": "HW004,HW012",
        "pri": "P0",
        "deliver": "预警规则原型 + DO/PU 齐套检查清单",
        "status": "有方案",
        "gate": "HOLD 抓取源确认；LFD 提醒对象/渠道",
        "artifacts": "0804 方案；主单跟进-MVP",
        "dev_ready": "规则表+原型齐备后交接",
        "owner": "产品·海外",
        "note": "整柜 DRAYEASY 预报接口另跟袁经理",
    },
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "船务组",
        "pkg": "滞箱/滞港费逻辑 + 免用箱登记",
        "ids": "CW007,CW005",
        "pri": "P0",
        "deliver": "费用规则原型 + 供应商多维筛选表",
        "status": "有方案",
        "gate": "与提单费用登记/邮件水单合并能力对齐",
        "artifacts": "0730 费用节点方案；提单管理费用登记",
        "dev_ready": "规则评审后交接",
        "owner": "产品·船务",
        "note": "国内外免用箱；尾端异常费用表格沉淀",
    },
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "船务组",
        "pkg": "运价上系统时机 + 配舱回写",
        "ids": "CW001",
        "pri": "P0",
        "deliver": "时序图 + 状态机说明补进方案/原型",
        "status": "有方案",
        "gate": "依赖进港时间口径（W1）",
        "artifacts": "0730 方案；运价管理 MVP",
        "dev_ready": "与运价管理同包交接更佳",
        "owner": "产品·船务",
        "note": "费用节点与进港预警联动",
    },
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "关务组",
        "pkg": "AN/费用单上传 + 报关查运单性能",
        "ids": "GW025,GW022",
        "pri": "P1",
        "deliver": "上传交互原型；性能需求说明（非 UI）",
        "status": "方案中",
        "gate": "上传格式/必填；慢查询给研发性能指标",
        "artifacts": "报关管理列表/资料准备",
        "dev_ready": "上传可进设计；性能单独立项",
        "owner": "产品·关务",
        "note": "截单预警触达关务待办需确认是否已有通道",
    },
    {
        "wave": "W2 方案转原型",
        "week": "08/24–08/28",
        "group": "海外对接组",
        "pkg": "时效表（提柜/派送/FBA）+ 轨迹结果字段",
        "ids": "（时效主线）",
        "pri": "P1/P2",
        "deliver": "MVP 定稿：周统计 + 西雅图/自提例外 + 扫描超7天→刘奎林",
        "status": "有原型",
        "gate": "轨迹公式落「结果字段」；导出 5 万与 PT002 联动",
        "artifacts": "海外对接组-时效表-提柜私卡快递FBA-MVP.html",
        "dev_ready": "规则表齐备可交接",
        "owner": "产品·海外",
        "note": "亚马逊时效=签收−提柜",
    },
    # —— W3：待设计 / 体验 ——
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "报价组",
        "pkg": "签入无业务员成本提醒/卡控",
        "ids": "BJ005",
        "pri": "P1",
        "deliver": "交互设计（提醒 vs 禁止出货策略）",
        "status": "待设计",
        "gate": "策略拍板：仅提醒 or 硬卡",
        "artifacts": "—",
        "dev_ready": "设计评审后",
        "owner": "产品·报价",
        "note": "货物签入后无费用场景",
    },
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "报价组",
        "pkg": "沃尔玛仓费用/下单/尾端流转",
        "ids": "BJ006",
        "pri": "P2",
        "deliver": "流程设计 + 低保真",
        "status": "待设计",
        "gate": "与尾端预估字段对齐",
        "artifacts": "—",
        "dev_ready": "设计后",
        "owner": "产品·报价",
        "note": "体验/沉淀类",
    },
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "报价组",
        "pkg": "整柜/散货询价导出",
        "ids": "BJ007",
        "pri": "P2",
        "deliver": "导出字段清单设计",
        "status": "待设计",
        "gate": "字段与权限",
        "artifacts": "—",
        "dev_ready": "字段表即可开发",
        "owner": "产品·报价",
        "note": "轻量交付",
    },
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "跨组/平台",
        "pkg": "导出 5 万条能力",
        "ids": "PT002",
        "pri": "P2",
        "deliver": "平台能力需求说明（异步导出/权限）",
        "status": "待设计",
        "gate": "与查货率/时效导出共用",
        "artifacts": "—",
        "dev_ready": "说明齐即可排平台",
        "owner": "产品·平台",
        "note": "支撑 FK001 / 时效导出",
    },
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "关务组",
        "pkg": "查验前后端同步（改港退差价等）",
        "ids": "GW020",
        "pri": "P0",
        "deliver": "跨角色同步原型（前端可见后端查验结果）",
        "status": "有方案",
        "gate": "与查验报表/清关通知串联",
        "artifacts": "关务 0729 查验方案；查验报表",
        "dev_ready": "原型后交接",
        "owner": "产品·关务",
        "note": "案例 GW028/031 作验收样例",
    },
    {
        "wave": "W3 待设计与体验",
        "week": "08/31–09/04",
        "group": "海外对接组",
        "pkg": "整柜还柜轨迹对客服开放",
        "ids": "HW015",
        "pri": "P1",
        "deliver": "权限与展示原型",
        "status": "方案中",
        "gate": "谁可见、何时隐藏规则",
        "artifacts": "主单跟进-MVP",
        "dev_ready": "小改动可快交",
        "owner": "产品·海外",
        "note": "勿自动隐藏",
    },
]

HEADERS = [
    "波次",
    "周次",
    "组别",
    "交付包",
    "需求ID",
    "优先级",
    "产品交付物",
    "当前状态",
    "准入门槛/待拍板",
    "关联原型·方案·纪要",
    "对开发就绪度",
    "产品负责人域",
    "备注",
]

WAVE_COLORS = {
    "W0 本周可交接": "CF1322",
    "W1 补原型与口径": "D46B08",
    "W2 方案转原型": "1890FF",
    "W3 待设计与体验": "389E0D",
}


def write_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "交付计划"
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="001529")
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = thin

    for r, p in enumerate(PACKAGES, 2):
        vals = [
            p["wave"],
            p["week"],
            p["group"],
            p["pkg"],
            p["ids"],
            p["pri"],
            p["deliver"],
            p["status"],
            p["gate"],
            p["artifacts"],
            p["dev_ready"],
            p["owner"],
            p["note"],
        ]
        color = WAVE_COLORS.get(p["wave"], "666666")
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin
            if c == 1:
                cell.font = Font(bold=True, color=color)

    widths = [16, 12, 10, 28, 18, 8, 28, 14, 32, 36, 22, 12, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{len(PACKAGES) + 1}"

    # 总览
    ov = wb.create_sheet("总览说明", 0)
    lines = [
        ["产品部门 · 对开发的原型与设计交付计划表"],
        ["生成日", "2026-08-10"],
        ["用途", "产品侧按周输出「可给开发对接」的原型/PRD/设计；不是研发开发排期"],
        ["来源", "产品部门-各组需求清单-评审版（102条）+ 调研纪要 0727–0807（含 8/6–8/7 原型评审收口）"],
        ["统计口径", "有原型57 / 有方案34 / 方案中5 / 待设计4 / 验收中1 / 待原型1；P0=41 P1=46 P2=15"],
        [],
        ["波次定义"],
        ["W0 本周可交接", "原型已齐或仅需小回改/复审；本周完成定稿即可进开发对接"],
        ["W1 补原型与口径", "缺口径或待原型；本周内补齐后方可进入开发队列"],
        ["W2 方案转原型", "已有方案文字，需转成可评审 HTML/交互后再交接"],
        ["W3 待设计与体验", "待设计或 P2 体验项；设计评审后轻量交接"],
        [],
        ["本周（W0）产品必做"],
        ["1", "风控查货按 8/6 回改 → 约郭佳玉复审"],
        ["2", "查验报表：功能栏备注落点（登记人=跟进人）"],
        ["3", "提单管理 UI 批注回改（去分组/装完色标/发送标记一体）"],
        ["4", "周一对照私卡：合约价旁「确认」去留"],
        ["5", "指定人启动「海外工单」端到端流程梳理"],
        ["6", "梳清进港时间口径（抓取源、用途、异常）"],
        [],
        ["并行待拍（阻塞开发前必须有结论）"],
        ["A", "查验费是否进提单预估"],
        ["B", "信号旗能否抓清关 IT / 保税号"],
        ["C", "截单预警是否已触达关务待办"],
        ["D", "进港时间口径"],
        [],
        ["交接开发最低交付包"],
        ["①", "可点击原型（HTML）或等效交互稿"],
        ["②", "一页 PRD：范围/非范围/字段/状态机/权限"],
        ["③", "验收样例（最好挂真实案例号）"],
        ["④", "待拍板清单（带默认建议，避免阻塞）"],
        [],
        ["关联入口"],
        ["需求清单", "产品部门-各组需求清单-评审版.html"],
        ["纪要收口", "调研纪要梳理_2026-08-06至08-07_卢慧恒_原型评审.html"],
        ["导航", "产品部门-导航.html"],
    ]
    ov["A1"].font = Font(bold=True, size=14, color="001529")
    for r, row in enumerate(lines, 1):
        for c, v in enumerate(row, 1):
            cell = ov.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if r == 1:
                cell.font = Font(bold=True, size=14, color="001529")
            elif len(row) == 1 and v and not str(v)[0].isdigit():
                cell.font = Font(bold=True, size=12, color="003A8C")
    ov.column_dimensions["A"].width = 28
    ov.column_dimensions["B"].width = 88

    # 按组摘要
    gws = wb.create_sheet("按组摘要")
    gh = ["组别", "W0可交接包数", "W1补齐", "W2转原型", "W3设计", "关键建议"]
    for c, h in enumerate(gh, 1):
        cell = gws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="001529")
        cell.font = Font(color="FFFFFF", bold=True)
    advice = {
        "风控组": "复审通过后整包交接查货；报表可并行字段对接",
        "关务组": "先切片分单号+双通道+改配；查验备注本周可交；案例沉淀模板",
        "报价组": "AI报价+看板可交；POD验收闭环；LTL/待设计进 W1–W3",
        "海外对接组": "预估+尾单本周可交；邮箱/HOLD/LFD 转原型；工单先梳流程",
        "船务组": "提单+运价本周可交；进港口径与滞箱滞港进 W1–W2",
        "跨组": "工单流程+导出5万平台能力；全流程评审作基线",
        "跨组/平台": "导出5万与时效/查货率共用；全流程作评审基线",
        "风控/关务": "信号旗可行性先行",
    }
    from collections import Counter, defaultdict

    by = defaultdict(Counter)
    for p in PACKAGES:
        g = p["group"]
        if g.startswith("W"):
            continue
        key = p["wave"][:2]
        by[g][key] += 1
    # normalize group names for summary
    groups_order = ["风控组", "关务组", "报价组", "海外对接组", "船务组", "跨组", "跨组/平台", "风控/关务"]
    seen = set()
    row = 2
    for g in groups_order:
        if g not in by:
            continue
        seen.add(g)
        c = by[g]
        gws.cell(row, 1, g)
        gws.cell(row, 2, c.get("W0", 0))
        gws.cell(row, 3, c.get("W1", 0))
        gws.cell(row, 4, c.get("W2", 0))
        gws.cell(row, 5, c.get("W3", 0))
        gws.cell(row, 6, advice.get(g, ""))
        row += 1
    for g, c in by.items():
        if g in seen:
            continue
        gws.cell(row, 1, g)
        gws.cell(row, 2, c.get("W0", 0))
        gws.cell(row, 3, c.get("W1", 0))
        gws.cell(row, 4, c.get("W2", 0))
        gws.cell(row, 5, c.get("W3", 0))
        gws.cell(row, 6, advice.get(g, ""))
        row += 1
    for i, w in enumerate([12, 14, 10, 12, 10, 56], 1):
        gws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT_XLSX)
    print("wrote", OUT_XLSX)


def esc(s):
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def write_html():
    rows_html = []
    cur_wave = None
    for p in PACKAGES:
        if p["wave"] != cur_wave:
            cur_wave = p["wave"]
            rows_html.append(
                f'<tr class="wave-hd"><td colspan="8"><span class="wtag">{esc(cur_wave)}</span>'
                f'　{esc(p["week"])}</td></tr>'
            )
        rows_html.append(
            f"""<tr>
  <td><b>{esc(p['group'])}</b><div class="sub">{esc(p['pri'])} · {esc(p['ids'])}</div></td>
  <td><b>{esc(p['pkg'])}</b><div class="sub">{esc(p['status'])}</div></td>
  <td>{esc(p['deliver'])}</td>
  <td>{esc(p['gate'])}</td>
  <td class="sm">{esc(p['artifacts'])}</td>
  <td>{esc(p['dev_ready'])}</td>
  <td>{esc(p['owner'])}</td>
  <td class="sm">{esc(p['note'])}</td>
</tr>"""
        )

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>产品部门 · 对开发 · 原型与设计计划表</title>
  <style>
    :root {{
      --bg:#f0f2f5; --card:#fff; --line:#e8e8e8; --text:rgba(0,0,0,.85); --sub:rgba(0,0,0,.45);
      --w0:#cf1322; --w1:#d46b08; --w2:#1890ff; --w3:#389e0d;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Microsoft YaHei","PingFang SC",Arial,sans-serif; background:var(--bg); color:var(--text); font-size:13px; }}
    a {{ color:#1890ff; text-decoration:none; }}
    a:hover {{ text-decoration:underline; }}
    .top {{
      position:sticky; top:0; z-index:10; background:#001529; color:#fff;
      padding:12px 16px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;
    }}
    .top h1 {{ margin:0; font-size:16px; font-weight:600; }}
    .top .meta {{ color:rgba(255,255,255,.7); font-size:12px; }}
    .top .acts {{ margin-left:auto; display:flex; gap:8px; flex-wrap:wrap; }}
    .top .acts a {{
      color:#fff; background:rgba(255,255,255,.12); padding:4px 10px; border-radius:2px; font-size:12px;
    }}
    .wrap {{ max-width:1280px; margin:0 auto; padding:14px 16px 48px; }}
    .banner {{
      background:#fff; border:1px solid var(--line); border-radius:6px; padding:12px 14px; margin-bottom:12px; line-height:1.7;
    }}
    .banner b {{ color:#003a8c; }}
    .kpis {{ display:grid; grid-template-columns:repeat(4,1fr); gap:8px; margin-bottom:12px; }}
    @media (max-width:900px) {{ .kpis {{ grid-template-columns:1fr 1fr; }} }}
    .kpi {{
      background:#fff; border:1px solid var(--line); border-radius:6px; padding:10px 12px;
      border-top:3px solid #d9d9d9;
    }}
    .kpi.w0 {{ border-top-color:var(--w0); }}
    .kpi.w1 {{ border-top-color:var(--w1); }}
    .kpi.w2 {{ border-top-color:var(--w2); }}
    .kpi.w3 {{ border-top-color:var(--w3); }}
    .kpi b {{ display:block; font-size:14px; margin-bottom:4px; }}
    .kpi span {{ font-size:12px; color:var(--sub); line-height:1.5; }}
    .todo {{
      background:#fffbe6; border:1px solid #ffe58f; border-radius:6px; padding:10px 12px; margin-bottom:12px;
    }}
    .todo h3 {{ margin:0 0 6px; font-size:13px; color:#614700; }}
    .todo ol {{ margin:0; padding-left:18px; line-height:1.7; font-size:12px; }}
    table {{
      width:100%; border-collapse:collapse; background:#fff; border:1px solid var(--line); border-radius:6px; overflow:hidden;
    }}
    th {{
      background:#fafafa; text-align:left; padding:8px 10px; font-size:12px; color:var(--sub); font-weight:600;
      border-bottom:1px solid var(--line); position:sticky; top:48px; z-index:5;
    }}
    td {{
      padding:8px 10px; border-bottom:1px solid #f0f0f0; vertical-align:top; line-height:1.55; font-size:12px;
    }}
    tr.wave-hd td {{
      background:#001529; color:#fff; font-weight:600; padding:8px 10px; border-bottom:none;
    }}
    .wtag {{
      display:inline-block; padding:1px 8px; border-radius:2px; background:rgba(255,255,255,.15); font-size:12px;
    }}
    .sub {{ color:var(--sub); font-size:11px; margin-top:2px; }}
    .sm {{ font-size:11px; color:#334155; }}
    .foot {{ margin-top:12px; font-size:11px; color:var(--sub); text-align:center; }}
  </style>
</head>
<body>
  <div class="top">
    <h1>产品 → 开发 · 原型与设计交付计划表</h1>
    <span class="meta">2026-08-10 起 · 来源：需求清单102条 + 0727–0807纪要</span>
    <div class="acts">
      <a href="产品部门-对开发-原型与设计计划表.xlsx">下载 Excel</a>
      <a href="产品部门-各组需求清单-评审版.html">需求清单</a>
      <a href="调研纪要梳理_2026-08-06至08-07_卢慧恒_原型评审.html">8/6–8/7收口</a>
      <a href="产品部门-导航.html">返回导航</a>
    </div>
  </div>
  <div class="wrap">
    <div class="banner">
      <b>这份表是什么：</b>产品侧按周把「可给开发对接」的原型 / PRD / 交互设计排出来（不是研发排期）。
      交接最低包：①可点原型 ②一页PRD（范围/字段/状态/权限）③验收样例 ④待拍板清单（带默认建议）。
      清单现状：有原型57 · 有方案34 · 方案中5 · 待设计4 · 验收中1 · 待原型1；P0=41。
    </div>
    <div class="kpis">
      <div class="kpi w0"><b>W0 · 本周可交接</b><span>风控查货复审、提单/运价、海外预估与尾单、关务报关切片、AI报价/看板、查验备注</span></div>
      <div class="kpi w1"><b>W1 · 补原型与口径</b><span>海外工单流程、进港时间、HW001柜维比对、信号旗可行性、LTL轨迹、清关轨迹文案</span></div>
      <div class="kpi w2"><b>W2 · 方案转原型</b><span>邮箱工作台、HOLD/LFD、滞箱滞港、运价上时机、AN上传、时效例外规则</span></div>
      <div class="kpi w3"><b>W3 · 待设计</b><span>签入成本卡控、沃尔玛流转、询价导出、5万导出平台、查验前后端同步</span></div>
    </div>
    <div class="todo">
      <h3>本周产品必做（对齐 8/7 17:30 收口）</h3>
      <ol>
        <li>风控查货按 8/6 回改 → 约郭佳玉复审</li>
        <li>查验报表：功能栏备注；登记人=跟进人</li>
        <li>提单管理 UI 批注回改（去分组 / 装完色标仅两色 / AMS·ISF 一体）</li>
        <li>周一对照私卡页：合约价旁「确认」去留</li>
        <li>指定人启动「海外工单」端到端流程梳理（梳完再对原型）</li>
        <li>梳清进港时间口径（抓取源、用途、异常）</li>
      </ol>
    </div>
    <table>
      <thead>
        <tr>
          <th style="width:12%">组别 / ID</th>
          <th style="width:14%">交付包</th>
          <th style="width:14%">产品交付物</th>
          <th style="width:16%">准入门槛</th>
          <th style="width:18%">原型·方案·纪要</th>
          <th style="width:12%">对开发就绪度</th>
          <th style="width:8%">负责域</th>
          <th style="width:6%">备注</th>
        </tr>
      </thead>
      <tbody>
        {''.join(rows_html)}
      </tbody>
    </table>
    <p class="foot">随复审结论滚动更新 · 详细需求条目见评审版清单 · Excel 含总览与按组摘要</p>
  </div>
</body>
</html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")
    print("wrote", OUT_HTML)


if __name__ == "__main__":
    write_xlsx()
    write_html()
