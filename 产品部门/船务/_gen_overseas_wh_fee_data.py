# -*- coding: utf-8 -*-
"""从 海外仓2026-06.xlsx + 计费规则.xlsx 导出海外仓预估费用原型数据。"""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
WH_XLSX = BASE / "海外仓2026-06.xlsx"
RULE_XLSX = BASE / "计费规则.xlsx"
OUT_JS = BASE / "海外仓预估费用-data.js"
OUT_JSON = BASE / "海外仓预估费用-data.json"


def cell(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    s = str(v).strip()
    return s if s else None


def num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return int(v) if isinstance(v, float) and v == int(v) else float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "无", "—", "-", "×"):
        return None
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except ValueError:
        return None


def parse_combo_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"providers": [], "warehouses": []}
    header = [cell(c) for c in rows[0]]
    providers = []
    for h in header[1:]:
        if not h:
            if providers:
                break
            continue
        if h == "BW" and providers:
            break
        providers.append(h)
    warehouses = []
    for row in rows[1:]:
        code = cell(row[0])
        if not code or code in ("仓库地址", "仓系地址"):
            continue
        prices = {}
        for j, p in enumerate(providers, 1):
            if j >= len(row):
                break
            prices[p] = num(row[j])
        if any(prices.get(p) is not None for p in providers):
            warehouses.append({"code": code, "prices": prices})
    return {"providers": providers, "warehouses": warehouses}


def parse_pallet_sheet(ws, max_rows=120):
    rows = list(ws.iter_rows(values_only=True))
    header = [cell(c) for c in rows[0]]
    providers = [h for h in header[1:] if h]
    # 找 SYCO 计费方式列截断
    if "SYCO计费方式" in providers:
        idx = providers.index("SYCO计费方式")
        providers = providers[:idx]
    warehouses = []
    for row in rows[1:max_rows + 1]:
        code = cell(row[0])
        if not code:
            continue
        prices = {}
        for j, p in enumerate(providers, 1):
            if j >= len(row):
                break
            prices[p] = num(row[j])
        if any(prices.get(p) is not None for p in providers):
            warehouses.append({"code": code, "prices": prices})
    return {"providers": providers, "warehouses": warehouses}


def parse_amz_direct(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = [cell(c) for c in rows[0]]
    providers = [h for h in header[1:] if h]
    warehouses = []
    for row in rows[1:]:
        code = cell(row[0])
        if not code:
            continue
        prices = {providers[j - 1]: num(row[j]) for j in range(1, len(providers) + 1) if j < len(row)}
        if any(v is not None for v in prices.values()):
            warehouses.append({"code": code, "prices": prices})
    return {"providers": providers, "warehouses": warehouses}


def parse_full_us_rules(ws, max_rows=30):
    rows = list(ws.iter_rows(values_only=True))
    items = []
    for row in rows[1 : max_rows + 1]:
        mode = cell(row[0])
        addr = cell(row[1])
        provider = cell(row[2])
        if not mode and not provider:
            continue
        items.append(
            {
                "mode": mode,
                "address": addr,
                "provider": provider,
                "fcl40In": num(row[3]) if len(row) > 3 else None,
                "fcl40Out": num(row[4]) if len(row) > 4 else None,
                "fcl45In": num(row[5]) if len(row) > 5 else None,
                "fcl45Out": num(row[6]) if len(row) > 6 else None,
                "floorFee": cell(row[7]) if len(row) > 7 else None,
                "supervisorFee": cell(row[8]) if len(row) > 8 else None,
                "storageIn": cell(row[9]) if len(row) > 9 else None,
                "storageOut": cell(row[10]) if len(row) > 10 else None,
                "upsOut": cell(row[11]) if len(row) > 11 else None,
                "fedexOut": cell(row[13]) if len(row) > 13 else None,
                "notes": cell(row[14]) if len(row) > 14 else None,
            }
        )
    return items


def parse_rule_text(ws, max_rows=25):
    lines = []
    for row in ws.iter_rows(max_row=max_rows, values_only=True):
        for c in row:
            if c and str(c).strip():
                lines.append(str(c).strip())
    return lines


def parse_rule_detail_sheet(ws, max_rows=20):
    rows = list(ws.iter_rows(values_only=True))
    header_row = 0
    for i, row in enumerate(rows[:5]):
        if row and cell(row[0]) in ("核算方式", "组合", "散板", "全包"):
            header_row = i
            break
        if row and cell(row[0]) == "核算方式":
            header_row = i
            break
    if header_row == 0 and len(rows) > 1 and cell(rows[1][0]) == "核算方式":
        header_row = 1
    header = [cell(c) for c in rows[header_row]]
    items = []
    for row in rows[header_row + 1 : header_row + 1 + max_rows]:
        if not any(row):
            continue
        item = {}
        for i, h in enumerate(header):
            if not h or i >= len(row):
                continue
            v = row[i]
            if v is None:
                continue
            item[str(h).strip()] = num(v) if isinstance(v, (int, float)) else cell(v)
        if item.get("核算方式") or item.get("代理") or item.get("系统名字"):
            items.append(item)
    return items


def parse_private_direct(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [cell(c) for c in rows[0]]
    items = []
    for row in rows[1:]:
        if not cell(row[0]):
            continue
        item = {}
        for i, h in enumerate(header):
            if not h or i >= len(row):
                continue
            v = row[i]
            if v is None:
                continue
            item[str(h).strip()] = num(v) if isinstance(v, (int, float)) else cell(v)
        items.append(item)
    return items


def main():
    data = {
        "source": ["海外仓2026-06.xlsx", "计费规则.xlsx"],
        "updated": "2026-06",
        "calcRules": [],
        "comboPrice": {},
        "palletPrice": {},
        "amzDirect": {},
        "fullUsSample": [],
        "providerDetailSample": [],
        "privateDirectSample": [],
    }

    # 计费规则
    wb_r = load_workbook(RULE_XLSX, data_only=True)
    data["calcRules"] = parse_rule_text(wb_r[wb_r.sheetnames[0]])

    r_names = wb_r.sheetnames
    if len(r_names) > 1:
        data["providerDetailSample"] = parse_rule_detail_sheet(wb_r[r_names[1]], 15)
    if len(r_names) > 5:
        data["privateDirectSample"] = parse_private_direct(wb_r[r_names[5]])

    wb_r.close()

    # 海外仓价格表
    wb = load_workbook(WH_XLSX, data_only=True)
    names = wb.sheetnames
    if len(names) > 0:
        data["comboPrice"] = parse_combo_sheet(wb[names[0]])
    if len(names) > 1:
        data["palletPrice"] = parse_pallet_sheet(wb[names[1]])
    if len(names) > 2:
        data["amzDirect"] = parse_amz_direct(wb[names[2]])
    if len(names) > 3:
        data["fullUsSample"] = parse_full_us_rules(wb[names[3]], 15)
    wb.close()

    OUT_JSON.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    js = "window.OVERSEAS_WH_FEE = " + json.dumps(data, ensure_ascii=False, separators=(",", ":")) + ";"
    OUT_JS.write_text(js, encoding="utf-8")
    print("combo warehouses:", len(data["comboPrice"].get("warehouses", [])))
    print("pallet warehouses:", len(data["palletPrice"].get("warehouses", [])))
    print("amz warehouses:", len(data["amzDirect"].get("warehouses", [])))
    print("calc rules:", len(data["calcRules"]))
    print("written:", OUT_JS.name)


if __name__ == "__main__":
    main()
