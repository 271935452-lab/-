# -*- coding: utf-8 -*-
"""生成产品部门各组需求清单 · Excel + HTML（评审用，含优先级与原型跳转）"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
JSON_PATH = ROOT / "_req_extract_7.31.json"
XLSX_OUT = ROOT / "产品部门-各组需求清单-评审版.xlsx"
HTML_OUT = ROOT / "产品部门-各组需求清单-评审版.html"

# 相对 HTML 的链接（HTML 在 产品部门/ 下）——产出原型/方案须登记于此并在 enrich/overlay 挂到条目
LINKS = {
    "风控方案": "各组sop/风控组-0729-查货管理-大致方案.html",
    "查货操作AI识图": "各组sop/风控组-查货操作-AI识图品名识别-原型.html",
    "查货率报表": "各组sop/风控组-操作查货率报表-MVP.html",
    "关务方案汇总": "各组sop/关务组-部门问题方案汇总-流程与原型.html",
    "关务查验": "各组sop/关务组-0729-查验分柜装授权-大致方案.html",
    "分单号": "各组sop/关务组-分单号规则表.html",
    "分单号自动生成": "../报关行对接/报关管理-列表-MVP.html?tab=push",
    "分单号规则配置": "../报关行对接/分单号规则配置-MVP.html",
    "做资料分票核准": "../报关行对接/报关管理-做资料分票核准-MVP.html",
    "资料核准": "各组sop/关务组-资料核准双通道-流程与页面设计.html",
    "改配": "各组sop/关务组-改配三类场景-轨迹与同步.html",
    "三表": "各组sop/关务组-三表-仓位装柜跟踪-流程与页面设计.html",
    "流程图汇总": "各组sop/关务组-流程图汇总.html",
    "做资料全流程": "各组sop/关务组-做资料全流程图.html",
    "船务方案": "各组sop/船务组-0730-费用节点提单配仓-大致方案.html",
    "私卡0731": "各组sop/私卡报价-0731-AI报价有效期运单绑定-大致方案.html",
    "AI报价规则配置": "各组sop/私卡报价-AI报价规则配置-MVP.html",
    "报价员绑运单": "../二期/ESS询价报价-报价员列表-MVP.html",
    "报价看板": "报价时效统计看板-MVP.html",
    "海外方案": "各组sop/海外对接组-0804-港前港后整柜价格-大致方案.html",
    "尾单原型": "各组sop/海外对接组-港后-尾单单票跟踪-原型.html",
    "协议价原型": "各组sop/海外对接组-客户协议价格维护-原型.html",
    "全流程": "各组sop/产品部门-标准化全流程-评审版.html",
    "系统流程设计": "各组sop/产品部门-各组系统流程设计.html",
    "价值流图": "各组sop/产品部门-价值流图.html",
    "整柜绑报价": "../整柜下单/整柜询价与报价绑定-MVP.html",
    "进口商": "../进口商管理-MVP.html",
    "清关管理": "../清关管理-MVP.html",
    "清关行港口配置": "../进口商-清关行港口配置-MVP.html",
    "柜子绑进口商": "../柜子编辑-绑定进口商-MVP.html",
    "报关列表": "../报关行对接/报关管理-列表-MVP.html",
    "报关资料准备": "../报关行对接/报关管理-资料准备-列表-MVP.html",
    "报关流程图": "../报关行对接/报关管理-流程图-MVP.html",
    "提单管理原型": "各组sop/船务组-提单管理-按周分组与分配-MVP.html",
    "运价管理原型": "各组sop/船务组-运价管理-我司与代理订舱-MVP.html",
    "查验报表原型": "各组sop/关务组-查验报表-提单运单维度-MVP.html",
    "出库签收预警": "../B21-出库签收预警-MVP.html",
    "POD下载": "../二期/订单管理-POD下载-MVP.html",
    "POD说明": "客户API-POD对接调整说明.md",
    "纪要0731": "调研纪要_2026-07-31.txt",
    "纪要0804": "调研纪要_2026-08-04_海外对接组_全天整合.txt",
    "纪要0730": "调研纪要_2026-07-30_上午下午.txt",
    "纪要0729": "调研纪要_2026-07-29_上午下午.txt",
    "导航": "产品部门-导航.html",
}


def clean(s: str) -> str:
    s = (s or "").replace("\xa0", " ").strip()
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s


def first_line(s: str, n: int = 80) -> str:
    s = clean(s).split("\n")[0].strip()
    return s if len(s) <= n else s[: n - 1] + "…"


def proto(*keys: str) -> str:
    parts = []
    for k in keys:
        if k in LINKS:
            parts.append(f"{k}|{LINKS[k]}")
    return "；".join(parts)


def load_sheet_rows(data: dict, sheet: str) -> list[list[str]]:
    rows = data.get(sheet, [])
    if not rows:
        return []
    # skip header
    return [r for r in rows[1:] if any(clean(c) for c in r)]


def from_problem_rows(group: str, rows: list[list[str]], source: str, default_pri: str = "P1") -> list[dict]:
    """Convert problem-summary rows into requirement dicts."""
    stub_titles = {"报价组", "船务组", "关务组", "海外对接组", "风控组", "跨组", "跨组/平台", "平台"}
    out = []
    for i, r in enumerate(rows, 1):
        title = clean(r[0] if len(r) > 0 else "")
        detail = clean(r[1] if len(r) > 1 else "")
        plan = clean(r[2] if len(r) > 2 else "")
        extra = clean(r[3] if len(r) > 3 else "")
        if not title and not detail:
            continue
        # 汇总表里的分组空行（仅写了「报价组」等）不当作需求
        if title in stub_titles and len("\n".join(x for x in [detail, plan, extra] if x)) < 8:
            continue
        if not title:
            title = first_line(detail, 60) or f"条目{i}"
        body = "\n".join(x for x in [detail, plan, extra] if x)
        out.append(
            {
                "group": group,
                "title": first_line(title, 100),
                "desc": body[:800],
                "source": source,
                "priority": default_pri,
                "status": "待设计",
                "proto": "",
                "note": "",
            }
        )
    return out


def enrich(items: list[dict]) -> list[dict]:
    """Assign priority / prototype / status by keyword heuristics + known research."""
    rules = [
        # (group_substr or None, title/desc keywords, pri, status, proto_keys, note)
        # —— 风控：查货操作 AI 原型优先挂 ——
        ("风控", ["智能体", "腾小信", "识图", "品名识别", "反倾销", "两用物项"], "P0", "有原型", ["查货操作AI识图", "风控方案", "纪要0729"], "AI识图/品名识别落查货操作"),
        ("风控", ["侵权", "授权", "知识产权", "品牌"], "P0", "有原型", ["查货操作AI识图", "风控方案", "关务查验"], "与关务授权/保函联动"),
        ("风控", ["附加费", "查货操作", "直客", "应收", "产品确认"], "P0", "有原型", ["查货操作AI识图", "风控方案", "纪要0729"], "查货勾选附加费+成本/应收分流"),
        ("风控", ["超品名"], "P1", "有原型", ["查货操作AI识图", "风控方案"], ""),
        ("风控", ["查货率", "100%"], "P1", "有原型", ["查货率报表", "查货操作AI识图", "风控方案"], "总票/空白/保函/漏查·导出与播报"),
        ("风控", ["扣件", "罚款", "500"], "P1", "有方案", ["查货操作AI识图", "风控方案"], ""),
        ("风控", [], "P1", "有方案", ["风控方案", "查货操作AI识图"], "默认挂风控方案+查货操作原型"),
        # —— 报价 ——
        ("报价", ["时效看板", "时长看板", "人工报价"], "P0", "有原型", ["报价看板", "纪要0731"], "月度考核≥90%"),
        ("报价", ["AI", "有效期", "内外州"], "P0", "有原型", ["AI报价规则配置", "报价员绑运单", "私卡0731", "纪要0731"], "新AI·本州外州配置"),
        ("报价", ["尾端派送", "尾端成本"], "P0", "有原型", ["报价员绑运单", "私卡0731", "AI报价规则配置"], "确认台登记·绑单/签入挂费"),
        ("报价", ["绑运单", "绑定运单", "运单绑定"], "P0", "有原型", ["报价员绑运单", "整柜绑报价", "私卡0731"], "散货多绑+尾端费"),
        ("报价", ["轨迹", "LTL", "抓取"], "P0", "方案中", ["私卡0731", "POD说明", "出库签收预警"], "LTTL轨迹周期内完成"),
        ("报价", ["出库", "签收", "pod", "POD", "预警"], "P0", "验收中", ["私卡0731", "POD说明", "出库签收预警", "POD下载"], "出库预警验收"),
        ("报价", ["业务员成本", "没有费用"], "P1", "待设计", ["私卡0731"], "签入无成本拦截"),
        ("报价", ["沃尔玛"], "P2", "待设计", ["私卡0731", "协议价原型"], "尾端价格衔接"),
        ("报价", ["询价数据导出"], "P2", "待设计", ["报价员绑运单"], ""),
        ("报价", [], "P1", "有方案", ["私卡0731"], "默认挂私卡0731"),
        # —— 关务 ——
        ("关务", ["分单号"], "P0", "有原型", ["分单号自动生成", "分单号规则配置", "分单号", "关务方案汇总", "报关列表"], "船司×港口规则配置+自动生成；货物调整重算；#01/#03 推送同步"),
        ("关务", ["收发通", "进口商", "清关行", "港口"], "P0", "有原型", ["进口商", "清关行港口配置", "柜子绑进口商", "三表"], "收件人/发件人+邮箱密码+清关行港口"),
        ("关务", ["取消IT", "取消 IT"], "P0", "有原型", ["清关管理", "提单管理原型", "船务方案"], "清关侧取消IT标记；订阅船务提醒"),
        ("关务", ["一键通知", "通知客户", "清关查验", "报关查验"], "P0", "有原型", ["清关管理", "报关列表", "关务查验"], "清关/报关查验页签均可一键通知客户"),
        ("关务", ["合并报关", "分柜装", "再分票"], "P0", "有原型", ["关务查验", "报关列表", "报关资料准备", "关务方案汇总"], "合并标记+客户合并后再分票（例10→3）"),
        ("关务", ["改运抵", "运抵"], "P0", "有方案", ["改配", "船务方案"], ""),
        ("关务", ["委托"], "P1", "有原型", ["报关列表", "报关资料准备", "关务方案汇总"], "对接#12按客户名称+报关行查询回写状态；未对接委托标记+截图；无委托不可报关"),
        ("关务", ["件毛体", "报关资料", "核对", "资料核准", "买单"], "P0", "有原型", ["做资料分票核准", "资料核准", "关务方案汇总", "做资料全流程"], "报关件取预录；买单=合计−报关件"),
        ("关务", ["磅单", "超重", "限重"], "P1", "有方案", ["做资料分票核准", "关务方案汇总", "做资料全流程"], ""),
        ("关务", ["ISF", "AMS", "退关"], "P1", "有方案", ["改配", "船务方案"], ""),
        ("关务", ["查验", "缉私", "两用"], "P0", "有方案", ["关务查验", "报关列表", "报关流程图"], ""),
        ("关务", ["AN文件", "买单"], "P1", "方案中", ["海外方案", "报关列表", "报关资料准备"], "与海外邮箱资料衔接"),
        ("关务", ["报关管理", "运单号", "清关放行", "轨迹"], "P1", "方案中", ["报关列表", "报关流程图", "报关资料准备"], ""),
        ("关务", ["体积", "箱型", "提单号", "舱单"], "P1", "有方案", ["关务方案汇总", "做资料全流程", "三表"], ""),
        ("关务", ["配送方式", "业务数据统计"], "P1", "有方案", ["关务方案汇总", "全流程"], ""),
        ("关务", ["自报关", "全流程组别"], "P1", "有方案", ["流程图汇总", "做资料全流程", "关务方案汇总"], ""),
        ("关务", [], "P1", "有方案", ["关务方案汇总", "流程图汇总"], "默认挂关务方案汇总"),
        # —— 海外 ——
        ("海外", ["私仓价格", "尾端", "预估", "费用预估"], "P0", "待原型", ["海外方案", "协议价原型"], "柜子维度预估比对，非尾单页"),
        ("海外", ["四桶", "没约", "尾单", "派送时效", "拆柜"], "P0", "有原型", ["尾单原型", "纪要0804"], ""),
        ("海外", ["协议价", "价格表", "批量"], "P0", "有原型", ["协议价原型", "海外方案"], ""),
        ("海外", ["大客户", "调成本"], "P0", "有方案", ["协议价原型", "海外方案", "纪要0804"], ""),
        ("海外", ["HOLD", "提柜预警", "DO", "PU"], "P0", "有方案", ["海外方案", "纪要0804"], "港前准备台"),
        ("海外", ["邮箱", "滞港", "滞箱", "AN"], "P0", "有方案", ["海外方案", "纪要0804"], "103邮箱工作台"),
        ("海外", ["包查验"], "P1", "有方案", ["海外方案", "关务查验"], ""),
        ("海外", ["DRAYEASY", "预报"], "P1", "有方案", ["海外方案"], "问袁经理API"),
        ("海外", ["特殊预约", "指令", "留言", "异常"], "P1", "有原型", ["尾单原型"], "异常/备注线上化"),
        ("海外", ["轨迹", "机器人"], "P1", "方案中", ["尾单原型", "POD说明", "出库签收预警"], ""),
        ("海外", ["LFD"], "P0", "有方案", ["海外方案", "船务方案"], ""),
        ("海外", ["卡派跟进", "POD"], "P1", "有原型", ["尾单原型", "POD下载"], ""),
        ("海外", ["改址", "拦截"], "P1", "有原型", ["尾单原型", "纪要0804"], ""),
        ("海外", ["直送", "分公司", "抓取"], "P1", "有方案", ["海外方案", "尾单原型"], ""),
        ("海外", [], "P1", "有方案", ["海外方案", "尾单原型"], "默认挂海外方案+尾单"),
        # —— 船务 ——
        ("船务", ["取消IT", "取消 IT"], "P0", "有原型", ["提单管理原型", "船务方案", "纪要0730"], "取消IT标记→关务清关待办；截单可补中转港"),
        ("船务", ["分配提单", "提单分配", "按周", "同船"], "P0", "有原型", ["提单管理原型", "船务方案", "纪要0730"], "按周+同船聚合；分配消息一键复制拖车代理群"),
        ("船务", ["AN", "进口商", "收发通"], "P0", "有原型", ["船务方案", "进口商", "清关行港口配置", "柜子绑进口商", "纪要0730"], ""),
        ("船务", ["AMS", "ISF"], "P1", "有方案", ["船务方案", "改配"], ""),
        ("船务", ["运价", "合约", "仓位"], "P0", "有原型", ["运价管理原型", "船务方案", "纪要0730"], "我司订舱价+代理订舱价+价差对比；合约导入/同步报价组"),
        ("船务", ["赶进港", "截单", "截关", "落箱", "预警"], "P0", "有方案", ["船务方案"], ""),
        ("船务", ["滞箱", "滞港", "免用箱"], "P0", "有方案", ["船务方案", "海外方案"], "与秋玉侧费用衔接"),
        ("船务", ["查验", "异常", "报表"], "P1", "有原型", ["查验报表原型", "船务方案", "关务查验"], "提单/运单双维度；国内国外分统计；导出模版"),
        ("船务", ["进港代码", "海放", "码放", "甩柜"], "P1", "有方案", ["船务方案", "改配"], ""),
        ("船务", ["供应商", "local", "放箱"], "P1", "有原型", ["运价管理原型", "船务方案"], "代理Local/收费标准在运价管理·代理订舱价"),
        ("船务", ["对账单", "费用", "提单格式", "船期"], "P2", "有方案", ["船务方案"], ""),
        ("船务", ["业务数据统计", "字段"], "P1", "有方案", ["船务方案", "全流程"], ""),
        ("船务", [], "P1", "有方案", ["船务方案"], "默认挂船务方案"),
        # —— 跨组 ——
        ("跨组", ["导出50000", "导出"], "P2", "待设计", ["全流程", "导航"], "平台能力"),
        ("跨组", [], "P2", "有原型", ["全流程", "系统流程设计", "价值流图", "导航"], ""),
    ]

    for it in items:
        text = (it["title"] + "\n" + it["desc"]).lower()
        g = it["group"]
        best = None
        for rg, kws, pri, status, pkeys, note in rules:
            if rg not in g and rg != "跨组":
                continue
            if rg == "跨组" and g not in ("跨组/平台", "重复机械工作", "菜单与取数"):
                continue
            if kws and not any(k.lower() in text for k in kws):
                continue
            best = (pri, status, pkeys, note)
            break
        if best:
            pri, status, pkeys, note = best
            it["priority"] = pri
            it["status"] = status
            if pkeys:
                it["proto"] = proto(*pkeys)
            if note:
                it["note"] = note
        # research overlays always for known titles
        overlay(it)
    return items


def overlay(it: dict) -> None:
    t = it["title"]
    g = it["group"]
    if g == "报价组" and "时长看板" in t:
        it.update(priority="P0", status="有原型", proto=proto("报价看板", "纪要0731"), note="考核看板")
    if g == "报价组" and "绑" in t:
        it.update(priority="P0", status="有原型", proto=proto("报价员绑运单", "整柜绑报价", "私卡0731"))
    if "私仓价格" in t or "尾端派送成本" in t:
        it.update(
            group="海外对接组",
            priority="P0",
            status="待原型",
            proto=proto("海外方案", "协议价原型", "纪要0804"),
            note="柜子维度费用预估（非尾单单票页）",
        )
    if "分单号" in t:
        it.update(
            priority="P0",
            status="有原型",
            proto=proto("分单号自动生成", "分单号规则配置", "分单号", "关务方案汇总", "报关列表"),
            note="自动生成+货物调整重算+推送报关行同步",
        )
    # 强制挂新产出原型（避免被宽规则冲掉）
    if g == "风控组":
        it["proto"] = merge_proto(it.get("proto"), "查货操作AI识图", "风控方案")
        if any(k in (t + it.get("desc", "")) for k in ("查货率", "漏查", "100%")):
            it["proto"] = merge_proto(it.get("proto"), "查货率报表")
            it["status"] = "有原型"
        if any(k in (t + it.get("desc", "")) for k in ("智能体", "识图", "品名", "侵权", "查货操作", "附加费")):
            it["status"] = "有原型" if it["status"] in ("待设计", "有方案", "方案中") else it["status"]
    if any(k in t + it.get("desc", "") for k in ("收发通", "进口商", "清关行")):
        it["proto"] = merge_proto(it.get("proto"), "进口商", "清关行港口配置", "柜子绑进口商")
        it["status"] = "有原型"
        if "邮箱" in t or "密码" in t or "收件人" in t:
            it["note"] = (it.get("note") or "") and it["note"] or "收件人含邮箱/密码"
    if g == "关务组" and any(k in t + it.get("desc", "") for k in ("取消IT", "取消 IT", "一键通知", "清关查验", "报关查验", "通知客户")):
        it["proto"] = merge_proto(it.get("proto"), "清关管理", "报关列表", "关务查验")
        it["status"] = "有原型"
    if g == "关务组" and any(k in t + it.get("desc", "") for k in ("报关", "委托", "查验")):
        it["proto"] = merge_proto(it.get("proto"), "报关列表", "报关资料准备", "报关流程图")
    if g == "船务组" and any(k in t + it.get("desc", "") for k in ("取消IT", "取消 IT", "分配提单", "提单分配", "按周", "同船")):
        it["proto"] = merge_proto(it.get("proto"), "提单管理原型", "船务方案", "纪要0730")
        it["status"] = "有原型"
        if "取消IT" in t or "取消 IT" in t:
            it["note"] = "取消IT标记→关务清关待办；截单可补中转港"
        elif any(k in t + it.get("desc", "") for k in ("分配", "按周", "同船")):
            it["note"] = "按周+同船聚合；分配消息一键复制拖车代理群"
    if g == "船务组" and any(k in t + it.get("desc", "") for k in ("运价", "合约价", "local", "Local", "代理收费", "订舱数据")):
        it["proto"] = merge_proto(it.get("proto"), "运价管理原型", "船务方案", "纪要0730")
        if it.get("status") in ("待设计", "有方案", "方案中", ""):
            it["status"] = "有原型"
        if "local" in (t + it.get("desc", "")).lower() or "代理收费" in t:
            it["note"] = "我司/代理订舱价+Local分项+价差对比"
    if any(k in t + it.get("desc", "") for k in (
        "查验报表", "查验异常", "查验统计", "半月度查验", "月度查验",
        "查验的异常", "查验进度", "查验原因", "查验内容",
    )):
        it["proto"] = merge_proto(it.get("proto"), "查验报表原型", "关务查验", "船务方案")
        if it.get("status") in ("待设计", "有方案", "方案中", ""):
            it["status"] = "有原型"
        it["note"] = (it.get("note") or "") or "提单+运单双维度；国内/国外分统计；原因/品名/HS/货值/天数"
    if "委托" in t and g == "关务组":
        it.update(
            priority="P1",
            status="有原型",
            note="对接#12按客户名称+报关行查询回写状态；未对接委托标记+截图；无委托不可报关",
        )
        it["proto"] = merge_proto(it.get("proto"), "报关列表", "关务方案汇总")
    if any(k in t + it.get("desc", "") for k in ("买单", "件毛体", "分票", "资料核准", "预录")):
        if "委托" not in t:
            it["proto"] = merge_proto(it.get("proto"), "做资料分票核准", "资料核准")
            if it.get("status") in ("待设计", "有方案", "方案中", ""):
                it["status"] = "有原型"
        else:
            it["proto"] = merge_proto(it.get("proto"), "做资料分票核准", "资料核准")
    if any(k in t + it.get("desc", "") for k in ("POD", "pod", "签收预警", "出库/签收")):
        it["proto"] = merge_proto(it.get("proto"), "出库签收预警", "POD下载")


def merge_proto(existing: str | None, *keys: str) -> str:
    """Merge proto keys into existing proto string without duplicates."""
    have = set()
    parts = []
    for part in (existing or "").split("；"):
        if "|" in part:
            n, _p = part.split("|", 1)
            if n not in have and n in LINKS:
                have.add(n)
                parts.append(f"{n}|{LINKS[n]}")
    for k in keys:
        if k in LINKS and k not in have:
            have.add(k)
            parts.append(f"{k}|{LINKS[k]}")
    return "；".join(parts)


def research_extras() -> list[dict]:
    """Requirements from research notes not fully covered by 7.31 sheet."""
    return [
        {
            "group": "报价组",
            "title": "新AI报价：本州/外州规则配置 + 确认 + 尾端派送费 + 散货多绑",
            "desc": "旧AI取消；新AI按本州/外州配置出价；人工确认并登记尾端派送费；散货可绑多运单，签入自动挂费供港后导出；有效期后台可配、复核台可改。",
            "source": "0731 调研 + 8/6 卢慧恒/袁华君确认",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("报价员绑运单", "AI报价规则配置", "私卡0731", "纪要0731"),
            "note": "测距离本期不做；含有效期与绑运单",
        },
        {
            "group": "海外对接组",
            "title": "港后尾单单票跟踪：提柜后四桶 + 派送时效 + Excel 核心列",
            "desc": "没约/>15天改快递/有约/已送达；拆柜·约仓·出库·签收；周报表。",
            "source": "调研纪要 2026-08-04",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("尾单原型", "纪要0804"),
            "note": "",
        },
        {
            "group": "海外对接组",
            "title": "港前 HOLD 抓取与提柜预警；DO/PU 资料齐套",
            "desc": "码头 HOLD 抓取；资料不齐不可约；本页只读齐套结果。",
            "source": "调研纪要 2026-08-04",
            "priority": "P0",
            "status": "有方案",
            "proto": proto("海外方案", "纪要0804"),
            "note": "港前准备台待原型",
        },
        {
            "group": "海外对接组",
            "title": "目的港邮箱工作台（~103）：滞港/滞箱/AN/水单→业务员成本",
            "desc": "邮箱洪峰处理；DO/AN 上传推仓；合同不签不放行。",
            "source": "调研纪要 2026-08-04",
            "priority": "P0",
            "status": "有方案",
            "proto": proto("海外方案", "纪要0804"),
            "note": "秋玉侧",
        },
        {
            "group": "海外对接组",
            "title": "客户协议价格维护 + 批量改时间/金额 + 大客户调成本",
            "desc": "价格表批量；特价仓留痕；申请价差额落成本。",
            "source": "调研纪要 2026-08-04 + 原型",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("协议价原型", "海外方案"),
            "note": "",
        },
        {
            "group": "海外对接组",
            "title": "私卡尾端费用预估自动比对（柜子维度）",
            "desc": "人工比对痛点；系统预估 vs 实际差异；不在尾单单票页。含「柜子计划私仓价格错误、需核对尾端派送成本」问题。",
            "source": "调研纪要 2026-08-04 + 部门问题汇总",
            "priority": "P0",
            "status": "待原型",
            "proto": proto("海外方案", "纪要0804"),
            "note": "待柜页截图",
        },
        {
            "group": "船务组",
            "title": "运价上系统时机：配舱后回写；费用节点与进港预警",
            "desc": "0730：合约/运价、赶进港、截单截关落箱预警；救火→防火。",
            "source": "调研纪要 2026-07-30",
            "priority": "P0",
            "status": "有方案",
            "proto": proto("船务方案", "纪要0730"),
            "note": "",
        },
        {
            "group": "关务组",
            "title": "资料核准双通道 + 改配三类场景轨迹同步",
            "desc": "0728 调研落地：核准双通道、改配轨迹、三表线上化。",
            "source": "调研纪要 2026-07-28/29",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("做资料分票核准", "资料核准", "改配", "三表", "流程图汇总", "做资料全流程"),
            "note": "双通道+买单轧差落报关管理",
        },
        {
            "group": "风控组",
            "title": "查货操作列表：AI识图 + 品名识别 + 添加成本费用",
            "desc": "在现网查货操作上叠加 AI识图/品名识别（智能体辅助），并增加添加成本费用入口（同行应收+成本 / 直客业务成本，超品名费×个数）；附加费可在查货时备注，减少找客服。",
            "source": "调研纪要 2026-07-29 + 现网截图原型 + 部门问题汇总",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("查货操作AI识图", "风控方案", "纪要0729"),
            "note": "豆包辅助，不作唯一卡控；合并原智能体/附加费/超品名费条目",
        },
        {
            "group": "风控组",
            "title": "操作查货率报表：总票/空白/保函/漏查 · 100%目标",
            "desc": "按仓汇总总票与查验空白；空白票核对保函（有=豁免，无=漏查）；查货率=实查/应查；支持下钻、模板导出、日报播报。覆盖原「仓库查货率100%」「查货率表格统计步骤」。",
            "source": "部门问题汇总 + 调研",
            "priority": "P1",
            "status": "有原型",
            "proto": proto("查货率报表", "查货操作AI识图", "风控方案"),
            "note": "替代手工筛表；合并原查货率重复条目",
        },
        {
            "group": "关务组",
            "title": "进口商主数据：收件人/发件人绑定 + 清关行·港口配置",
            "desc": "发货人抬头改为收件人；发件人维护与多对多绑定；结合分析表/邮箱分配做进口商×清关行×港口配置。",
            "source": "进口商分析统计表 + 清关行及邮箱分配 + 原型",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("进口商", "清关行港口配置", "柜子绑进口商"),
            "note": "柜子绑定侧同步展示收件人",
        },
        {
            "group": "跨组/平台",
            "title": "产品部门标准化全流程（泳道 + 价值流）评审基线",
            "desc": "跨组主流程与系统落点总图，评审主入口。",
            "source": "产品架构",
            "priority": "P0",
            "status": "有原型",
            "proto": proto("全流程", "系统流程设计", "价值流图", "导航"),
            "note": "总览",
        },
    ]


def _norm_title_key(t: str) -> str:
    t = clean(t)
    t = re.sub(r"^[\d\.、．\s]+", "", t)
    t = re.sub(r"[\s\u3000]+", "", t)
    return t.lower()


def _item_score(it: dict) -> tuple:
    """Higher = prefer keep when deduping."""
    st = {"有原型": 4, "验收中": 3, "有方案": 2, "方案中": 1, "待原型": 1, "待设计": 0}.get(it.get("status") or "", 0)
    proto_n = len([p for p in (it.get("proto") or "").split("；") if p.strip()])
    src_bonus = 2 if any(k in (it.get("source") or "") for k in ("调研", "原型", "8/6", "0731", "产品架构")) else 0
    pri = {"P0": 3, "P1": 2, "P2": 1}.get(it.get("priority") or "", 0)
    return (st, proto_n, src_bonus, pri, len(it.get("desc") or ""))


def dedupe_items(items: list[dict]) -> list[dict]:
    """Remove stub rows and merge near-duplicate requirements within a group."""
    # 1) exact normalized title → keep highest score, merge proto
    best: dict[tuple[str, str], dict] = {}
    order: list[tuple[str, str]] = []
    for it in items:
        key = (it["group"], _norm_title_key(it["title"]))
        if key not in best:
            best[key] = it
            order.append(key)
            continue
        cur = best[key]
        winner, loser = (it, cur) if _item_score(it) > _item_score(cur) else (cur, it)
        winner["proto"] = merge_proto(
            winner.get("proto"),
            *[p.split("|", 1)[0] for p in (loser.get("proto") or "").split("；") if "|" in p],
        )
        if loser.get("note") and loser["note"] not in (winner.get("note") or ""):
            winner["note"] = ((winner.get("note") or "") + "；" + loser["note"]).strip("；")
        best[key] = winner
    items = [best[k] for k in order]

    # 2) topic absorb：调研落地原型吞掉表内碎片条目
    absorb_rules = [
        {
            "group": "报价组",
            "keep": lambda it: "新AI报价" in it["title"],
            "drop": lambda it: (
                ("AI" in it["title"] and "报价" in it["title"] and "新AI" not in it["title"])
                or ("绑定运单" in it["title"] or "绑运单" in it["title"])
                or ("有效期" in it["title"] and "报价" in it["title"])
            ),
        },
        {
            "group": "风控组",
            "keep": lambda it: "操作查货率报表" in it["title"],
            "drop": lambda it: "查货率" in it["title"] and "操作查货率报表" not in it["title"],
        },
        {
            "group": "风控组",
            "keep": lambda it: "查货操作列表" in it["title"] and "AI识图" in it["title"],
            "drop": lambda it: (
                it["title"].strip() in ("智能体",)
                or "超品名" in it["title"]
                or it["title"].startswith("附加费手动")
                or ("主要是查货操作" in it["title"] and "加收附加费" in it["title"])
            ),
        },
        {
            "group": "海外对接组",
            "keep": lambda it: "私卡尾端费用预估" in it["title"] or "费用预估自动比对" in it["title"],
            "drop": lambda it: ("私仓价格" in it["title"] and "预估" not in it["title"])
            or ("尾端派送成本" in it["title"] and "预估" not in it["title"]),
        },
    ]

    drop_ids = set()
    for rule in absorb_rules:
        keepers = [it for it in items if it["group"] == rule["group"] and rule["keep"](it)]
        if not keepers:
            continue
        keeper = max(keepers, key=_item_score)
        for it in items:
            if it is keeper or it["group"] != rule["group"]:
                continue
            if not rule["drop"](it):
                continue
            frag = first_line(it["title"], 40)
            note = keeper.get("note") or ""
            if frag and frag not in note:
                keeper["note"] = (note + ("；" if note else "") + "已合并：" + frag).strip("；")
            keeper["proto"] = merge_proto(
                keeper.get("proto"),
                *[p.split("|", 1)[0] for p in (it.get("proto") or "").split("；") if "|" in p],
            )
            drop_ids.add(id(it))

    return [it for it in items if id(it) not in drop_ids]


def build_all(data: dict) -> list[dict]:
    items: list[dict] = []
    items += from_problem_rows("风控组", load_sheet_rows(data, "风控组"), "部门问题汇总表7.31 · 风控组", "P1")
    items += from_problem_rows("报价组", load_sheet_rows(data, "报价组"), "部门问题汇总表7.31 · 报价组", "P1")
    items += from_problem_rows("关务组", load_sheet_rows(data, "关务问题"), "部门问题汇总表7.31 · 关务问题", "P1")
    items += from_problem_rows("关务组", load_sheet_rows(data, "国内查验"), "部门问题汇总表7.31 · 国内查验（案例沉淀）", "P2")
    items += from_problem_rows("海外对接组", load_sheet_rows(data, "海外组问题"), "部门问题汇总表7.31 · 海外组问题", "P1")
    items += from_problem_rows("船务组", load_sheet_rows(data, "船务组"), "部门问题汇总表7.31 · 船务组", "P1")

    # cross sheets
    for sheet, g in [
        ("重复的机械性的工作", "跨组/平台"),
        ("各业务线的功能菜单页面", "跨组/平台"),
        ("取消IT等时间节点提醒功能", "跨组/平台"),
    ]:
        items += from_problem_rows(g, load_sheet_rows(data, sheet), f"部门问题汇总表7.31 · {sheet}", "P2")

    items = enrich(items)
    # prepend research extras (dedupe by title loosely)
    seen = {_norm_title_key(i["title"]) for i in items}
    for ex in research_extras():
        key = _norm_title_key(ex["title"])
        if key not in seen:
            items.insert(0, ex)
            seen.add(key)

    items = dedupe_items(items)

    # assign IDs
    counters: dict[str, int] = {}
    prefix = {
        "风控组": "FK",
        "报价组": "BJ",
        "关务组": "GW",
        "海外对接组": "HW",
        "船务组": "CW",
        "跨组/平台": "PT",
    }
    for it in items:
        g = it["group"]
        counters[g] = counters.get(g, 0) + 1
        it["id"] = f"{prefix.get(g, 'X')}{counters[g]:03d}"
    return items


def write_excel(items: list[dict]) -> None:
    wb = Workbook()
    # summary
    ws0 = wb.active
    ws0.title = "总览"
    headers = [
        "需求ID",
        "组别",
        "优先级",
        "状态",
        "需求标题",
        "来源",
        "问题/方案摘要",
        "原型/方案（名称）",
        "原型路径",
        "备注",
    ]
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    head_fill = PatternFill("solid", fgColor="003A8C")
    head_font = Font(color="FFFFFF", bold=True, size=11)
    fills = {
        "P0": PatternFill("solid", fgColor="FFF1F0"),
        "P1": PatternFill("solid", fgColor="FFF7E6"),
        "P2": PatternFill("solid", fgColor="F6FFED"),
    }
    pri_font = {
        "P0": Font(color="CF1322", bold=True),
        "P1": Font(color="D46B08", bold=True),
        "P2": Font(color="389E0D"),
    }

    def write_sheet(ws, rows_items):
        ws.append(headers)
        for c in ws[1]:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for it in rows_items:
            names, paths = [], []
            for part in (it.get("proto") or "").split("；"):
                if "|" in part:
                    n, p = part.split("|", 1)
                    names.append(n)
                    paths.append(p)
            ws.append(
                [
                    it["id"],
                    it["group"],
                    it["priority"],
                    it["status"],
                    it["title"],
                    it["source"],
                    clean(it["desc"])[:500],
                    " / ".join(names),
                    "\n".join(paths),
                    it.get("note") or "",
                ]
            )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            pri = row[2].value
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if pri in fills:
                row[2].fill = fills[pri]
                row[2].font = pri_font[pri]
        widths = [10, 12, 8, 10, 36, 28, 48, 28, 40, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    # order groups
    order = ["风控组", "关务组", "报价组", "海外对接组", "船务组", "跨组/平台"]
    write_sheet(ws0, sorted(items, key=lambda x: (order.index(x["group"]) if x["group"] in order else 99, x["priority"], x["id"])))

    for g in order:
        subset = [i for i in items if i["group"] == g]
        if not subset:
            continue
        ws = wb.create_sheet(g.replace("/", "-")[:31])
        write_sheet(ws, sorted(subset, key=lambda x: (x["priority"], x["id"])))

    # priority matrix
    ws = wb.create_sheet("优先级统计")
    ws.append(["组别", "P0", "P1", "P2", "合计", "有原型", "有方案", "待设计/待原型"])
    for c in ws[1]:
        c.fill = head_fill
        c.font = head_font
    for g in order:
        subset = [i for i in items if i["group"] == g]
        def cnt(pred):
            return sum(1 for i in subset if pred(i))
        ws.append(
            [
                g,
                cnt(lambda i: i["priority"] == "P0"),
                cnt(lambda i: i["priority"] == "P1"),
                cnt(lambda i: i["priority"] == "P2"),
                len(subset),
                cnt(lambda i: i["status"] == "有原型"),
                cnt(lambda i: i["status"] in ("有方案", "方案中", "验收中")),
                cnt(lambda i: i["status"] in ("待设计", "待原型")),
            ]
        )
    for i, w in enumerate([14, 8, 8, 8, 8, 10, 10, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_OUT)


def write_html(items: list[dict]) -> None:
    order = ["风控组", "关务组", "报价组", "海外对接组", "船务组", "跨组/平台"]
    stats = []
    for g in order:
        subset = [i for i in items if i["group"] == g]
        stats.append(
            {
                "g": g,
                "n": len(subset),
                "p0": sum(1 for i in subset if i["priority"] == "P0"),
                "p1": sum(1 for i in subset if i["priority"] == "P1"),
                "p2": sum(1 for i in subset if i["priority"] == "P2"),
                "proto": sum(1 for i in subset if i["status"] == "有原型"),
            }
        )

    def proto_html(s: str) -> str:
        if not s:
            return '<span class="muted">—</span>'
        bits = []
        for part in s.split("；"):
            if "|" in part:
                n, p = part.split("|", 1)
                bits.append(f'<a href="{p}" target="_blank" rel="noopener">{n}</a>')
        return " · ".join(bits) if bits else "—"

    sections = []
    for g in order:
        subset = sorted(
            [i for i in items if i["group"] == g],
            key=lambda x: (x["priority"], x["id"]),
        )
        if not subset:
            continue
        cards = []
        for it in subset:
            cards.append(
                f"""
        <article class="card pri-{it['priority'].lower()}" id="{it['id']}">
          <header>
            <span class="id">{it['id']}</span>
            <span class="pri">{it['priority']}</span>
            <span class="st">{it['status']}</span>
            <h3>{it['title']}</h3>
          </header>
          <p class="src">来源：{it['source']}</p>
          <p class="desc">{clean(it['desc'])[:420].replace(chr(10), '<br/>')}</p>
          <div class="links"><b>原型/方案：</b> {proto_html(it.get('proto') or '')}</div>
          {"<p class='note'>备注：" + it['note'] + "</p>" if it.get("note") else ""}
        </article>"""
            )
        sections.append(
            f'<section class="group" id="g-{g.replace("/", "-")}" data-group="{g}"><h2>{g} <small>{len(subset)} 条</small></h2><div class="grid">{"".join(cards)}</div></section>'
        )

    stat_html = "".join(
        f'<a class="stat" href="#g-{s["g"].replace("/", "-")}"><b>{s["g"]}</b><span>合计 {s["n"]}</span>'
        f'<span class="p0">P0 {s["p0"]}</span><span class="p1">P1 {s["p1"]}</span>'
        f'<span class="p2">P2 {s["p2"]}</span><span>原型 {s["proto"]}</span></a>'
        for s in stats
    )

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>产品部门 · 各组需求清单（评审版）</title>
  <style>
    :root {{
      --bg:#f0f2f5; --card:#fff; --line:#e8e8e8; --text:rgba(0,0,0,.85); --sub:rgba(0,0,0,.45);
      --blue:#1890ff; --p0:#cf1322; --p1:#d46b08; --p2:#389e0d;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:"Microsoft YaHei","PingFang SC",Arial,sans-serif; background:var(--bg); color:var(--text); font-size:13px; }}
    a {{ color:var(--blue); text-decoration:none; }}
    a:hover {{ text-decoration:underline; }}
    .top {{
      position:sticky; top:0; z-index:10; background:#001529; color:#fff;
      padding:10px 16px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;
    }}
    .top h1 {{ margin:0; font-size:16px; font-weight:600; }}
    .top .meta {{ color:rgba(255,255,255,.7); font-size:12px; }}
    .top .acts {{ margin-left:auto; display:flex; gap:8px; flex-wrap:wrap; }}
    .top .acts a {{
      color:#fff; background:rgba(255,255,255,.12); padding:4px 10px; border-radius:2px; font-size:12px;
    }}
    .wrap {{ max-width:1200px; margin:0 auto; padding:14px 16px 48px; }}
    .banner {{
      background:#fff; border:1px solid var(--line); border-radius:6px; padding:12px 14px; margin-bottom:12px; line-height:1.65;
    }}
    .banner b {{ color:#003a8c; }}
    .filters {{
      display:flex; flex-wrap:wrap; gap:8px; margin-bottom:12px; align-items:center;
      background:#fff; border:1px solid var(--line); border-radius:6px; padding:10px 12px;
    }}
    .filters label {{ color:var(--sub); font-size:12px; }}
    .filters select, .filters input {{
      height:28px; border:1px solid #d9d9d9; border-radius:2px; padding:0 8px; font:inherit;
    }}
    .stats {{ display:grid; grid-template-columns:repeat(3,1fr); gap:8px; margin-bottom:14px; }}
    @media (max-width:900px) {{ .stats {{ grid-template-columns:1fr 1fr; }} }}
    .stat {{
      background:#fff; border:1px solid var(--line); border-radius:6px; padding:10px 12px;
      display:flex; flex-direction:column; gap:4px; color:inherit;
    }}
    .stat:hover {{ border-color:var(--blue); text-decoration:none; }}
    .stat b {{ font-size:14px; }}
    .stat span {{ font-size:12px; color:var(--sub); }}
    .stat .p0 {{ color:var(--p0); font-weight:700; }}
    .stat .p1 {{ color:var(--p1); font-weight:700; }}
    .stat .p2 {{ color:var(--p2); font-weight:700; }}
    .group {{ margin-bottom:18px; }}
    .group h2 {{ margin:0 0 10px; font-size:16px; }}
    .group h2 small {{ color:var(--sub); font-weight:400; margin-left:6px; }}
    .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:10px; }}
    @media (max-width:900px) {{ .grid {{ grid-template-columns:1fr; }} }}
    .card {{
      background:var(--card); border:1px solid var(--line); border-radius:6px; padding:12px 14px;
      border-left:4px solid #d9d9d9;
    }}
    .card.pri-p0 {{ border-left-color:var(--p0); background:#fffafa; }}
    .card.pri-p1 {{ border-left-color:var(--p1); }}
    .card.pri-p2 {{ border-left-color:var(--p2); }}
    .card header {{ display:flex; flex-wrap:wrap; gap:6px; align-items:center; margin-bottom:6px; }}
    .card h3 {{ margin:4px 0 0; font-size:14px; width:100%; font-weight:600; }}
    .id {{ font-size:11px; color:#003a8c; font-weight:700; background:#e6f7ff; padding:1px 6px; border-radius:2px; }}
    .pri {{ font-size:11px; font-weight:700; padding:1px 6px; border-radius:2px; border:1px solid; }}
    .pri-p0 .pri, .card .pri {{ }}
    .card.pri-p0 .pri {{ color:var(--p0); background:#fff1f0; border-color:#ffa39e; }}
    .card.pri-p1 .pri {{ color:var(--p1); background:#fff7e6; border-color:#ffd591; }}
    .card.pri-p2 .pri {{ color:var(--p2); background:#f6ffed; border-color:#b7eb8f; }}
    .st {{ font-size:11px; color:var(--sub); background:#fafafa; border:1px solid var(--line); padding:1px 6px; border-radius:2px; }}
    .src {{ margin:0 0 6px; font-size:11px; color:var(--sub); }}
    .desc {{ margin:0 0 8px; font-size:12px; color:#334155; line-height:1.55; max-height:7.5em; overflow:auto; }}
    .links {{ font-size:12px; }}
    .note {{ margin:6px 0 0; font-size:11px; color:#614700; background:#fffbe6; border:1px solid #ffe58f; padding:6px 8px; border-radius:2px; }}
    .muted {{ color:var(--sub); }}
    .card.hide {{ display:none; }}
    .group.hide {{ display:none; }}
  </style>
</head>
<body>
  <div class="top">
    <h1>产品部门 · 各组需求清单（评审版）</h1>
    <span class="meta">来源：部门问题汇总表7.31 + 调研纪要（0727–0804）· 点击原型跳转评审</span>
    <div class="acts">
      <a href="产品部门-各组需求清单-评审版.xlsx">下载 Excel</a>
      <a href="产品部门-导航.html">返回导航</a>
      <a href="各组sop/产品部门-标准化全流程-评审版.html">全流程评审版</a>
    </div>
  </div>
  <div class="wrap">
    <div class="banner">
      <b>用法：</b>按组别浏览 → 优先看红色 <b>P0</b> → 点「原型/方案」打开对应 HTML 投屏。
      优先级口径：P0=调研确认痛点且影响出货/费用/合规；P1=明确优化可排期；P2=体验/案例沉淀/平台能力。
      Excel 含总览、各组页、优先级统计。共 <b>{len(items)}</b> 条。
      <div style="margin-top:8px;font-size:12px;line-height:1.7">
        <b>已产出原型索引（均已挂到对应条目）：</b>
        <a href="各组sop/风控组-查货操作-AI识图品名识别-原型.html" target="_blank">查货操作·AI识图</a> ·
        <a href="各组sop/风控组-操作查货率报表-MVP.html" target="_blank">操作查货率报表</a> ·
        <a href="../进口商管理-MVP.html" target="_blank">进口商</a> ·
        <a href="../进口商-清关行港口配置-MVP.html" target="_blank">清关行·港口配置</a> ·
        <a href="../柜子编辑-绑定进口商-MVP.html" target="_blank">柜子绑进口商</a> ·
        <a href="各组sop/海外对接组-港后-尾单单票跟踪-原型.html" target="_blank">尾单跟踪</a> ·
        <a href="各组sop/海外对接组-客户协议价格维护-原型.html" target="_blank">协议价</a> ·
        <a href="../二期/ESS询价报价-报价员列表-MVP.html" target="_blank">报价员·AI确认/多绑</a> ·
        <a href="各组sop/私卡报价-AI报价规则配置-MVP.html" target="_blank">AI报价规则配置</a> ·
        <a href="../整柜下单/整柜询价与报价绑定-MVP.html" target="_blank">整柜绑报价</a> ·
        <a href="报价时效统计看板-MVP.html" target="_blank">报价看板</a> ·
        <a href="../报关行对接/报关管理-列表-MVP.html" target="_blank">报关列表</a> ·
        <a href="../清关管理-MVP.html" target="_blank">清关管理·取消IT/查验通知</a> ·
        <a href="../报关行对接/报关管理-资料准备-列表-MVP.html" target="_blank">报关资料准备</a> ·
        <a href="各组sop/关务组-查验报表-提单运单维度-MVP.html" target="_blank">查验报表·提单/运单</a> ·
        <a href="各组sop/船务组-运价管理-我司与代理订舱-MVP.html" target="_blank">运价管理·我司/代理订舱</a> ·
        <a href="各组sop/船务组-提单管理-按周分组与分配-MVP.html" target="_blank">提单管理·按周/取消IT/分配</a> ·
        <a href="../B21-出库签收预警-MVP.html" target="_blank">出库签收预警</a> ·
        <a href="../二期/订单管理-POD下载-MVP.html" target="_blank">POD下载</a> ·
        <a href="各组sop/产品部门-标准化全流程-评审版.html" target="_blank">全流程</a>
      </div>
    </div>
    <div class="filters">
      <label>组别</label>
      <select id="fGroup">
        <option value="">全部</option>
        {''.join(f'<option value="{g}">{g}</option>' for g in order)}
      </select>
      <label>优先级</label>
      <select id="fPri">
        <option value="">全部</option>
        <option value="P0">P0</option>
        <option value="P1">P1</option>
        <option value="P2">P2</option>
      </select>
      <label>状态</label>
      <select id="fSt">
        <option value="">全部</option>
        <option>有原型</option>
        <option>有方案</option>
        <option>方案中</option>
        <option>验收中</option>
        <option>待原型</option>
        <option>待设计</option>
      </select>
      <label>搜索</label>
      <input id="fQ" placeholder="标题/摘要关键词" style="min-width:200px" />
      <button type="button" id="btnReset" style="height:28px;padding:0 10px;border:1px solid #d9d9d9;background:#fff;border-radius:2px;cursor:pointer;">重置</button>
    </div>
    <div class="stats">{stat_html}</div>
    {''.join(sections)}
  </div>
  <script>
    (function () {{
      function apply() {{
        var g = document.getElementById('fGroup').value;
        var p = document.getElementById('fPri').value;
        var s = document.getElementById('fSt').value;
        var q = (document.getElementById('fQ').value || '').trim().toLowerCase();
        document.querySelectorAll('.group').forEach(function (sec) {{
          var showSec = !g || sec.getAttribute('data-group') === g;
          var any = false;
          sec.querySelectorAll('.card').forEach(function (card) {{
            var ok = showSec;
            if (ok && p && !card.classList.contains('pri-' + p.toLowerCase())) ok = false;
            if (ok && s && card.querySelector('.st').textContent !== s) ok = false;
            if (ok && q) {{
              var text = card.textContent.toLowerCase();
              if (text.indexOf(q) < 0) ok = false;
            }}
            card.classList.toggle('hide', !ok);
            if (ok) any = true;
          }});
          sec.classList.toggle('hide', !any);
        }});
      }}
      ['fGroup','fPri','fSt','fQ'].forEach(function (id) {{
        document.getElementById(id).addEventListener('input', apply);
        document.getElementById(id).addEventListener('change', apply);
      }});
      document.getElementById('btnReset').addEventListener('click', function () {{
        fGroup.value = fPri.value = fSt.value = fQ.value = '';
        apply();
      }});
    }})();
  </script>
</body>
</html>
"""
    HTML_OUT.write_text(html, encoding="utf-8")


def main() -> None:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    items = build_all(data)
    write_excel(items)
    write_html(items)
    # quick stats
    from collections import Counter

    c = Counter(i["group"] for i in items)
    p = Counter(i["priority"] for i in items)
    print("groups", dict(c))
    print("priority", dict(p))
    print("xlsx", XLSX_OUT)
    print("html", HTML_OUT)
    print("total", len(items))


if __name__ == "__main__":
    main()
