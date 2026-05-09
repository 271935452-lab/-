# -*- coding: utf-8 -*-
"""在各场景子表 Sxxx 的步骤列区插入「步骤4·规则包命中」，原「命中规则」起顺延为步骤5–8。

若「step34合并.xlsx」被 Excel 占用无法覆盖保存，脚本会另存为
「step34合并_rulepkg命中列.xlsx」；关闭占用后可手动替换主文件。
"""
import re
import shutil
from openpyxl import load_workbook

SRC = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并.xlsx"
BAK = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_before_rulepkg_hit_col.xlsx"
OUT_ALT = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并_rulepkg命中列.xlsx"


def find_hdr_row(ws):
    for r in range(1, min(ws.max_row + 1, 45)):
        if ws.cell(r, 10).value == "步骤1·主档对齐":
            return r
    return None


def split_old_hit_cell(val):
    """原单列「步骤4·命中规则」拆成：规则包命中（编排/包编码） + 包内子规则命中文案。"""
    if val is None:
        return None, None
    if not isinstance(val, str):
        return None, val
    s = val.strip()
    if not s:
        return None, None
    if " · 命中 " in s:
        left, right = s.split(" · 命中 ", 1)
        return left.strip(), ("命中 " + right.strip()).strip()
    m_rp = re.search(r"（见上行·(RP_\w+)）", s)
    seg_m = re.match(r"^(【顺序执行·段\d+】)", s)
    if m_rp and seg_m:
        pkg = seg_m.group(1) + "规则包 " + m_rp.group(1)
        rest = s[len(seg_m.group(1)) :].strip()
        return pkg, rest
    return None, s


def bump_step_refs(s):
    if not isinstance(s, str):
        return s
    s = s.replace("步骤6/7", "步骤7/8")
    s = s.replace("Σ步骤7", "Σ步骤8")
    s = s.replace("（见步骤7）", "（见步骤8）")
    s = s.replace("见步骤7", "见步骤8")
    return s


def process_sheet(ws):
    hdr = find_hdr_row(ws)
    if not hdr:
        return False
    ws.insert_cols(13)
    ws.cell(hdr, 13).value = "步骤4·规则包命中"
    ws.cell(hdr, 14).value = "步骤5·命中规则"
    ws.cell(hdr, 15).value = "步骤6·基数加工"
    ws.cell(hdr, 16).value = "步骤7·提成试算"
    ws.cell(hdr, 17).value = "步骤8·扣减后应发"

    max_r = ws.max_row
    for r in range(hdr + 1, max_r + 1):
        pkg, rule = split_old_hit_cell(ws.cell(r, 14).value)
        if pkg:
            ws.cell(r, 13).value = pkg
        if rule is not None:
            ws.cell(r, 14).value = rule

    max_c = ws.max_column
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, str):
                cell.value = bump_step_refs(cell.value)
    return True


def main():
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=False)
    pat = re.compile(r"^S\d+$")
    n = 0
    for name in wb.sheetnames:
        if pat.match(name):
            if process_sheet(wb[name]):
                n += 1
    try:
        wb.save(SRC)
        saved = SRC
    except PermissionError:
        wb.save(OUT_ALT)
        saved = OUT_ALT
        print("警告：源文件被占用，已另存为副本，关闭 Excel 后可手动替换原文件。")
    wb.close()
    print("Updated sheets:", n)
    print("Saved:", saved)
    print("Backup:", BAK)


if __name__ == "__main__":
    main()
