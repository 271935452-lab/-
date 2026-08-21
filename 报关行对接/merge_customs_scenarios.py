"""合并报关系统场景.xlsx 与 报关行场景.png 内容，写入 xlsx。"""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(r"d:\Cursor\头程项目\报关行对接")
XLSX = BASE / "报关系统场景.xlsx"
PNG = BASE / "报关行场景.png"

# 报关系统场景（来自原 Sheet1）
SYSTEM_SCENARIOS = [
    {
        "来源": "报关系统",
        "场景分类": "报关资料制作",
        "子场景": "客户初次报关，或对报关资料制作不熟练",
        "场景描述": (
            "例如：客户初次制作报关资料不熟练，推送后问题过多，会进行多次沟通修改"
            "或多次重新上传报关资料，是否会按照原单号覆盖。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关系统",
        "场景分类": "分单号绑定",
        "子场景": "原先漏勾选报关，绑定分单号后勾选",
        "场景描述": (
            "在绑定分单号后，会有客户漏勾选报关，上传了资料，询问后才进行勾选，"
            "需增加一票报关。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关系统",
        "场景分类": "改买单",
        "子场景": "改买单",
        "场景描述": (
            "例：因客观原因，原本打算报关客户，需改为买单出，后这个单号依旧存在，"
            "但是不能用这个客户的抬头报了（分单号不变，但是实际内容需要进行变更，"
            "不能再使用原先预录单）。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关系统",
        "场景分类": "测试出现问题",
        "子场景": "沟通困难",
        "场景描述": (
            "例1：绑定分单号与客户知道的单号不一致，发过去客户会有疑惑，"
            "或者觉得不是自己的不认同。"
            "例2：因需要改名发客户，与报关行沟通又出现问题，容易出现沟通错误。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关系统",
        "场景分类": "测试出现问题",
        "子场景": "下载文件重复",
        "场景描述": "预录单修改后会有多份预录单，不便寻找。",
        "优先级/备注": "",
    },
]

# 报关行场景（来自 报关行场景.png）
BROKER_SCENARIOS = [
    {
        "来源": "报关行",
        "场景分类": "签入或终配舱报关单号推送",
        "子场景": "客户合并报关正确",
        "场景描述": "报关单号推送；报关单号抓取资料。",
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "签入或终配舱报关单号推送",
        "子场景": "客户合并报关调整",
        "场景描述": (
            "例如：123一单、456一单已推送；调整为1245一单、36一单。"
            "前面的两单需要跟报关行沟通作废（申报之前）；"
            "将新的重新推送过去，我们系统以新的为准。"
        ),
        "优先级/备注": "优",
    },
    {
        "来源": "报关行",
        "场景分类": "签入或终配舱报关单号推送",
        "子场景": "客户合并报关需要拆",
        "场景描述": (
            "例如：123456一单已推送；调整为1245一单、36一单。"
            "前面的一单需要跟报关行沟通作废（申报之前）；"
            "将新的重新推送过去，我们系统以新的为准。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "签入或终配舱报关单号推送",
        "子场景": "配仓减少",
        "场景描述": (
            "例如：12、34、56三单已推送；调整为56删除。"
            "56的一单需要跟报关行沟通作废；其他不影响。"
        ),
        "优先级/备注": "优",
    },
    {
        "来源": "报关行",
        "场景分类": "签入或终配舱报关单号推送",
        "子场景": "配仓增加",
        "场景描述": (
            "例如：12、34、56三单已推送；新增78增加；"
            "78重新推送；其他不影响。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "分单号录入正确",
        "场景描述": "分单号推送；分单号抓取资料。",
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "分单号录入错误",
        "场景描述": "分单号推送；分单号修改推送；分单号抓取资料。",
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "客户合并报关调整",
        "场景描述": (
            "已拦截。例如：123一单、456一单已推送；调整为1245一单、36一单。"
            "需要修改分单号重新推送，原单作废。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "客户合并报关需要拆",
        "场景描述": (
            "例如：123456一单拆分为2单；绑定分单号的时候分为1245一单、36一单推送；"
            "以我们系统推送的为准。"
        ),
        "优先级/备注": "优",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "配仓减少",
        "场景描述": (
            "已拦截。例如：12、34、56三单已推送；调整为56删除。"
            "56的分单号需要沿用，会被报关行拦截；"
            "分单号需要调整重推，前面的全部作废。"
        ),
        "优先级/备注": "",
    },
    {
        "来源": "报关行",
        "场景分类": "分单号推送",
        "子场景": "配仓增加",
        "场景描述": (
            "例如：12、34、56三单已推送；新增78；"
            "78重新推送；其他不影响。"
        ),
        "优先级/备注": "",
    },
]

COLUMNS = ["来源", "场景分类", "子场景", "场景描述", "优先级/备注"]


def style_sheet(ws, row_count: int):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [10, 24, 24, 72, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in range(2, row_count + 2):
        for col in range(1, len(COLUMNS) + 1):
            ws.cell(row=row, column=col).alignment = wrap
        ws.row_dimensions[row].height = 72


def write_table(ws, scenarios):
    for r_idx, item in enumerate(scenarios, start=2):
        for c_idx, key in enumerate(COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=item[key])
    style_sheet(ws, len(scenarios))


def main():
    merged = SYSTEM_SCENARIOS + BROKER_SCENARIOS
    df_merged = pd.DataFrame(merged, columns=COLUMNS)
    df_system = pd.DataFrame(SYSTEM_SCENARIOS, columns=COLUMNS)
    df_broker = pd.DataFrame(BROKER_SCENARIOS, columns=COLUMNS)

    with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
        df_merged.to_excel(writer, sheet_name="场景合并", index=False)
        df_system.to_excel(writer, sheet_name="报关系统场景", index=False)
        df_broker.to_excel(writer, sheet_name="报关行场景", index=False)

    wb = load_workbook(XLSX)

    for name, scenarios in [
        ("场景合并", merged),
        ("报关系统场景", SYSTEM_SCENARIOS),
        ("报关行场景", BROKER_SCENARIOS),
    ]:
        write_table(wb[name], scenarios)

    img_ws = wb.create_sheet("报关行场景原图")
    img_ws["A1"] = "报关行场景（原 PNG 对照）"
    img_ws["A1"].font = Font(bold=True, size=12)
    img = XLImage(str(PNG))
    max_width = 1200
    if img.width > max_width:
        ratio = max_width / img.width
        img.width = max_width
        img.height = int(img.height * ratio)
    img_ws.add_image(img, "A3")
    img_ws.column_dimensions["A"].width = 30

    wb.save(XLSX)

    print(f"完成：已合并 {len(SYSTEM_SCENARIOS)} 条报关系统场景 + {len(BROKER_SCENARIOS)} 条报关行场景")
    print(f"输出文件：{XLSX}")
    print("Sheet：场景合并 / 报关系统场景 / 报关行场景 / 报关行场景原图")


if __name__ == "__main__":
    main()
