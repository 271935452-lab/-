# -*- coding: utf-8 -*-
"""Generate 人员映射 × 规则中心 calculation scenario workbook from 数据映射与逻辑加工表.xls."""
from __future__ import annotations

from pathlib import Path

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
XLS_PATH = ROOT / "数据映射与逻辑加工表.xls"
OUT_PATH = ROOT / "人员映射与规则中心_计算场景全量梳理.xlsx"
SHEET_INDEX = 5  # 人员映射表

YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED_HEAD = PatternFill("solid", fgColor="C00000")
STEP_HEAD = PatternFill("solid", fgColor="DDEBF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
HEADER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
CELL_WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="FFBFBFBF")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_date(wb: xlrd.book.Book, v):
    if v == "" or v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s in ("长期", "—", "-"):
            return s or "长期"
        return s
    try:
        return xlrd.xldate.xldate_as_datetime(v, wb.datemode).strftime("%Y-%m-%d")
    except Exception:
        return str(v)


def infer_scene_kind(acc: str, emp: str, role: str, remark: str, rule_code: str) -> str:
    r = (remark or "").strip()
    rl = (role or "").strip()
    if rule_code and str(rule_code).strip():
        return "指定规则编码锁定规则包"
    if "离职" in r or "止提" in r:
        return "离职/止提特殊口径（映射备注约定）"
    if "主管" in rl and ("组员" in r or "实发" in r or "平均票数" in r):
        return "主管汇总（组员维度聚合后再阶梯）"
    if rl == "同行经理":
        return "同行经理（柜数/计费重基数 + 超额累进）"
    if rl == "同行业务员":
        return "同行业务员（计费重×单价等）"
    if "渠道" in rl:
        return "渠道客服（票数/计费重阶梯等）"
    if rl in ("分公司直客客服", "总部直客客服") or "客服" in rl:
        return "客服类（应收/票数/整柜等多基数）"
    if rl in ("直客业务员", "直客业务员主管"):
        return "直客业务员（净利润阶梯等）"
    return "通用（净利润/固定比例等）"


def default_steps(
    acc: str,
    emp: str,
    branch: str,
    dept: str,
    grp: str,
    role: str,
    mapping_start: str,
    mapping_end: str,
    calc_start: str,
    calc_end: str,
    status: str,
    rule_code: str,
    remark: str,
    kind: str,
) -> list[tuple[str, str]]:
    rc = (rule_code or "").strip()
    lock = (
        f"指定编码「{rc}」→ 规则中心直接锁定对应规则包版本"
        if rc
        else "编码为空 → 按分公司/部门/小组/提成角色向规则中心请求候选规则包并择优命中"
    )

    return [
        (
            "步骤1·主档对齐\n解析系统账号名称+员工姓名 → employee_id / 工号",
            f"命中「{acc}+{emp}」映射主档；组织上下文：分公司={branch}，部门={dept or '—'}，小组={grp or '—'}",
        ),
        (
            "步骤2·映射生效区间\n映射生效起始~失效 ∩ 计薪月是否非空",
            f"区间 {mapping_start} ~ {mapping_end}；计薪月需在区间内方可继续（否则本场景不产生归因上下文）",
        ),
        (
            "步骤3·提成计算区间\n提成计算起止 ∩ 计薪月",
            f"区间 {calc_start} ~ {calc_end}；与映射生效相互独立，财务口径以本区间为准",
        ),
        (
            "步骤4·人员状态门禁\n在职/离职当月止提/停发",
            f"人员状态={status or '—'}；备注摘要：{(remark[:80] + '…') if len(remark) > 80 else remark or '—'}",
        ),
        ("步骤5·规则包定位\n指定规则编码 vs 自动匹配", lock),
        (
            "步骤6·规则包版本筛选\n生效月份、发布状态、组织维度",
            f"候选包按分公司「{branch}」过滤；匹配角色「{role}」对应业务板块；剔除草稿/停用或未覆盖账期版本",
        ),
        (
            "步骤7·规则命中与基数\n条件表达式 → 基数方案 → 计算方式",
            f"场景类型：{kind}；命中后按规则绑定基数方案（毛利/净利/计费重/票数/柜数等）执行阶梯或固定比例",
        ),
        (
            "步骤8·输出\n扣减链 / 上月结转 / 主管拆分 → 试算结果",
            "合并客服提成扣、表外导入扣、师傅扣及规则公式扣减；主管岗可按组员聚合后再套阶梯（见映射备注）",
        ),
    ]


def load_mapping_rows(wb: xlrd.book.Book, sh: xlrd.sheet.Sheet) -> list[dict]:
    rows_out: list[dict] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for r in range(2, sh.nrows):
        remark = str(sh.cell_value(r, 0) or "").strip()
        acc = str(sh.cell_value(r, 1) or "").strip()
        emp = str(sh.cell_value(r, 2) or "").strip()
        if acc == "示例":
            continue
        if not acc and not emp:
            continue
        if not emp:
            continue
        branch = str(sh.cell_value(r, 3) or "").strip()
        dept = str(sh.cell_value(r, 4) or "").strip()
        grp = str(sh.cell_value(r, 5) or "").strip()
        role = str(sh.cell_value(r, 14) or "").strip()
        key = (acc, emp, branch, dept, grp, role)
        if key in seen:
            continue
        seen.add(key)
        rows_out.append(
            {
                "remark": remark,
                "account": acc or "（账号列空）",
                "employee": emp,
                "branch": branch,
                "dept": dept,
                "group": grp,
                "map_start": fmt_date(wb, sh.cell_value(r, 6)),
                "map_end": fmt_date(wb, sh.cell_value(r, 7)),
                "calc_start": fmt_date(wb, sh.cell_value(r, 8)),
                "calc_end": fmt_date(wb, sh.cell_value(r, 9)),
                "hire": fmt_date(wb, sh.cell_value(r, 10)),
                "regular": fmt_date(wb, sh.cell_value(r, 11)),
                "leave": fmt_date(wb, sh.cell_value(r, 12)),
                "emp_status": str(sh.cell_value(r, 13) or "").strip(),
                "role": role,
                "rule_code": str(sh.cell_value(r, 15) or "").strip(),
            }
        )
    return rows_out


def write_scenario_block(ws, start_row: int, idx: int, rec: dict) -> int:
    """Write one scenario block starting at start_row; return next free row."""
    kind = infer_scene_kind(
        rec["account"],
        rec["employee"],
        rec["role"],
        rec["remark"],
        rec["rule_code"],
    )
    params = (
        f"计薪月=T（与底表归属月一致）；系统账号={rec['account']}；员工姓名={rec['employee']}；"
        f"分公司={rec['branch']}；部门={rec['dept'] or '—'}；小组={rec['group'] or '—'}；"
        f"提成岗位={rec['role']}；映射生效={rec['map_start']}~{rec['map_end']}；"
        f"提成计算={rec['calc_start']}~{rec['calc_end']}；人员状态={rec['emp_status'] or '—'}；"
        f"指定规则编码={'「'+rec['rule_code']+'」' if rec['rule_code'] else '空（自动匹配）'}；"
        f"场景归类={kind}"
    )
    if rec["remark"]:
        params += f"；映射备注摘要={' '.join(rec['remark'].split())[:200]}"

    steps = default_steps(
        rec["account"],
        rec["employee"],
        rec["branch"],
        rec["dept"],
        rec["group"],
        rec["role"],
        rec["map_start"],
        rec["map_end"],
        rec["calc_start"],
        rec["calc_end"],
        rec["emp_status"],
        rec["rule_code"],
        rec["remark"],
        kind,
    )

    r_title = start_row
    r_param = start_row + 2
    r_head = start_row + 4
    r_data = start_row + 5
    r_sum = start_row + 7

    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=13)
    c = ws.cell(row=r_title, column=1, value=f"场景{idx:03d}·{rec['account']}+{rec['employee']}·{rec['role']}")
    c.font = TITLE_FONT
    c.fill = YELLOW
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=r_param, start_column=1, end_row=r_param, end_column=13)
    p = ws.cell(row=r_param, column=1, value=params)
    p.alignment = CELL_WRAP

    headers_raw = ("员工昵称(系统账号)", "员工姓名", "所属分公司", "提成岗位")
    for i, h in enumerate(headers_raw, start=1):
        cell = ws.cell(row=r_head, column=i, value=h)
        cell.fill = RED_HEAD
        cell.font = WHITE_BOLD
        cell.alignment = HEADER_WRAP
        cell.border = thin_border()

    col_step = 6
    for j, (hdr, _) in enumerate(steps):
        cell = ws.cell(row=r_head, column=col_step + j, value=hdr)
        cell.fill = STEP_HEAD
        cell.font = Font(bold=True, size=10)
        cell.alignment = HEADER_WRAP
        cell.border = thin_border()

    ws.cell(row=r_data, column=1, value=rec["account"]).alignment = CELL_WRAP
    ws.cell(row=r_data, column=2, value=rec["employee"]).alignment = CELL_WRAP
    ws.cell(row=r_data, column=3, value=rec["branch"]).alignment = CELL_WRAP
    ws.cell(row=r_data, column=4, value=rec["role"]).alignment = CELL_WRAP
    for j, (_, val) in enumerate(steps):
        c2 = ws.cell(row=r_data, column=col_step + j, value=val)
        c2.alignment = CELL_WRAP
        c2.border = thin_border()

    ws.merge_cells(start_row=r_sum, start_column=1, end_row=r_sum, end_column=4)
    sum_label = ws.cell(row=r_sum, column=1, value="本场景结论（示意）")
    sum_label.font = Font(bold=True)
    sum_label.fill = PatternFill("solid", fgColor="FCE4D6")
    ws.merge_cells(start_row=r_sum, start_column=6, end_row=r_sum, end_column=13)
    conclusion = (
        f"归并为「{kind}」。规则中心在步骤5–7完成规则包/规则/基数方案绑定；"
        f"若映射备注中含专项口径（如票数均值、整柜识别、产品维度），在规则条件表达式与销售产品中落地。"
    )
    sc = ws.cell(row=r_sum, column=6, value=conclusion)
    sc.alignment = CELL_WRAP
    sc.fill = PatternFill("solid", fgColor="FCE4D6")

    for row in (r_data, r_sum):
        for col in range(1, 14):
            ws.cell(row=row, column=col).border = thin_border()

    return start_row + 10


def main():
    wb_rd = xlrd.open_workbook(str(XLS_PATH))
    sh = wb_rd.sheet_by_index(SHEET_INDEX)
    records = load_mapping_rows(wb_rd, sh)

    out_wb = Workbook()
    idx_sheet = out_wb.active
    idx_sheet.title = "索引"

    headers_idx = ["序号", "系统账号名称", "员工姓名", "分公司", "提成岗位", "场景归类", "工作表名"]
    for c, h in enumerate(headers_idx, 1):
        idx_sheet.cell(row=1, column=c, value=h).font = Font(bold=True)

    for i, rec in enumerate(records, start=1):
        sheet_name = f"S{i:03d}"
        ws_new = out_wb.create_sheet(title=sheet_name)
        write_scenario_block(ws_new, start_row=1, idx=i, rec=rec)

        row = i + 1
        idx_sheet.cell(row=row, column=1, value=i)
        idx_sheet.cell(row=row, column=2, value=rec["account"])
        idx_sheet.cell(row=row, column=3, value=rec["employee"])
        idx_sheet.cell(row=row, column=4, value=rec["branch"])
        idx_sheet.cell(row=row, column=5, value=rec["role"])
        idx_sheet.cell(row=row, column=6, value=infer_scene_kind(
            rec["account"], rec["employee"], rec["role"], rec["remark"], rec["rule_code"]
        ))
        idx_sheet.cell(row=row, column=7, value=sheet_name)

    for ws in out_wb.worksheets:
        if ws.title == "索引":
            ws.column_dimensions["A"].width = 6
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 18
            ws.column_dimensions["F"].width = 36
            ws.column_dimensions["G"].width = 10
            continue
        for col in range(1, 14):
            ws.column_dimensions[get_column_letter(col)].width = 18 if col >= 6 else 14

    out_wb.save(OUT_PATH)
    print(f"Wrote {len(records)} scenarios -> {OUT_PATH}")


if __name__ == "__main__":
    main()
