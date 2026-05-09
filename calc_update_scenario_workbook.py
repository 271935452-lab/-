# -*- coding: utf-8 -*-
"""
将「测试数据_运单底表_逻辑加工表人员映射」「测试数据_手工万能调整池_逻辑加工表人员映射」
中的真实数值汇入提成试算，并写回「人员映射与规则中心_计算场景全量梳理.xlsx」。
口径说明写在各场景表参数区；系数对齐原型规则编码（演示账期 2026-03）。
"""
from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).resolve().parent
XLS_MAPPING = ROOT / "数据映射与逻辑加工表.xls"
WB_DATA = ROOT / "财务资料" / "测试数据_运单底表_逻辑加工表人员映射.xlsx"
ADJ_DATA = ROOT / "财务资料" / "测试数据_手工万能调整池_逻辑加工表人员映射.xlsx"
OUT_BOOK = ROOT / "人员映射与规则中心_计算场景全量梳理.xlsx"
OUT_BOOK_SAVE = ROOT / "人员映射与规则中心_计算场景全量梳理_含底表测算.xlsx"
MAPPING_SHEET_IDX = 5

YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED_HEAD = PatternFill("solid", fgColor="C00000")
STEP_HEAD = PatternFill("solid", fgColor="DDEBF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUMMARY_FILL = PatternFill("solid", fgColor="FCE4D6")
CELL_WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="FFBFBFBF")


def thin_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_date(wb: xlrd.book.Book, v):
    if v == "" or v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s in ("长期", "—", "-"):
            return s or "长期"
        return s
    try:
        return xlrd.xldate.xldate_as_datetime(v, wb.datemode).strftime("%Y-%m-%d")
    except Exception:
        return str(v)


def load_personnel_rows() -> list[dict]:
    wb = xlrd.open_workbook(str(XLS_MAPPING))
    sh = wb.sheet_by_index(MAPPING_SHEET_IDX)
    rows_out: list[dict] = []
    seen: set[tuple[str, str, str, str, str, str]] = set()
    for r in range(2, sh.nrows):
        remark = str(sh.cell_value(r, 0) or "").strip()
        acc = str(sh.cell_value(r, 1) or "").strip()
        emp = str(sh.cell_value(r, 2) or "").strip()
        if acc == "示例":
            continue
        if not acc and not emp:
            continue
        if not emp:
            continue
        branch = str(sh.cell_value(r, 3) or "").strip()
        dept = str(sh.cell_value(r, 4) or "").strip()
        grp = str(sh.cell_value(r, 5) or "").strip()
        role = str(sh.cell_value(r, 14) or "").strip()
        key = (acc, emp, branch, dept, grp, role)
        if key in seen:
            continue
        seen.add(key)
        rows_out.append(
            {
                "remark": remark,
                "account": acc or "（账号列空）",
                "employee": emp,
                "branch": branch,
                "dept": dept,
                "group": grp,
                "map_start": fmt_date(wb, sh.cell_value(r, 6)),
                "map_end": fmt_date(wb, sh.cell_value(r, 7)),
                "calc_start": fmt_date(wb, sh.cell_value(r, 8)),
                "calc_end": fmt_date(wb, sh.cell_value(r, 9)),
                "emp_status": str(sh.cell_value(r, 13) or "").strip(),
                "role": role,
                "rule_code": str(sh.cell_value(r, 15) or "").strip(),
            }
        )
    return rows_out


def read_matrix(path: Path, sheet_index: int = 0) -> tuple[list[str], list[list[Any]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    rows = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        rows.append(row)
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return headers, rows[1:]


def hidx(headers: list[str], *names: str) -> int:
    for n in names:
        for i, h in enumerate(headers):
            if h.replace("*", "").strip() == n.replace("*", "").strip():
                return i
    raise KeyError(names)


def to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


@dataclass
class AdjLine:
    account: str
    employee: str
    month: str
    waybill: str
    fee_cat: str  # 扣毛利 / 扣提成
    adj_type: str  # 加项 / 减项
    amount: float
    remark: str = ""


@dataclass
class WaybillRow:
    wb_no: str
    month: str
    bw: float
    containers: float
    tickets: float
    sales_acc: str
    cs_acc: str
    receivable: float
    gross_profit: float
    calc_net: float
    rule: str
    cs_has_sales: str


def parse_adjustments(headers: list[str], body: list) -> list[AdjLine]:
    hi = {h.replace("*", "").strip(): i for i, h in enumerate(headers)}
    lines = []
    for row in body:
        if not row or all(x is None or str(x).strip() == "" for x in row):
            continue
        acc = str(row[hi["系统账户名称"]] or "").strip()
        emp = str(row[hi["员工姓名"]] or "").strip()
        if not acc:
            continue
        month = str(row[hi["月度"]] or "").strip()
        wb = str(row[hi["运单号"]] or "").strip()
        fee_cat = str(row[hi["费用性质"]] or "").strip()
        adj_type = str(row[hi["调整类型"]] or "").strip()
        amt = to_float(row[hi["成本金额"]])
        rm = str(row[hi["备注"]] if "备注" in hi else "") or ""
        lines.append(AdjLine(acc, emp, month, wb, fee_cat, adj_type, amt, rm))
    return lines


def parse_waybills(headers: list[str], body: list) -> list[WaybillRow]:
    idx = {h.strip(): i for i, h in enumerate(headers)}
    out = []
    for row in body:
        if not row or not row[idx["运单号"]]:
            continue
        out.append(
            WaybillRow(
                wb_no=str(row[idx["运单号"]]).strip(),
                month=str(row[idx["到货月份"]]).strip(),
                bw=to_float(row[idx["计费重(kg)"]]),
                containers=to_float(row[idx["整柜数"]]),
                tickets=to_float(row[idx["票数"]]),
                sales_acc=str(row[idx["业务员系统账号"]] or "").strip(),
                cs_acc=str(row[idx["客服系统账号"]] or "").strip(),
                receivable=to_float(row[idx["应收"]]),
                gross_profit=to_float(row[idx["毛利"]]),
                calc_net=to_float(row[idx["计算净利润"]]),
                rule=str(row[idx["规则编码_演示"]] or "").strip(),
                cs_has_sales=str(row[idx["客服是否有业务员"]] or "").strip(),
            )
        )
    return out


def first_emp_for_account(personnel: list[dict]) -> dict[str, str]:
    m: dict[str, str] = {}
    for rec in personnel:
        a = rec["account"]
        if a not in m:
            m[a] = rec["employee"]
    return m


def adj_sums_for(
    adjustments: list[AdjLine],
    account: str,
    waybill: str | None,
    month: str,
) -> tuple[float, float]:
    """Returns (扣毛利加项合计影响净利基数, 扣提成减项合计影响最终提成)."""
    gm = 0.0
    ct = 0.0
    mn = month.replace("/", "-")[:7]
    for a in adjustments:
        if a.account != account or a.month.replace("/", "-")[:7] != mn:
            continue
        if waybill is not None:
            aw = str(a.waybill or "").strip()
            if aw and aw != waybill:
                continue
            if not aw:
                continue
        if a.fee_cat == "扣毛利" and a.adj_type == "加项":
            gm += a.amount
        elif a.fee_cat == "扣提成" and a.adj_type == "减项":
            ct += a.amount
    return gm, ct


def monthly_orphan_adjustments(adjustments: list[AdjLine], account: str, month: str) -> tuple[float, float]:
    """Adjustments with empty waybill for this account-month."""
    gm = 0.0
    ct = 0.0
    mn = month.replace("/", "-")[:7]
    for a in adjustments:
        if a.account != account:
            continue
        if a.month.replace("/", "-")[:7] != mn:
            continue
        if str(a.waybill or "").strip():
            continue
        if a.fee_cat == "扣毛利" and a.adj_type == "加项":
            gm += a.amount
        elif a.fee_cat == "扣提成" and a.adj_type == "减项":
            ct += a.amount
    return gm, ct


def round_money(x: float) -> float:
    return round(x + 1e-9, 2)


def calc_pair(
    w: WaybillRow,
    adjustments: list[AdjLine],
    role: str,
    account: str,
    emp: str,
    month: str,
) -> dict[str, Any] | None:
    """Returns commission line for one role on one waybill (sales or cs)."""
    gm_wb, ct_wb = adj_sums_for(adjustments, account, w.wb_no, month)
    gm_orphan, ct_orphan = 0.0, 0.0

    net = w.calc_net
    recv = w.receivable
    bw = w.bw
    tk = w.tickets
    ctr = w.containers
    rule = w.rule

    if role == "业务员":
        net_adj = net - gm_wb
        pre = 0.0
        detail = ""

        if rule == "DIRECT_SALES_GP_TIER":
            pct = 0.385
            pre = net_adj * pct
            detail = f"基数=计算净利润{net:.2f}−手工毛利扣{gm_wb:.2f}={net_adj:.2f}；×{pct:.3f}(演示阶梯中档)"
        elif rule == "CHANNEL_MANAGER_SETTLE":
            pct = 0.062
            ticket_part = tk * 1.0
            pre = net_adj * pct + ticket_part
            detail = f"基数={net_adj:.2f}×{pct}+票数×1({ticket_part:.2f})；渠道业务员毛利+票数奖励（演示）"
        elif rule == "PEER_SALES_VOLUME_TIER":
            if str(w.cs_has_sales).strip() == "有业务员":
                vol = bw * 0.28 + net_adj * 0.12
                pre = vol
                detail = f"同行多规则并行演示：计费重×0.28+净利×0.12→{pre:.2f}"
            else:
                vol = bw * 0.35 + net_adj * 0.08
                pre = vol
                detail = f"空运口径：计费重×0.35+净利×0.08→{pre:.2f}"
        elif rule == "CS_NO_SALES_RATE":
            pre = bw * 0.42 + net_adj * 0.06
            detail = f"无业务员客服场景·业务员侧：计费重×0.42+净利×0.06→{pre:.2f}"
        elif rule == "CS_WITH_SALES_RATE":
            pre = net_adj * 0.41
            detail = f"整柜+应收组合演示：净利基数×0.41→{pre:.2f}"
        elif rule == "PEER_MANAGER_CONTAINER_TIER":
            pre = net_adj * 0.085 + ctr * 2150
            detail = f"经理柜数阶梯演示：净利×0.085+整柜数×2150→{pre:.2f}"
        elif rule == "PEER_UNDER300_NO_COMM":
            tons = bw / 1000.0
            if tons <= 300:
                pre = 0.0
                detail = f"计费重折合{tons:.1f}吨≤300→按映射备注不发提成"
            else:
                pre = net_adj * 0.095 + ctr * 1800
                detail = f"超300吨部分：净利×0.095+柜×1800→{pre:.2f}"
        elif rule == "DIRECT_MANAGER_REVENUE_RATE":
            perf = 0.03
            pre = recv * 0.0015 * (1 - perf)
            detail = f"主管组员应收演示：应收×0.15%×(1−绩效{perf:.0%})→{pre:.2f}"
        elif rule == "DIRECT_KA_BONUS":
            pct = 0.372
            pre = net_adj * pct
            detail = f"KA客户专项演示：净利基数×{pct:.3f}→{pre:.2f}"
        else:
            return None

        final = max(0.0, pre - ct_wb)
        return {
            "person_role": "业务员",
            "account": account,
            "employee": emp,
            "waybill": w.wb_no,
            "rule": rule,
            "step6": detail,
            "step7_pre": round_money(pre),
            "step8": round_money(final),
            "gm_wb": gm_wb,
            "ct_wb": ct_wb,
            "recv": recv,
            "net": net,
            "bw": bw,
            "tk": tk,
            "ctr": ctr,
        }

    if role == "客服":
        net_adj = net
        pre = 0.0
        detail = ""

        if rule == "CHANNEL_MANAGER_SETTLE" and account == w.cs_acc:
            pre = tk * 1.2 + recv * 0.0018
            detail = f"渠道客服演示：票数×1.2+应收×0.18%→{pre:.2f}"
        elif rule == "PEER_SALES_VOLUME_TIER" and str(w.cs_has_sales).strip() == "有业务员" and account == w.cs_acc:
            pre = recv * 0.001 + tk * 0.35
            detail = f"有业务员客服：应收×0.1%+票数×0.35→{pre:.2f}"
        elif rule == "CS_NO_SALES_RATE" and account == w.cs_acc:
            pre = recv * 0.0028 + tk * 0.5
            detail = f"无业务员客服：应收×0.28%+票数×0.5→{pre:.2f}"
        elif rule == "CS_WITH_SALES_RATE" and account == w.cs_acc:
            pre = ctr * 920 + recv * 0.00075
            detail = f"分公司直客客服整柜演示：柜×920+应收×0.075%→{pre:.2f}"
        elif rule == "DIRECT_KA_BONUS" and account == w.cs_acc:
            base = recv * 0.012 - gm_wb * 0.6
            pre = max(0.0, base)
            detail = f"关联客服演示：应收×1.2%−毛利加项×60%摊回→{pre:.2f}"
        else:
            return None

        final = max(0.0, pre - ct_wb)
        return {
            "person_role": "客服",
            "account": account,
            "employee": emp,
            "waybill": w.wb_no,
            "rule": rule,
            "step6": detail,
            "step7_pre": round_money(pre),
            "step8": round_money(final),
            "gm_wb": gm_wb,
            "ct_wb": ct_wb,
            "recv": recv,
            "net": net,
            "bw": bw,
            "tk": tk,
            "ctr": ctr,
        }

    return None


def build_commission_lines(
    waybills: list[WaybillRow],
    adjustments: list[AdjLine],
    acct_emp: dict[str, str],
    month: str,
) -> dict[tuple[str, str], list[dict]]:
    bucket: dict[tuple[str, str], list[dict]] = defaultdict(list)
    mn = month.replace("/", "-")[:7]

    for w in waybills:
        if w.month.replace("/", "-")[:7] != mn:
            continue
        sa = w.sales_acc
        emp_s = acct_emp.get(sa, "")
        line_s = calc_pair(w, adjustments, "业务员", sa, emp_s, month)
        if line_s:
            bucket[(sa, emp_s)].append(line_s)

        ca = w.cs_acc
        if ca:
            emp_c = acct_emp.get(ca, "")
            line_c = calc_pair(w, adjustments, "客服", ca, emp_c, month)
            if line_c:
                bucket[(ca, emp_c)].append(line_c)

    # 万能调整池·无运单号行：单独演示毛利加项/提成减项对月度应发的折算影响
    orphan_accounts = {
        a.account
        for a in adjustments
        if not str(a.waybill or "").strip() and a.month.replace("/", "-")[:7] == mn
    }
    for acc in orphan_accounts:
        emp = acct_emp.get(acc, "")
        gm_o, ct_o = monthly_orphan_adjustments(adjustments, acc, month)
        if gm_o == 0 and ct_o == 0:
            continue
        key = (acc, emp)
        impact_g = round_money(-gm_o * 0.385)
        impact_ct = round_money(-ct_o)
        bucket[key].append(
            {
                "person_role": "月度手工池",
                "account": acc,
                "employee": emp,
                "waybill": "—",
                "rule": "万能调整池·无运单号",
                "step6": (
                    f"无运单号记录：毛利加项合计{gm_o:.2f}→演示按净利润提成系数0.385折算影响{impact_g:.2f}；"
                    f"提成减项合计{ct_o:.2f}"
                ),
                "step7_pre": impact_g,
                "step8": round_money(impact_g + impact_ct),
                "gm_wb": gm_o,
                "ct_wb": ct_o,
                "recv": 0,
                "net": 0,
                "bw": 0,
                "tk": 0,
                "ctr": 0,
            }
        )

    return bucket


def rebuild_sheet(ws, rec: dict, idx: int, lines: list[dict] | None, month: str):
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))

    for row in ws.iter_rows():
        for c in row:
            if hasattr(c, "value"):
                c.value = None

    acc = rec["account"]
    emp = rec["employee"]
    title = f"场景{idx:03d}·{acc}+{emp}·{rec['role']}"
    param_extra = (
        f"【真实测算】账期 {month}；运单底表：{WB_DATA.name}；万能调整池：{ADJ_DATA.name}。"
        f"步骤7/8为按列公式演示值，可与底表「计算净利润/应收/票数」手工复核。"
    )

    start_row = 1
    r_title = start_row
    r_param = start_row + 2
    r_head = start_row + 4

    headers_raw = ("员工昵称(系统账号)", "员工姓名", "运单号", "角色", "应收", "计算净利润", "计费重kg", "票数")
    col_step = len(headers_raw) + 2
    step_headers = [
        "步骤1·主档对齐",
        "步骤2·映射∩账期",
        "步骤3·提成区间",
        "步骤4·状态门禁",
        "步骤5·命中规则",
        "步骤6·基数加工",
        "步骤7·提成试算",
        "步骤8·扣减后应发",
    ]
    last_col = col_step + len(step_headers) - 1

    ws.merge_cells(start_row=r_title, start_column=1, end_row=r_title, end_column=last_col)
    c = ws.cell(row=r_title, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = YELLOW

    base_params = (
        f"计薪月={month}；分公司={rec['branch']}；部门={rec['dept'] or '—'}；小组={rec['group'] or '—'}；"
        f"提成岗位={rec['role']}；映射生效={rec['map_start']}~{rec['map_end']}；提成计算={rec['calc_start']}~{rec['calc_end']}。"
    )
    ws.merge_cells(start_row=r_param, start_column=1, end_row=r_param, end_column=last_col)
    ws.cell(row=r_param, column=1, value=base_params + param_extra).alignment = CELL_WRAP

    for i, h in enumerate(headers_raw, start=1):
        cell = ws.cell(row=r_head, column=i, value=h)
        cell.fill = RED_HEAD
        cell.font = WHITE_BOLD
        cell.alignment = HEADER_WRAP
        cell.border = thin_border()

    for j, h in enumerate(step_headers):
        cell = ws.cell(row=r_head, column=col_step + j, value=h)
        cell.fill = STEP_HEAD
        cell.font = Font(bold=True, size=10)
        cell.alignment = HEADER_WRAP
        cell.border = thin_border()

    if not lines:
        r_blank = r_head + 1
        ws.cell(row=r_blank, column=1, value=acc)
        ws.cell(row=r_blank, column=2, value=emp)
        ws.merge_cells(start_row=r_blank, start_column=col_step, end_row=r_blank, end_column=last_col)
        ws.cell(
            row=r_blank,
            column=col_step,
            value="本期测试底表未出现该系统账号的分摊行；下方仍为映射×规则中心链路示意（无数值测算）。",
        ).alignment = CELL_WRAP
        return

    total_final = 0.0
    r = r_head + 1
    for ln in lines:
        ws.cell(row=r, column=1, value=ln["account"])
        ws.cell(row=r, column=2, value=ln["employee"])
        ws.cell(row=r, column=3, value=ln["waybill"])
        ws.cell(row=r, column=4, value=ln["person_role"])
        ws.cell(row=r, column=5, value=ln["recv"])
        ws.cell(row=r, column=6, value=ln["net"])
        ws.cell(row=r, column=7, value=ln["bw"])
        ws.cell(row=r, column=8, value=ln["tk"])

        steps_val = [
            f"命中映射「{acc}+{emp}」",
            "映射生效∩2026-03：通过（示意）",
            "提成计算区间：覆盖（示意）",
            f"{rec['emp_status'] or '在职'}·发放正常（示意）",
            ln["rule"],
            ln["step6"],
            ln["step7_pre"],
            ln["step8"],
        ]
        for j, v in enumerate(steps_val):
            cell = ws.cell(row=r, column=col_step + j, value=v)
            cell.alignment = CELL_WRAP
            cell.border = thin_border()

        for col in range(1, 9):
            ws.cell(row=r, column=col).alignment = CELL_WRAP
            ws.cell(row=r, column=col).border = thin_border()

        total_final += ln["step8"]
        r += 1

    ws.cell(row=r, column=1, value="合计（演示应发提成）").font = Font(bold=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    ws.merge_cells(start_row=r, start_column=col_step, end_row=r, end_column=last_col)
    ws.cell(row=r, column=col_step, value=f"Σ步骤8 = {round_money(total_final)} CNY（同一账号多运单行求和）").font = Font(bold=True)
    for c in range(1, last_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = SUMMARY_FILL
        cell.border = thin_border()


def write_summary_sheet(wb, bucket: dict[tuple[str, str], list[dict]], month: str):
    name = f"提成汇总_{month.replace('-', '')}"
    if name in wb.sheetnames:
        ws_old = wb[name]
        wb.remove(ws_old)
    ws = wb.create_sheet(name, 1)
    ws.cell(row=1, column=1, value=f"账期 {month} · 基于测试底表与万能调整池汇总").font = Font(bold=True)
    hdr = ["系统账号", "员工姓名", "底表分摊行数", "应发提成合计(CNY)", "规则编码（逐笔见各场景表）"]
    for c, h in enumerate(hdr, 1):
        ws.cell(row=3, column=c, value=h).font = Font(bold=True)

    row = 4
    items = sorted(bucket.items(), key=lambda x: (-sum(l["step8"] for l in x[1]), x[0][0]))
    for (acc, emp), lines in items:
        if not lines:
            continue
        ws.cell(row=row, column=1, value=acc)
        ws.cell(row=row, column=2, value=emp)
        ws.cell(row=row, column=3, value=len(lines))
        total = round_money(sum(l["step8"] for l in lines))
        ws.cell(row=row, column=4, value=total)
        rules = "；".join(sorted({l["rule"] for l in lines}))
        ws.cell(row=row, column=5, value=rules)
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 48


def main():
    month = "2026-03"
    personnel = load_personnel_rows()
    acct_emp = first_emp_for_account(personnel)

    wh, wb_body = read_matrix(WB_DATA)
    ah, adj_body = read_matrix(ADJ_DATA)
    waybills = parse_waybills(wh, wb_body)
    adjustments = parse_adjustments(ah, adj_body)

    bucket = build_commission_lines(waybills, adjustments, acct_emp, month)

    book = load_workbook(OUT_BOOK)

    write_summary_sheet(book, bucket, month)

    for i, rec in enumerate(personnel, start=1):
        sheet_name = f"S{i:03d}"
        if sheet_name not in book.sheetnames:
            continue
        ws = book[sheet_name]
        key = (rec["account"], rec["employee"])
        lines = bucket.get(key)
        rebuild_sheet(ws, rec, i, lines, month)

    idx = book["索引"]
    idx.cell(row=1, column=8, value="底表应发提成(CNY)")
    idx.cell(row=1, column=9, value="有无测试分摊")
    for r in range(2, idx.max_row + 1):
        acc = idx.cell(row=r, column=2).value
        emp = idx.cell(row=r, column=3).value
        if not acc:
            continue
        lines = bucket.get((str(acc), str(emp)), [])
        total = round_money(sum(l["step8"] for l in lines))
        idx.cell(row=r, column=8, value=total if lines else "—")
        idx.cell(row=r, column=9, value="有" if lines else "无")

    try:
        book.save(OUT_BOOK)
        saved = OUT_BOOK
    except PermissionError:
        book.save(OUT_BOOK_SAVE)
        saved = OUT_BOOK_SAVE
    print(f"Updated {saved} ({len(personnel)} scenarios, {len(bucket)} commission keys)")


if __name__ == "__main__":
    main()
