# -*- coding: utf-8 -*-
"""从 待排期需求.xlsx 的「待排期」sheet 生成迭代排期 Excel"""
from collections import Counter
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = r"d:\Cursor\头程项目\其他\待排期需求.xlsx"
OUT = r"d:\Cursor\头程项目\其他\待排期迭代计划.xlsx"
SPRINT_START = datetime(2026, 5, 26)

# 高效团队：单任务工时约为常规估算的 40%~55%
BE_CAP = 15  # 3后端 × 5天
FE_CAP = 5   # 1前端 × 5天
TE_CAP = 7.5 # 1.5测试 × 5天

PRIORITY_MAP = {"紧急": 4, "高": 3, "中": 2, "低": 1, "": 2}
COMPLEXITY_MAP = {"简单": 1, "中等": 2, "复杂": 3, "": 2}
GROUP_NAMES = {
    1: "1-高优+简单",
    2: "2-全部简单",
    3: "3-高优+中等",
    4: "4-高优+复杂",
    5: "5-中优+中等",
    6: "6-中优+复杂",
    7: "7-低优+中等",
    8: "8-低优+复杂",
    9: "9-待补全",
}

# (后端人日, 前端人日, 测试人日) — 高效团队口径
EFFORT = {
    ("简单", False): (0.5, 0.25, 0.25),
    ("简单", True): (0.5, 0.5, 0.25),
    ("中等", False): (1.5, 0.5, 0.5),
    ("中等", True): (1.5, 1, 0.5),
    ("复杂", False): (4, 1, 1),
    ("复杂", True): (5, 2, 1),
}


def sort_group(p, c):
    if c == "简单" and p in ("高", "紧急"):
        return 1
    if c == "简单":
        return 2
    if p in ("高", "紧急") and c == "中等":
        return 3
    if p in ("高", "紧急") and c == "复杂":
        return 4
    if p == "中" and c == "中等":
        return 5
    if p == "中" and c == "复杂":
        return 6
    if p == "低" and c == "中等":
        return 7
    if p == "低" and c == "复杂":
        return 8
    return 9


def needs_frontend(module, desc, dept):
    kws = (
        "APP", "客户端", "小程序", "页面", "按钮", "弹窗", "报表", "筛选",
        "导出", "上传", "拖拽", "标签", "颜色", "排版", "轨迹", "工单",
    )
    text = (module or "") + (desc or "")
    if "客服" in dept or "营销" in dept:
        if "对接" not in (desc or "")[:30]:
            return True
    return any(k in text for k in kws)


def read_pending_items():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["待排期"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    current_dept = ""
    items = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        desc = str(vals[idx["需求描述"]] or "").strip()
        if desc.startswith("共") and "个" in desc:
            current_dept = str(vals[idx["提出时间"]] or "").replace("部门", "")
            continue

        propose_time = vals[idx["提出时间"]]
        module = str(vals[idx["功能名称"]] or "").strip()
        rid = str(vals[idx["需求编号"]] or "").strip()
        if not desc and not module and not rid:
            continue

        p = str(vals[idx["优先级"]] or "").strip() or "中"
        c = str(vals[idx["复杂程度"]] or "").strip() or "中等"
        dept = current_dept or str(vals[idx["归纳部门"]] or "").replace("部门", "")
        status = str(vals[idx["需求状态"]] or "").strip()
        proposer = str(vals[idx["需求提出人"]] or "").strip()
        owner = str(vals[idx["需求负责人"]] or "").strip()
        fe = needs_frontend(module, desc, dept)
        be, fr, te = EFFORT.get((c, fe), EFFORT.get((c, False), (1.5, 0.5, 0.5)))

        items.append(
            {
                "提出时间": propose_time,
                "需求状态": status,
                "需求提出人": proposer,
                "功能名称": module,
                "归纳部门": dept,
                "需求描述": desc.replace("\n", " "),
                "优先级": p,
                "复杂程度": c,
                "需求编号": rid,
                "需求负责人": owner,
                "排序组": sort_group(p, c),
                "排序组说明": GROUP_NAMES[sort_group(p, c)],
                "后端人日": be,
                "前端人日": fr,
                "测试人日": te,
                "需前端": "是" if fe else "否",
            }
        )

    items.sort(
        key=lambda x: (
            x["排序组"],
            -PRIORITY_MAP.get(x["优先级"], 2),
            x["复杂程度"] == "简单",
            x["需求编号"] or "9999",
        )
    )
    return items


def _assign_sprint_meta(it, week_num):
    start = SPRINT_START + timedelta(weeks=week_num - 1)
    end = start + timedelta(days=4)
    it["迭代序号"] = week_num
    it["迭代周期"] = f"{start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}"
    it["迭代名称"] = f"Sprint {week_num:02d}"


def schedule_weeks(items, be_cap=BE_CAP, fe_cap=FE_CAP):
    """Sprint 01 固定纳入全部「紧急」需求；其余从 Sprint 02 起按产能排期。"""
    urgent = [it for it in items if it["优先级"] == "紧急"]
    rest = [it for it in items if it["优先级"] != "紧急"]

    urgent.sort(
        key=lambda x: (
            COMPLEXITY_MAP.get(x["复杂程度"], 2),
            x["需求编号"] or "9999",
        )
    )

    weeks = []

    if urgent:
        s01 = {"week": 1, "items": [], "be": 0.0, "fe": 0.0, "urgent_only": True}
        for it in urgent:
            _assign_sprint_meta(it, 1)
            s01["items"].append(it)
            s01["be"] += it["后端人日"]
            s01["fe"] += it["前端人日"]
        weeks.append(s01)

    w = 2 if urgent else 1
    be_used = fe_used = 0
    current = {"week": w, "items": [], "be": 0.0, "fe": 0.0, "urgent_only": False}

    for it in rest:
        if current["items"] and (
            be_used + it["后端人日"] > be_cap or fe_used + it["前端人日"] > fe_cap
        ):
            weeks.append(current)
            w += 1
            current = {"week": w, "items": [], "be": 0.0, "fe": 0.0, "urgent_only": False}
            be_used = fe_used = 0

        _assign_sprint_meta(it, current["week"])
        current["items"].append(it)
        current["be"] += it["后端人日"]
        current["fe"] += it["前端人日"]
        be_used += it["后端人日"]
        fe_used += it["前端人日"]

    if current["items"]:
        weeks.append(current)
    return weeks


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def write_excel(items, weeks):
    wb = Workbook()

    # Sheet1: 迭代排期明细
    ws1 = wb.active
    ws1.title = "迭代排期明细"
    headers1 = [
        "迭代序号", "迭代名称", "迭代周期", "排序组", "排序组说明",
        "需求编号", "提出时间", "需求状态", "需求提出人", "功能名称", "归纳部门",
        "需求描述", "优先级", "复杂程度", "需前端",
        "后端人日", "前端人日", "测试人日", "需求负责人",
    ]
    ws1.append(headers1)
    style_header(ws1, 1, len(headers1))

    group_fills = {
        1: "E2EFDA", 2: "E2EFDA", 3: "FFF2CC", 4: "FCE4D6",
        5: "DDEBF7", 6: "DDEBF7", 7: "EDEDED", 8: "EDEDED", 9: "F4CCCC",
    }
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for it in sorted(items, key=lambda x: (x["迭代序号"], x["排序组"], x["需求编号"] or "9999")):
        ws1.append([
            it["迭代序号"], it["迭代名称"], it["迭代周期"],
            it["排序组"], it["排序组说明"],
            it["需求编号"], it["提出时间"], it["需求状态"], it["需求提出人"],
            it["功能名称"], it["归纳部门"], it["需求描述"],
            it["优先级"], it["复杂程度"], it["需前端"],
            it["后端人日"], it["前端人日"], it["测试人日"], it["需求负责人"],
        ])
        r = ws1.max_row
        fill_color = group_fills.get(it["排序组"], "FFFFFF")
        for c in range(1, len(headers1) + 1):
            cell = ws1.cell(row=r, column=c)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill_color)

    widths1 = [8, 10, 22, 6, 14, 8, 12, 10, 12, 14, 10, 50, 8, 8, 8, 8, 8, 8, 12]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # Sheet2: 迭代汇总
    ws2 = wb.create_sheet("迭代汇总")
    headers2 = [
        "迭代序号", "迭代名称", "迭代周期", "需求数",
        "后端人日", "前端人日", "测试人日",
        f"后端负荷({BE_CAP})", f"前端负荷({FE_CAP})", f"测试负荷({TE_CAP})",
        "主要排序组", "本迭代需求编号",
    ]
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))

    for wk in weeks:
        start = SPRINT_START + timedelta(weeks=wk["week"] - 1)
        end = start + timedelta(days=4)
        period = f"{start.strftime('%Y-%m-%d')} ~ {end.strftime('%Y-%m-%d')}"
        te = sum(it["测试人日"] for it in wk["items"])
        if wk.get("urgent_only"):
            main_label = "0-紧急专项(S01)"
        else:
            groups = Counter(it["排序组"] for it in wk["items"])
            main_group = groups.most_common(1)[0][0]
            main_label = GROUP_NAMES.get(main_group, "")
        ids = ", ".join(it["需求编号"] or "-" for it in wk["items"])
        ws2.append([
            wk["week"],
            f"Sprint {wk['week']:02d}",
            period,
            len(wk["items"]),
            wk["be"],
            wk["fe"],
            te,
            f"{wk['be']/BE_CAP*100:.0f}%",
            f"{wk['fe']/FE_CAP*100:.0f}%",
            f"{te/TE_CAP*100:.0f}%",
            main_label,
            ids,
        ])
        r = ws2.max_row
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths2 = [8, 10, 22, 8, 8, 8, 8, 12, 12, 12, 14, 40]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Sheet3: 统计说明
    ws3 = wb.create_sheet("排期说明")
    lines = [
        ["待排期迭代计划", ""],
        ["数据来源", "待排期需求.xlsx → 待排期（不含「进行中」）"],
        ["需求总数", len(items)],
        ["迭代总数", len(weeks)],
        ["起始日期", SPRINT_START.strftime("%Y-%m-%d")],
        ["排期模式", "高效团队（单任务工时约为常规 40%~55%）"],
        ["团队配置", "3后端 + 1前端 + 1.5测试"],
        ["每周产能", f"后端{BE_CAP}人日 / 前端{FE_CAP}人日 / 测试{TE_CAP}人日"],
        ["工时口径", "简单0.5 / 中等1.5 / 复杂4~5（后端人日）"],
        ["", ""],
        ["特殊规则", "全部「紧急」优先级需求固定排入 Sprint 01"],
        ["", ""],
        ["优先级排序规则(S02起)", ""],
        ["1", "简单 + 高/紧急"],
        ["2", "全部简单"],
        ["3", "高/紧急 + 中等"],
        ["4", "高/紧急 + 复杂"],
        ["5", "中 + 中等"],
        ["6", "中 + 复杂"],
        ["7", "低 + 中等"],
        ["8", "低 + 复杂"],
        ["", ""],
        ["排序组统计", "数量"],
    ]
    urgent_count = sum(1 for it in items if it["优先级"] == "紧急")
    lines.append(["紧急需求数(S01)", urgent_count])
    lines.append(["", ""])
    gc = Counter(it["排序组"] for it in items)
    for g in sorted(gc):
        lines.append([GROUP_NAMES[g], gc[g]])

    for row in lines:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 50

    wb.save(OUT)
    return len(items), len(weeks)


if __name__ == "__main__":
    data = read_pending_items()
    wks = schedule_weeks(data)
    n, w = write_excel(data, wks)
    urgent_n = sum(1 for it in data if it["优先级"] == "紧急")
    print(f"已生成: {OUT}")
    print(f"待排期需求 {n} 条, 共 {w} 个迭代, 其中紧急 {urgent_n} 条均在 Sprint 01")
