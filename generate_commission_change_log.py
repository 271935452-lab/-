# -*- coding: utf-8 -*-
"""生成提成管理 · 需求变更记录表"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(r"d:\Cursor\头程项目\验收\提成管理-需求变更记录.xlsx")

HEADERS = [
    "序号",
    "需求编号",
    "功能模块",
    "变更类型",
    "优先级",
    "变更描述",
    "变更前/现状",
    "变更后/目标",
    "状态",
    "备注",
]

# (module, change_type, priority, desc, before, after, status, note)
ROWS = [
    # —— 已完成 ——
    ("规则包/规则层", "优化", "P0", "规则层编辑时分公司支持多选", "分公司单选", "分公司多选", "已完成", "最高优先级"),
    ("人员映射表", "需求新增", "P1", "人员映射表新增删除功能", "不支持删除", "支持删除", "已完成", ""),
    (
        "人员映射表",
        "需求新增",
        "P1",
        "人员映射表删除约束",
        "—",
        "未存在发放提成数据的映射可删除",
        "已完成",
        "与删除功能配套",
    ),
    ("提成试算", "需求新增", "P1", "提成试算删除功能", "不支持删除", "不存在发放提成状态的数据可删除", "已完成", ""),
    (
        "基数方案",
        "优化",
        "P0",
        "基数维护新增「收货方数/80」",
        "无此基数项",
        "取运单收货方数之和÷80，只取整（如 5.6→5）",
        "已完成",
        "最高优先级",
    ),
    (
        "人员映射表",
        "优化",
        "P1",
        "指定规则包展示规则包名称",
        "可能仅展示 ID 或不直观",
        "列表/详情显示规则包名称",
        "已完成",
        "",
    ),
    (
        "提成试算",
        "优化",
        "P0",
        "试算归档数据保存且禁止重复试算",
        "试算结果未持久化或可重复试算",
        "每次归档提成数据保存；已归档不支持再次试算",
        "已完成",
        "最高优先级",
    ),
    (
        "提成试算/月度归档",
        "优化",
        "P1",
        "账期改为月度范围",
        "账期可能为单点月份",
        "提成试算、月度归档账期均为月度范围",
        "已完成",
        "月度归档已完成；试算待确认是否同步完成",
    ),
    (
        "基数方案",
        "需求变更",
        "P1",
        "变量重命名：毛利、计费重",
        "毛利；计费重",
        "业务员毛利；收货计费重",
        "已完成",
        "",
    ),
    (
        "人员映射表",
        "优化",
        "P0",
        "提成规则包实时匹配标识",
        "无实时匹配开关",
        "选「是」实时匹配规则包；选「否」不自动匹配",
        "已完成",
        "最高优先级",
    ),
    (
        "基数/销售产品",
        "需求变更",
        "P1",
        "真实柜数按销售产品配置调整",
        "原柜数取数逻辑",
        "按图片/销售产品维度调整真实柜数口径",
        "已完成",
        "",
    ),
    (
        "提成试算",
        "优化",
        "P1",
        "试算列表新增备注字段",
        "无备注或不可见",
        "展示月度调整表备注；多条拼接显示",
        "已完成",
        "",
    ),
    (
        "提成试算",
        "需求变更",
        "P1",
        "操作费/内陆费负号、字段更名、新增件数",
        "操作费/内陆费正数；固定提成点；无件数",
        "操作费/内陆费带负号；固定提成点→提成系数；增加件数字段",
        "已完成",
        "",
    ),
    (
        "规则匹配",
        "优化",
        "P1",
        "人员映射规则包降级匹配",
        "部门+小组未命中可能直接失败",
        "有部门/小组时先按部门+小组匹配；未命中再按分公司+其他条件匹配",
        "已完成",
        "",
    ),
    (
        "运单底表",
        "需求变更",
        "P1",
        "毛利字段更名",
        "毛利",
        "业务员毛利",
        "已完成",
        "",
    ),
    # —— 待优化 ——
    (
        "同行试算",
        "缺陷修复/优化",
        "P0",
        "阶梯汇总展示多规则分产品计费明细",
        "阶梯汇总未展示分产品明细",
        "例：海派特价 0.2 计费重 100×0.2=20；空派 0.2 计费重 100×0.2=20；合计应发 40",
        "待开发",
        "0625 验收 P0",
    ),
    (
        "人员映射表/月度调整表",
        "需求变更",
        "P0",
        "系统账户名称改为员工昵称",
        "系统账户名称",
        "员工昵称；列表、导入/导出模板同步；工号仍为主键",
        "待开发",
        "0625 验收 P0",
    ),
    (
        "运单底表",
        "需求新增",
        "P0",
        "筛选增加部门、小组",
        "无部门/小组筛选",
        "筛选条件增加部门、小组（来源人员映射表）",
        "待开发",
        "0625 验收 P0",
    ),
    (
        "基数方案",
        "需求变更",
        "P0",
        "收货方数/80 取组员汇总",
        "可能取主管本人汇总",
        "取组员收货方数汇总÷80（主管场景不含本人）",
        "待开发",
        "0625 验收 P0",
    ),
    (
        "人员映射表",
        "需求变更",
        "P1",
        "导出模板提成岗位改字典下拉",
        "提成岗位自由文本",
        "导出/导入模板提成岗位为系统字典下拉，禁止自由文本",
        "待开发",
        "0625 验收",
    ),
    (
        "业务数据统计",
        "需求变更",
        "P1",
        "直客业务员单含退件数据",
        "业务数据统计可能不含退件",
        "直客业务员相关统计包含退件数据",
        "待开发",
        "",
    ),
    (
        "运单底表",
        "需求新增",
        "P1",
        "方数列增加合计",
        "列表无方数合计",
        "运单底表方数增加合计行/合计展示",
        "待开发",
        "",
    ),
    (
        "提成角色",
        "需求变更",
        "P1",
        "新增提成角色",
        "无同行主管、KA客服、KA助理",
        "同行主管（同同行经理）；KA客服（运单业务员）；KA助理（运单客服）",
        "待开发",
        "",
    ),
    (
        "提成试算",
        "需求新增",
        "P1",
        "试算列表数据权限",
        "权限未细分",
        "业务员/客服看自己；上级看下级（人员映射）；其他非业务员类角色看全部",
        "待开发",
        "",
    ),
    (
        "毛利调整",
        "需求新增",
        "P2",
        "毛利调整的数据计算",
        "无或未完善",
        "实现毛利调整相关数据计算逻辑",
        "待开发",
        "",
    ),
    (
        "毛利调整",
        "需求新增",
        "P2",
        "毛利负数结转至下次提成扣",
        "负数毛利可能不结转",
        "毛利为负时结转至下次提成扣减",
        "待开发",
        "",
    ),
    (
        "提成试算",
        "需求新增",
        "P1",
        "阶梯汇总明细字段统一为提成系数",
        "固定比例/毛利阶梯明细可能仍用旧字段名",
        "阶梯汇总明细、固定比例、毛利阶梯计费明细中统一显示「提成系数」",
        "待开发",
        "",
    ),
]

THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="0B1F3A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SUB_FONT = Font(size=10, color="595959")
DONE_FILL = PatternFill("solid", fgColor="F6FFED")
TODO_FILL = PatternFill("solid", fgColor="FFF7E6")
P0_FILL = PatternFill("solid", fgColor="FFF1F0")


def write_change_log_sheet(wb, rows, sheet_title="需求变更记录"):
    """写入需求变更记录 Sheet（可被验收总览复用）"""
    ws = wb.create_sheet(sheet_title)

    ws.merge_cells("A1:J1")
    ws["A1"] = "提成管理 · 需求变更记录"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    done = sum(1 for x in rows if x[6] == "已完成")
    todo = sum(1 for x in rows if x[6] == "待开发")
    p0_todo = sum(1 for x in rows if x[6] == "待开发" and x[2] == "P0")
    ws.merge_cells("A2:J2")
    ws["A2"] = (
        f"整理日期：2026-06-16 · 已完成 {done} 项 · 待优化 {todo} 项（P0 {p0_todo} 项）"
        " · 来源：0625 提成验收与待办清单"
    )
    ws["A2"].font = SUB_FONT

    start_row = 4
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(rows, 1):
        r = start_row + i
        module, ctype, pri, desc, before, after, status, note = row
        req_id = f"TC-{i:03d}"
        values = [i, req_id, module, ctype, pri, desc, before, after, status, note]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if col == 5 and val == "P0":
                c.fill = P0_FILL
            if col == 9:
                c.fill = DONE_FILL if val == "已完成" else TODO_FILL

    widths = [6, 10, 18, 12, 8, 36, 28, 36, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[start_row].height = 32

    ws2 = wb.create_sheet("变更汇总")
    ws2["A1"] = "状态"
    ws2["B1"] = "数量"
    ws2["A1"].font = ws2["B1"].font = Font(bold=True)
    summary = [
        ("已完成", done),
        ("待开发", todo),
        ("P0 待开发", p0_todo),
    ]
    for i, (a, b) in enumerate(summary, 2):
        ws2.cell(row=i, column=1, value=a)
        ws2.cell(row=i, column=2, value=b)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 10


def main():
    wb = Workbook()
    write_change_log_sheet(wb, ROWS)
    wb.remove(wb.active)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"已生成: {OUT}")


if __name__ == "__main__":
    main()
