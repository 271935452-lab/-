# -*- coding: utf-8 -*-
"""按同一运单号、员工、费用性质、费用类型合并；备注仅拼接原文。"""
from pathlib import Path
from collections import OrderedDict

import openpyxl
from openpyxl.utils import get_column_letter

SRC = Path(__file__).with_name("总导入-2026-06提成管理导入表外成本导入模板 .xlsx")
OUT = Path(__file__).with_name(
    "总导入-2026-06提成管理导入表外成本导入模板_按运单员工费用合并.xlsx"
)


def uniq_join(items, sep="；", limit=None):
    seen = []
    for x in items:
        if x and x not in seen:
            seen.append(x)
    if not seen:
        return None
    if limit is not None and len(seen) > limit:
        return sep.join(seen[:limit]) + f"…等{len(seen)}项"
    return sep.join(seen)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    # 分组键：运单号 + 员工昵称 + 员工姓名 + 月度 + 费用性质 + 费用类型
    # 源表无「调整类型」列，费用维度以费用性质+费用类型为准
    groups = OrderedDict()
    for r in range(2, ws.max_row + 1):
        nick = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        month = ws.cell(r, 3).value
        waybill = ws.cell(r, 4).value
        nature = ws.cell(r, 5).value
        fee_type = ws.cell(r, 6).value
        amount = ws.cell(r, 7).value
        customer = ws.cell(r, 8).value
        cs = ws.cell(r, 9).value
        remark = ws.cell(r, 10).value
        waybill_opt = ws.cell(r, 11).value

        key = (waybill, nick, name, month, nature, fee_type)
        if key not in groups:
            groups[key] = {
                "nick": nick,
                "name": name,
                "month": month,
                "waybill": waybill,
                "nature": nature,
                "fee_type": fee_type,
                "amount": 0.0,
                "customers": [],
                "css": [],
                "remarks": [],
                "waybill_opts": [],
            }
        g = groups[key]
        if amount is not None and amount != "":
            g["amount"] += float(amount)
        if customer not in (None, ""):
            g["customers"].append(str(customer))
        if cs not in (None, ""):
            g["css"].append(str(cs))
        if remark not in (None, ""):
            g["remarks"].append(str(remark))
        if waybill_opt not in (None, ""):
            g["waybill_opts"].append(str(waybill_opt))

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = wb.sheetnames[0]

    for c, h in enumerate(headers, 1):
        out_ws.cell(1, c, h)

    for i, g in enumerate(groups.values(), 2):
        out_ws.cell(i, 1, g["nick"])
        out_ws.cell(i, 2, g["name"])
        out_ws.cell(i, 3, g["month"])
        out_ws.cell(i, 4, g["waybill"])
        out_ws.cell(i, 5, g["nature"])
        out_ws.cell(i, 6, g["fee_type"])
        out_ws.cell(i, 7, round(g["amount"], 2))
        out_ws.cell(i, 8, uniq_join(g["customers"]))
        out_ws.cell(i, 9, uniq_join(g["css"]))
        # 备注：只拼接原备注，不追加其他内容
        out_ws.cell(i, 10, uniq_join(g["remarks"]))
        out_ws.cell(i, 11, uniq_join(g["waybill_opts"]))

    for name in wb.sheetnames[1:]:
        src_ws = wb[name]
        dst = out_wb.create_sheet(name)
        for row in src_ws.iter_rows():
            for cell in row:
                dst.cell(cell.row, cell.column, cell.value)

    widths = [12, 12, 10, 18, 10, 22, 12, 16, 12, 40, 14]
    for i, w in enumerate(widths, 1):
        out_ws.column_dimensions[get_column_letter(i)].width = w

    out_wb.save(OUT)

    orig_total = sum(float(ws.cell(r, 7).value or 0) for r in range(2, ws.max_row + 1))
    new_total = sum(
        float(out_ws.cell(r, 7).value or 0) for r in range(2, out_ws.max_row + 1)
    )
    print(f"源行数: {ws.max_row - 1}")
    print(f"合并后行数: {out_ws.max_row - 1}")
    print(f"源金额合计: {round(orig_total, 2)}")
    print(f"新表金额合计: {round(new_total, 2)}")
    print(f"差额: {round(new_total - orig_total, 4)}")
    print(f"输出: {OUT}")


if __name__ == "__main__":
    main()
