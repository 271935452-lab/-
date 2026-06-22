# -*- coding: utf-8 -*-
"""生成项目进度总览 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"d:\Cursor\头程项目\项目进度总览.xlsx"

HEADERS = [
    "序号",
    "模块/事项",
    "目前使用情况",
    "待解决问题",
    "开发完成时间",
    "最终验收时间",
    "验收人",
    "项目状态",
]

ROWS = [
    [
        1,
        "整柜询价/下单",
        "客服已可下单；产品侧刚测完，处于优化阶段",
        "测试暴露的问题仍在优化中",
        "待确认（优化中）",
        "待确认",
        "待确认",
        "测试/优化中",
    ],
    [
        2,
        "散货询价",
        "已在使用中",
        "与整柜联动、区域推广策略待明确",
        "已上线（使用中）",
        "待确认",
        "待确认",
        "已在使用",
    ],
    [
        3,
        "整柜推广",
        "刚上线；与袁经理沟通是否先在苏州试点",
        "推广范围、试点条件待袁经理确认",
        "已上线",
        "待确认（苏州试点未定）",
        "袁经理（推广决策）",
        "测试/优化中",
    ],
    [
        4,
        "提成管理",
        "与财务联调测试中",
        "测试阶段问题待财务侧反馈并修复",
        "待确认",
        "待确认",
        "财务 + 待确认业务负责人",
        "测试/优化中",
    ],
    [
        5,
        "小小科技",
        "在使用中；反馈已解决部分问题",
        "剩余问题范围、是否全部闭环待确认",
        "待确认",
        "待确认",
        "待确认",
        "已在使用",
    ],
    [
        6,
        "待办中心",
        "已打通：询价报价、报关资料两类待办",
        "原有待办需清理；其他流程待逐步完善",
        "部分完成（持续迭代）",
        "待确认",
        "待确认",
        "部分完成",
    ],
    [
        7,
        "服务商模版匹配",
        "方案已定，开发/配置进行中",
        "方案落地与联调细节",
        "预计 2026-07 完成",
        "待确认（建议 7 月内）",
        "待确认",
        "开发中（7月）",
    ],
    [
        8,
        "现金流量表",
        "开发中",
        "需求/联调细节待确认",
        "预计 2026-06 月底",
        "待确认",
        "待确认",
        "开发中（6月底）",
    ],
    [
        9,
        "财务自动核销",
        "未开始",
        "需求尚未对齐",
        "未排期",
        "未排期",
        "待确认",
        "未开始",
    ],
    [
        10,
        "报关行对接",
        "开发中（原型/方案已有）",
        "对接联调、上线前验收",
        "预计 2026-06 月底上线",
        "待确认（建议与上线同步）",
        "待确认",
        "开发中（6月底）",
    ],
    [
        11,
        "进口商管理",
        "开发中",
        "功能完整性、与业务联调",
        "预计 2026-06 月底",
        "待确认",
        "待确认",
        "开发中（6月底）",
    ],
    [
        12,
        "工单系统",
        "未开始，未排期",
        "需求、优先级、资源均未定",
        "未排期",
        "未排期",
        "待确认",
        "未开始",
    ],
    [
        13,
        "APP 业务重构",
        "未开始，未排期",
        "范围、排期均未定",
        "未排期",
        "未排期",
        "待确认",
        "未开始",
    ],
]

STATUS_SUMMARY = [
    ["项目状态", "模块", "说明"],
    ["已在使用", "散货询价", "正式使用中"],
    ["已在使用", "小小科技", "使用中，部分问题已解决"],
    ["已在使用", "待办中心", "询价报价、报关资料待办已拉通"],
    ["已在使用", "整柜", "刚上线；客服已可下单"],
    ["测试/优化中", "整柜询价/下单", "产品测试完成，问题优化中"],
    ["测试/优化中", "提成管理", "与财务联调测试"],
    ["测试/优化中", "整柜推广", "苏州试点方案与袁经理沟通中"],
    ["开发中（6月底）", "现金流量表", "预计 2026-06 月底完成"],
    ["开发中（6月底）", "报关行对接", "预计 2026-06 月底上线"],
    ["开发中（6月底）", "进口商管理", "预计 2026-06 月底完成"],
    ["开发中（7月）", "服务商模版匹配", "预计 2026-07 完成"],
    ["未开始", "财务自动核销", "待对需求"],
    ["未开始", "工单系统", "未排期"],
    ["未开始", "APP 业务重构", "未排期"],
]

STATUS_COLORS = {
    "已在使用": "C6EFCE",
    "测试/优化中": "FFEB9C",
    "部分完成": "BDD7EE",
    "开发中（6月底）": "FCE4D6",
    "开发中（7月）": "FCE4D6",
    "未开始": "FFC7CE",
}

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, row, col_count):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = center
        cell.border = border


def style_body(ws, start_row, end_row, col_count, status_col=None):
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = wrap if c != 1 else center
            if status_col and c == status_col:
                val = cell.value or ""
                color = STATUS_COLORS.get(val)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# Sheet 1: 项目进度总览
ws1 = wb.active
ws1.title = "项目进度总览"
ws1.merge_cells("A1:H1")
title = ws1["A1"]
title.value = "头程项目 · 进度总览（截至 2026-06-16）"
title.font = Font(bold=True, size=14)
title.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 28

for col, h in enumerate(HEADERS, 1):
    ws1.cell(row=2, column=col, value=h)
style_header(ws1, 2, len(HEADERS))

for i, row in enumerate(ROWS, 3):
    for col, val in enumerate(row, 1):
        ws1.cell(row=i, column=col, value=val)

style_body(ws1, 3, 3 + len(ROWS) - 1, len(HEADERS), status_col=8)
set_widths(ws1, [6, 18, 32, 32, 22, 22, 22, 16])

# Sheet 2: 状态分组
ws2 = wb.create_sheet("状态分组")
for r, row in enumerate(STATUS_SUMMARY, 1):
    for c, val in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=val)
style_header(ws2, 1, 3)
style_body(ws2, 2, len(STATUS_SUMMARY), 3, status_col=1)
set_widths(ws2, [18, 20, 40])

# Sheet 3: 填写说明
ws3 = wb.create_sheet("填写说明")
notes = [
    ["字段", "说明"],
    ["目前使用情况", "当前业务/测试/上线状态"],
    ["待解决问题", "阻塞项、优化项、待确认事项"],
    ["开发完成时间", "开发侧计划完成或实际上线时间"],
    ["最终验收时间", "业务/财务正式签收时间（待补充）"],
    ["验收人", "负责验收的角色或姓名（待补充）"],
    ["项目状态", "汇总状态，便于筛选：已在使用 / 测试优化中 / 开发中 / 未开始"],
    ["", ""],
    ["备注", "原文未提供验收人与最终验收时间，表中标注「待确认」处请项目负责人补充"],
]
for r, row in enumerate(notes, 1):
    for c, val in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=val)
style_header(ws3, 1, 2)
style_body(ws3, 2, len(notes), 2)
set_widths(ws3, [18, 60])

wb.save(OUTPUT)
print(f"已生成: {OUTPUT}")
