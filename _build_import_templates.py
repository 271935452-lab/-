# -*- coding: utf-8 -*-
"""从《数据映射与逻辑加工表.xls》拆出 3 个独立导入模板（xlsx）。"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = r"d:\Cursor\头程项目"

FILES = {
    "导入模板_手工万能调整池.xlsx": "pool",
    "导入模板_主管绩效.xlsx": "perf",
    "导入模板_人员映射.xlsx": "person",
}


def header_style(cell, required=False):
    cell.font = Font(bold=True, size=11, color="FFFFFF" if required else "111827")
    cell.fill = PatternFill("solid", fgColor="1D4ED8" if required else "E2E8F0")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    side = Side(style="thin", color="94A3B8")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def thin_border(cell):
    side = Side(style="thin", color="CBD5E1")
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def autosize(ws, max_w=48):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        m = 10
        for row in range(1, min(ws.max_row + 1, 80)):
            v = ws.cell(row=row, column=col).value
            if v is not None:
                m = max(m, min(max_w, len(str(v)) + 2))
        ws.column_dimensions[letter].width = m


def add_sheet_help(wb, title, lines):
    ws = wb.create_sheet("填写说明", 1)
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    r = 3
    for line in lines:
        ws.cell(row=r, column=1, value=line)
        ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 92


def build_pool():
    wb = Workbook()
    ws = wb.active
    ws.title = "手工万能调整池"
    headers = [
        ("*系统账户名称", True),
        ("*员工姓名", True),
        ("*月度", True),
        ("运单号", False),
        ("*费用性质", True),
        ("*费用类型", True),
        ("*成本金额", True),
        ("*币种", True),
        ("*调整类型", True),
        ("客户名称", False),
        ("对应客服", False),
        ("备注", False),
        ("运单号非必填", False),
    ]
    for c, (h, req) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        header_style(cell, req)
    # 示例行（与源表口径一致，便于财务对照）
    ex1 = ["", "", "2026-06", "", "扣毛利", "", "", "", "加项", "", "", "", ""]
    ex2 = ["", "", "", "", "扣提成", "", "", "", "减项", "", "", "", ""]
    for r, row in enumerate([ex1, ex2], 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            thin_border(cell)
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    autosize(ws)

    add_sheet_help(
        wb,
        "手工万能调整池 · 导入说明",
        [
            "来源：数据映射与逻辑加工表 · 工作表「手工万能调整池」拆出；字段名与必填标记（*）与源表一致。",
            "用途：表外/万能调整类数据批量导入；与原型「导入表外成本」「月度调整表」链路对齐时，请与研发确认字段与费用字典枚举是否一致。",
            "*月度：建议格式 YYYY-MM（例 2026-06）。",
            "运单号：可为空；最后一列表头「运单号非必填」为源表提示列，导入时若系统不需要该列可忽略或映射为说明字段。",
            "*费用性质：示例为「扣毛利」「扣提成」；若系统使用「毛利扣」「提成扣」等编码，以接口/字典为准。",
            "*调整类型：示例为「加项」「减项」；具体枚举以产品定义为准。",
            "删除示例行后再导入正式数据。",
        ],
    )
    wb.save(f"{BASE}\\导入模板_手工万能调整池.xlsx")
    print("OK", "导入模板_手工万能调整池.xlsx")


def build_perf():
    wb = Workbook()
    ws = wb.active
    ws.title = "主管绩效"
    headers = ["系统账户名称", "员工姓名", "月度", "绩效系数"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        header_style(cell, False)
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=4, value="")
    for r in range(1, 3):
        for c in range(1, 5):
            thin_border(ws.cell(row=r, column=c))
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24
    autosize(ws)

    add_sheet_help(
        wb,
        "主管绩效模版 · 导入说明",
        [
            "来源：数据映射与逻辑加工表 · 工作表「主管绩效模版」拆出。",
            "用途：按员工+月度导入绩效系数；与原型「导入绩效系数弹窗页」一致时，百分比需转为数字系数（如 108% → 1.08）。",
            "月度：建议 YYYY-MM。",
            "若需「必填」标记或增加分公司等扩展列，可在定稿需求后追加列并同步本模板。",
        ],
    )
    wb.save(f"{BASE}\\导入模板_主管绩效.xlsx")
    print("OK", "导入模板_主管绩效.xlsx")


def build_person():
    wb = Workbook()
    ws = wb.active
    ws.title = "人员映射"
    headers = [
        "员工工号",
        "员工昵称",
        "员工姓名",
        "所属部门",
        "所属小组",
        "生效起始时间",
        "失效截止时间",
        "提成计算开始时间",
        "提成计算结束时间",
        "入职时间",
        "转正时间",
        "离职时间",
        "提成发放状态",
        "人员状态",
        "提成岗位",
        "指定规则（预留字段）",
        "备注",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        header_style(cell, False)
    # 一行示例（文本日期，避免 Excel 序列号）
    example = [
        "E0010086",
        "黑猫警长",
        "黄丽青",
        "深圳客服部",
        "",
        "2026-01-01",
        "长期",
        "2026-01-01",
        "2026-12-31",
        "2024-06-01",
        "2024-09-01",
        "",
        "正常",
        "在职",
        "直客业务员",
        "",
        "示例行可删",
    ]
    for c, v in enumerate(example, 1):
        cell = ws.cell(row=2, column=c, value=v)
        thin_border(cell)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36
    autosize(ws)

    add_sheet_help(
        wb,
        "人员映射表 · 导入说明",
        [
            "来源：数据映射与逻辑加工表 · 工作表「人员映射表」表头行拆出（源表上方空行已去掉）。",
            "用途：与原型「人员映射表」「批量导入人员映射」字段对齐；时间列建议统一为 YYYY-MM-DD 或文本「长期」。",
            "提成岗位、人员状态、提成发放状态：取值需与系统字典一致（源表中有：直客业务员、同行客服、客服中心主管等示例）。",
            "指定规则（预留字段）：可填规则包编码或留空走自动匹配，以实现为准。",
            "删除示例行后再导入正式数据。",
        ],
    )
    wb.save(f"{BASE}\\导入模板_人员映射.xlsx")
    print("OK", "导入模板_人员映射.xlsx")


def main():
    build_pool()
    build_perf()
    build_person()


if __name__ == "__main__":
    main()
