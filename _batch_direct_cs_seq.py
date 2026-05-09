# -*- coding: utf-8 -*-
"""
为索引中所有「直客业务员」（不含主管）场景子表补充：
· A3 顺序执行规则包说明
· 段1 客服演示行（若无）
· 业务员步骤4（列M）追加「顺序执行·段2」前缀
· 合计行（Σ步骤8）补充客服段说明；若无合计行则追加一行
"""
import re
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter

SRC = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并.xlsx"
BAK = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_before_batch_cs_seq.xlsx"

RULE_NOTE_APPEND = (
    "\n【规则包】编排类型=顺序执行（示意）：①客服子规则段先行（计提记入上行客服演示账号）；"
    "②业务员子规则段匹配入职/提成区间并计提；③万能池「提成扣」等在发放层冲减（见步骤8）。"
    "包编码示意 RP_{sid}_SEQ；互斥组/相加组在其他映射另行演示。"
)

CS_RATE = 0.0075


def find_hdr_row(ws):
    for r in range(1, min(ws.max_row + 1, 35)):
        v = ws.cell(r, 10).value
        if v and isinstance(v, str) and "步骤1" in v:
            return r
    return None


def find_total_row(ws, hdr_row):
    lim = max(ws.max_row, hdr_row + 8) + 6
    for r in range(hdr_row + 1, lim):
        for c in (1, 10):
            v = ws.cell(r, c).value
            if v and isinstance(v, str):
                if "合计" in v or "Σ步骤" in v:
                    return r
    return None


def shift_merge_ranges(merge_strs, hdr_row):
    """Insert happened at hdr_row+1: rows originally >= hdr_row+1 move down by 1."""
    out = []
    for s in merge_strs:
        c1, r1, c2, r2 = range_boundaries(s)
        if r1 >= hdr_row + 1:
            r1 += 1
            r2 += 1
        out.append(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")
    return out


def filter_sales_row_step_merges(ws, merge_strs, hdr):
    """剔除非合计行的单行 J:P 合并，避免挡住步骤列分项写入。"""
    keep = []
    for s in merge_strs:
        c1, r1, c2, r2 = range_boundaries(s)
        if r1 == r2 and c1 == 10 and c2 == 17 and r1 > hdr + 1:
            a = ws.cell(r1, 1).value
            if not a or not isinstance(a, str) or "合计" not in a:
                continue
        keep.append(s)
    return keep


def sid_from_sheet(ws):
    t = str(ws.cell(1, 1).value or "")
    m = re.search(r"场景(\d{3})", t)
    if m:
        return "S" + m.group(1)
    m = re.search(r"S(\d{3})", t)
    if m:
        return "S" + m.group(1)
    return "S000"


def ensure_a3(ws, sid):
    cell = ws["A3"]
    v = cell.value or ""
    if not isinstance(v, str):
        v = str(v)
    if "顺序执行" not in v:
        cell.value = v.strip() + RULE_NOTE_APPEND.format(sid=sid)
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def pick_revenue_waybill(ws, sales_row, sid):
    e = ws.cell(sales_row, 5).value
    rev = None
    try:
        if e is not None and str(e).strip() != "":
            rev = float(e)
    except (TypeError, ValueError):
        rev = None
    if rev is None or rev == 0:
        rev = 52800.0
    wb = ws.cell(sales_row, 3).value
    if wb and isinstance(wb, str) and wb.strip():
        waybill = wb.strip()
    else:
        waybill = f"WB-DIRECT-{sid}-202603001"
    return rev, waybill


def cs_row_values(sid, revenue, cs_amt, waybill, pkg_code):
    nick = f"顺序包客服_{sid}"
    name_b = "（演示）顺序执行·段1"
    return [
        nick,
        name_b,
        waybill,
        "客服",
        revenue,
        None,
        None,
        None,
        None,
        f"命中映射「{nick}+{name_b}」（示意）",
        "映射生效∩2026-03：通过（示意）",
        "提成区间：覆盖（示意）；在职·发放正常（示意）",
        f"【顺序执行·段1】规则包 {pkg_code}",
        "命中 CS_WITH_SALES_RATE（客服计提先行）",
        f"基数=本运单归因应收 {revenue:.2f}（净应收汇总示意）",
        cs_amt,
        cs_amt,
    ]


def patch_sales_row(ws, r, cs_amt, pkg_code):
    prefix_rule = (
        f"前置段1客服计提 {cs_amt:.2f} 已完成（见上行·{pkg_code}）；"
    )
    cur = ws.cell(r, 14).value
    cur = "" if cur is None else str(cur)
    cur = re.sub(
        r"^【顺序执行·段2】前置段1客服计提 [\d.]+\s*已完成（见上行·RP_S\d+_SEQ）；",
        "",
        cur,
    )
    cur = re.sub(
        r"^前置段1客服计提 [\d.]+\s*已完成（见上行·RP_S\d+_SEQ）；",
        "",
        cur.strip(),
    )
    ws.cell(r, 13).value = f"【顺序执行·段2】规则包 {pkg_code}"
    if not cur.startswith("前置段1客服计提"):
        cur = prefix_rule + cur
    ws.cell(r, 14).value = cur


def read_sales_final(ws, r):
    for col in (17, 16):
        v = ws.cell(r, col).value
        try:
            if v is not None and str(v).strip() != "":
                return float(v)
        except (TypeError, ValueError):
            continue
    return 0.0


def patch_total_row(ws, total_row, cs_amt, sales_final_display):
    ws.cell(total_row, 1).value = "合计（演示）"
    ws.cell(total_row, 10).value = (
        f"Σ步骤8（业务员本人合计）= {sales_final_display:.2f} CNY。"
        f"顺序包段1客服应发 {cs_amt:.2f} 记入上行客服演示账号，不与业务员本人合计相加。"
    )
    ws.cell(total_row, 10).alignment = Alignment(wrap_text=True, vertical="top")


def ensure_row_capacity(ws, row_idx):
    while ws.max_row < row_idx:
        ws.append([])


def apply_merges(ws, merge_specs):
    seen = set()
    for m in merge_specs:
        if m in seen:
            continue
        seen.add(m)
        try:
            ws.merge_cells(m)
        except Exception:
            pass


def process_one_sheet(ws, sid):
    hdr = find_hdr_row(ws)
    if not hdr:
        return "no-header"

    merges_orig = [str(x) for x in ws.merged_cells.ranges]
    for m in merges_orig:
        ws.unmerge_cells(m)

    had_cs = ws.cell(hdr + 1, 4).value == "客服"
    pkg_code = f"RP_{sid}_SEQ"

    merges_use = list(merges_orig)

    if not had_cs:
        rev, waybill = pick_revenue_waybill(ws, hdr + 1, sid)
        cs_amt = round(rev * CS_RATE, 2)
        ws.insert_rows(hdr + 1)
        merges_use = shift_merge_ranges(merges_orig, hdr)
        vals = cs_row_values(sid, rev, cs_amt, waybill, pkg_code)
        for col, val in enumerate(vals, start=1):
            ws.cell(hdr + 1, col).value = val

    ensure_a3(ws, sid)

    cs_amt = ws.cell(hdr + 1, 17).value
    try:
        cs_amt = float(cs_amt)
    except (TypeError, ValueError):
        cs_amt = round(52800 * CS_RATE, 2)

    total_row = find_total_row(ws, hdr)
    sales_first = hdr + 2
    sales_last = (total_row - 1) if total_row else hdr + 2

    sales_sum = 0.0
    any_sales = False
    for r in range(sales_first, sales_last + 1):
        role = ws.cell(r, 4).value
        if role == "客服":
            continue
        any_sales = True
        patch_sales_row(ws, r, cs_amt, pkg_code)
        sales_sum += read_sales_final(ws, r)

    display_final = sales_sum if any_sales else read_sales_final(ws, sales_first)

    total_row2 = find_total_row(ws, hdr)
    extra_merges = []
    created_total = False
    if total_row2:
        tr = total_row2
    else:
        tr = hdr + 3
        ensure_row_capacity(ws, tr)
        created_total = True

    patch_total_row(ws, tr, cs_amt, display_final)
    if created_total:
        extra_merges.extend([f"A{tr}:H{tr}", f"J{tr}:Q{tr}"])

    merges_use = filter_sales_row_step_merges(ws, merges_use, hdr)

    specs = []
    seen = set()
    for m in merges_use + extra_merges:
        if m not in seen:
            seen.add(m)
            specs.append(m)
    apply_merges(ws, specs)
    return "ok"


def main():
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=False)
    idx_name = wb.sheetnames[0]
    idx = wb[idx_name]

    targets = []
    for r in range(2, idx.max_row + 1):
        role = idx.cell(r, 5).value
        sid = idx.cell(r, 7).value
        if not role or not sid:
            continue
        rs = str(role)
        if "直客业务员" not in rs or "主管" in rs:
            continue
        targets.append(str(sid))

    seen = set()
    for sid in targets:
        if sid in seen:
            continue
        seen.add(sid)
        if sid not in wb.sheetnames:
            print(sid, "missing-sheet")
            continue
        ws = wb[sid]
        sid_cell = sid_from_sheet(ws)
        st = process_one_sheet(ws, sid_cell)
        print(sid, st)

    wb.save(SRC)
    wb.close()
    print("Backup:", BAK)


if __name__ == "__main__":
    main()
