# -*- coding: utf-8 -*-
"""One-off generator: 提成计算全链路场景说明.xlsx — run from project root if needed."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Use ASCII filename so paths are stable on all Windows locales; title is in sheet "0_使用说明"
OUT = "commission_full_pipeline_scenarios_V1.xlsx"


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(bold=True, color="FFFFFF", size=11)
    side = Side(style="thin", color="CBD5E1")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(left=side, right=side, top=side, bottom=side)


def autosize_columns(ws, max_width=64):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 200)):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max_width, max(10, max_len + 2))


def main():
    wb = Workbook()
    # --- Sheet 0: 说明 ---
    meta = wb.active
    meta.title = "0_使用说明"
    meta["A1"] = "文档名称"
    meta["B1"] = "提成计算全链路 · 场景与步骤（V1 知识库）"
    meta["A2"] = "适用范围"
    meta["B2"] = "开发自测、测试用例设计、财务/运营做规则与映射配置时的口径对齐"
    meta["A3"] = "来源"
    meta["B3"] = "根目录 MVP 原型页 + 二期/ 下规则与试算相关 HTML；为归纳稿，以最终 PRD/研发实现为准"
    meta["A4"] = "维护"
    meta["B4"] = "随原型迭代更新本文件版本号与日期"
    meta["A5"] = "工作表"
    meta["B5"] = "1=端到端步骤 | 2=各场景计算要点 | 3=试算与异常 | 4=场景矩阵(二期) | 5=页面索引 | 6=配置检查清单"
    meta["A6"] = "文件"
    meta["B6"] = OUT + "（同内容另存为「提成计算全链路场景说明_V1.xlsx」亦可）"
    for r in range(1, 6):
        meta.row_dimensions[r].height = 22
    meta.column_dimensions["A"].width = 14
    meta.column_dimensions["B"].width = 88

    # --- Sheet 1: 端到端 ---
    ws1 = wb.create_sheet("1_端到端计算步骤", 1)
    steps = [
        ("1", "主数据", "运单/订单进入运单管理；沉淀运单号、客户类型、业务员/客服、应收、成本、产品、所属分公司等。", "运单管理（外系统）+ 业务数据统计", "原始业务单", "运单主数据", "业务/中台", ""),
        ("2", "运单级毛利", "毛利 = 应收 − 业务员成本（币种/取整以财务主数据为准）。", "月度提成归档与运单明细-MVP.html (PRD)", "运单应收、成本", "行级毛利", "系统", "与底表字段一致"),
        ("3", "回款", "实收 = 运单下已核销账单金额之和；回款率 = 实收/应收（应收=0 时按产品约定）；截止时点默认账期月末。", "同左", "账单核销状态、应收", "实收、回款率", "系统", ""),
        ("4", "表外调整（运单）", "表外毛利扣汇总 = 导入费用性质=毛利扣 之和；表外提成扣汇总 = 费用性质=提成扣 之和。计算净利润 = 毛利 − 表外毛利扣汇总。", "导入表外成本弹窗 + 表外导入成本列表-MVP", "导入行(运单关联)", "毛利扣/提成扣汇总、计算净利润", "财务", "未关联运单走分摊/其他链路以 PRD 为准"),
        ("5", "人员归属", "按计薪/归属月取有效人员映射：分公司+部门+小组+提成角色；可指定规则包或走自动匹配。", "人员映射表-MVP、引用规则包弹窗", "员工+区间+角色+可选规则包", "归属上下文、命中规则包集合", "HR/销售支持", "映射区间与提成区间包含关系以页面 PRD 为准"),
        ("6", "规则命中", "在有效规则包内按优先级/互斥组/条件(含 service_mode 等)命中单条业务规则；未关联规则进异常列表页签。", "二期/规则包-规则子母表-MVP、编辑规则-MVP、规则引擎中心列表", "规则包版本、组织维度、规则条件", "命中的规则编码、基数方案引用", "规则运营", "与 提成场景全覆盖配置页 场景一一可对照"),
        ("7", "基数", "按基数方案(公式构建器) 将运单/聚合指标转为提成基数；含系统成本项、导入成本项、组件方向(基数/扣减)与缺失策略。", "二期/基数维护-公式构建器", "基数字段、运单及调整聚合结果", "基数金额", "规则运营+开发", ""),
        ("8", "比例与阶梯", "按规则：固定比例 或 毛利/营收 阶梯；渠道场景可有对内提成+对外结算、团队池、主辅/客服 split。", "二期/编辑规则-MVP、提成场景全覆盖", "基数、档配置", "未封顶提成额", "—", "封顶/保底/负毛利在下一步"),
        ("9", "分摊", "有业务员/无业务员客服、主辅比例；渠道双轨、团队池等按场景取 allocate 配置。", "二期/统一试算-客服与渠道场景", "服务关系、分片", "个人/池子可计提份额", "—", ""),
        ("10", "风险与负毛利", "按规则: 负毛利不计提/冲减/公司承担/结转；导入缺失、多命中冲突等走阻断或预警策略。", "二期/统一试算结果-正向与异常", "规则上异常策略、负毛利配置", "试算状态、需财务调整清单", "财务", "与筛选「负毛利策略」等联动"),
        ("11", "绩效与其他月度因子", "绩效系数等按 员工+月 导入，汇总进月度调整主表，乘入或调整提成链路（以最终实现为准）。", "导入绩效系数弹窗、表外导入成本列表", "员工+月+系数", "调整后的提成相关金额/系数", "财务", ""),
        ("12", "试算", "对选中账期人员与运单跑统一试算，输出正向/异常/需调整。", "二期/统一试算结果-正向与异常、提成试算结果", "上游全部", "人维度/单维度试算结果", "测试/财务", ""),
        ("13", "发放门槛与归档", "提成试算状态=已发放 等条件满足时生成可归档批次；底表从运单管理、业务统计、人员映射、月度调整、规则包命中等汇总。", "月度提成归档与运单明细-MVP", "当前批次数据", "归档批次+冻结明细", "财务", "调度节奏见该页 PRD"),
        ("14", "追溯", "计提结果报表快照用于审计对比。", "二期/计提结果报表-快照追溯-MVP", "归档快照", "报表", "财务/审计", ""),
    ]
    h1 = ["步骤", "阶段", "做什么(业务)", "原型/页面(路径)", "主要输入", "主要输出", "主责角色", "备注"]
    ws1.append(h1)
    for row in steps:
        ws1.append(list(row))
    style_header(ws1)
    for r in range(2, ws1.max_row + 1):
        for c in range(1, 9):
            ws1.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 28
    for r in range(2, ws1.max_row + 1):
        ws1.row_dimensions[r].height = 60
    for col in "ABCDEFGH":
        ws1.column_dimensions[col].width = 18
    ws1.column_dimensions["C"].width = 42
    ws1.column_dimensions["D"].width = 36
    ws1.column_dimensions["E"].width = 24
    ws1.column_dimensions["F"].width = 24

    # --- Sheet 2: 计算口径（运单+归属+规则内）---
    ws2 = wb.create_sheet("2_关键口径与公式", 2)
    ws2.append(["主题", "公式/规则(开发可读)", "说明", "页面依据"])
    formulas = [
        ("运单毛利", "gross_profit = receivable - sales_cost", "财务主数据定币种与取整", "月度提成归档… PRD"),
        ("实收", "SUM(bill.amount WHERE waybill 且 status=已核销)", "挂账在运单下", "同左"),
        ("回款率", "IF receivable=0 THEN 约定空/0 ELSE received/receivable", "展示百分号", "同左"),
        ("表外毛利扣", "SUM(导入行 WHERE 运单匹配 AND 费用性质=毛利扣)", "可正负", "导入表外成本+归档 PRD"),
        ("表外提成扣", "SUM(… 费用性质=提成扣)", "可正负", "同左"),
        ("计算净利润", "calc_net_profit = gross_profit - 表外毛利扣汇总", "提成链路前置", "同左"),
        ("客服分片", "若 业务员=客服 同一账号 → 无业务员客服；否则 有业务员客服", "影响 service_mode 命中", "归档底表+客服场景"),
        ("命中规则(抽象)", "filter(有效规则包) → order by 优先级/互斥 → first match 条件(组织+业务线+service_mode+客户…)", "冲突策略见各场景", "规则包 PRD+场景页"),
        ("阶梯示例", "毛利于 [下限,上上限) 内取对应档比例; 全档或超额档模式见规则", "同场景「阶梯/比例」配置", "提成场景全覆盖"),
        ("主辅/客服", "主辅 split: 基数×比例后分别套规则/上限", "有/无业务员两场景比例不同", "CS_WITH_SALES / CS_NO_SALES"),
        ("渠道", "可拆: 对内提成基数×比例 + 对外结算(协议)", "见 CHANNEL_MANAGER_SETTLE", "场景页"),
        ("试算负毛利/异常", "按规则: 负毛利不计提/自动扣/结转/公司承担/需财务调单", "统一试算筛选项", "统一试算结果页"),
    ]
    for f in formulas:
        ws2.append(f)
    style_header(ws2)
    for r in range(2, ws2.max_row + 1):
        for c in range(1, 5):
            ws2.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 52
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 32
    ws2.freeze_panes = "A2"

    # --- Sheet 3: 试算与异常 ---
    ws3 = wb.create_sheet("3_试算与异常类型", 3)
    ws3.append(["结果/视图项", "含义", "与计算关系", "原型"])
    trial = [
        ("纯正向提成", "基数与比例正常命中，无自动扣项", "标准视图主路径", "统一试算结果-正向与异常"),
        ("自动扣提成", "按规则对提成额自动扣减(如负提成结转类)", "异常处理视图+扩展列", "同左"),
        ("自动扣毛利", "在毛利/基数侧扣减", "同左", "同左"),
        ("公司承担", "不向下游转嫁，实发为0等策略", "负毛利策略一种", "同左 筛选:负毛利策略"),
        ("需财务调整", "系统无法自动闭合，要人工单", "提交财务调整单", "同左"),
        ("负毛利策略-负提成结转扣减", "历史负提成在后续月扣回", "策略枚举", "同左 筛选器"),
        ("负毛利策略-负毛利结转扣减", "毛利为负时结转处理", "策略枚举", "同左"),
        ("重算状态", "已重算/未重算", "便于回归", "同左"),
    ]
    for t in trial:
        ws3.append(t)
    style_header(ws3)
    for r in range(2, ws3.max_row + 1):
        for c in range(1, 5):
            ws3.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 36
    ws3.column_dimensions["C"].width = 32
    ws3.column_dimensions["D"].width = 36
    ws3.freeze_panes = "A2"

    # --- Sheet 4: 8 scenes from 二期页面 ---
    ws4 = wb.create_sheet("4_场景矩阵_二期", 4)
    ws4.append([
        "分组", "场景名", "业务线代码", "角色", "规则编码", "service_mode/要点",
        "基数方案", "基数文字定义", "计算模式", "主辅/渠道", "负毛利/冲突策略摘要", "试算样例(页面)",
        "来源页面",
    ])
    scenes = [
        ("直客-业务", "直客业务员-常规毛利阶梯", "direct", "业务员", "DIRECT_SALES_GP_TIER", "全部场景; 入职等条件", "COMMISSION_GP_STD", "毛利-系统成本-导入成本", "GP_TIERED 三档", "个人100%", "负毛利不计提;冲突阻断", "160000/40%→64000", "二期/提成场景全覆盖配置页.html"),
        ("直客-管理", "直客主管-营收固定比例", "direct", "主管", "DIRECT_MANAGER_REVENUE_RATE", "全部; 月净营收>0", "DIRECT_REVENUE_BASE", "净营收", "REVENUE_RATE 0.15%", "个人100%", "按净营收判断", "500000/0.15%→750", "同左"),
        ("客服-分片", "客服-有业务员客服", "service", "客服", "CS_WITH_SALES_RATE", "有业务员客服; 主辅拆分", "CS_REVENUE_BASE", "净营收", "REV 1.00%", "主70%/客30%", "负营收冲减", "120000/1%→1200", "同左"),
        ("客服-分片", "客服-无业务员客服", "service", "客服", "CS_NO_SALES_RATE", "无业务员客服", "CS_REVENUE_BASE", "净营收", "REV 1.20%", "客服100%", "负营收冲减", "80000/1.2%→960", "同左"),
        ("同行-销售", "同行销售-量级阶梯", "peer", "同行销售", "PEER_SALES_VOLUME_TIER", "账期<=60天", "COMMISSION_GP_PEER", "毛利-专项成本-导入", "GP 三档", "个人;团队池", "负毛利不计提", "90000/30%→27000", "同左"),
        ("渠道-结算", "渠道经理-对内+对外", "channel", "渠道经理", "CHANNEL_MANAGER_SETTLE", "渠道A/B/C", "COMMISSION_GP_CHANNEL", "毛利-返点-导入", "CHANNEL_SETTLEMENT 阶梯", "双轨+可选团队池", "负毛利按协议", "110000/25%→27500", "同左"),
        ("直客-特例", "直客KA专项加点", "direct", "业务员", "DIRECT_KA_BONUS", "KA客户", "COMMISSION_GP_STD", "毛利-成本-导入", "GP_FLAT 5%", "主80%/服20%封顶3万", "负毛利不计提;冲突入异常", "待核对", "同左 状态待完善"),
        ("客服-风险", "客服跨月冲销", "service", "客服", "CS_REVERSE_ADJUST", "有业务员;冲销单", "CS_REVENUE_BASE", "净营收(含冲销)", "REV 1.00% 主辅", "主70%客30%", "允许负值冲减;导入阻断", "-5000→-50 风险", "同左 状态风险"),
    ]
    for s in scenes:
        ws4.append(s)
    style_header(ws4)
    for r in range(2, ws4.max_row + 1):
        for c in range(1, 14):
            ws4.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 14):
        ws4.column_dimensions[get_column_letter(col)].width = 16
    ws4.column_dimensions["B"].width = 26
    ws4.column_dimensions["F"].width = 22
    ws4.column_dimensions["G"].width = 22
    ws4.column_dimensions["M"].width = 32
    ws4.freeze_panes = "A2"

    # --- Sheet 5: 页面索引 ---
    ws5 = wb.create_sheet("5_页面索引", 5)
    ws5.append(["相对路径", "模块", "在链路中的位置"])
    pages = [
        ("月度提成归档与运单明细-MVP.html", "月底归档/运单底表", "步骤2-4,12-13 输出与冻结批次"),
        ("表外导入成本列表-MVP.html", "月度调整主表", "步骤4,11 表外+绩效汇总入口"),
        ("导入表外成本弹窗页-MVP.html", "表外成本导入", "步骤4 明细预检与入库"),
        ("导入绩效系数弹窗页-MVP.html", "绩效系数导入", "步骤11"),
        ("人员映射表-MVP.html", "人员与映射", "步骤5"),
        ("新增/编辑/批量导入人员映射-MVP.html", "映射维护", "步骤5"),
        ("引用规则包-独立弹窗页-MVP.html", "指定规则包", "步骤5-6"),
        ("二期/规则包-规则子母表-MVP.html", "规则包列表", "步骤6-7 查看包与规则"),
        ("二期/编辑规则-MVP.html", "单条规则", "步骤7-8 条件与计算"),
        ("二期/规则包编辑-弹窗页-MVP.html", "规则包主档", "步骤6"),
        ("二期/基数维护-公式构建器.html", "基数方案", "步骤7"),
        ("二期/规则引擎中心列表.html", "规则总览(引擎视角)", "步骤6 辅助"),
        ("二期/提成场景全覆盖配置页.html", "场景与场景试算", "步骤6-8 业务场景全览"),
        ("二期/提成规则统一配置页.html", "统一配置(若启用)", "步骤6-8 配置入口"),
        ("二期/统一试算结果-正向与异常.html", "试算+异常/财务调单", "步骤10-12"),
        ("二期/提成试算结果.html", "试算(另一呈现)", "步骤12"),
        ("二期/计提结果报表-快照追溯-MVP.html", "快照审计", "步骤14"),
        ("index.html", "导航", "入口"),
    ]
    for p in pages:
        ws5.append(p)
    style_header(ws5)
    for r in range(2, ws5.max_row + 1):
        for c in range(1, 4):
            ws5.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.column_dimensions["A"].width = 46
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 44
    ws5.freeze_panes = "A2"

    # --- Sheet 6: 检查清单 ---
    ws6 = wb.create_sheet("6_配置与测试检查", 6)
    ws6.append(["环节", "检查项", "期望", "可参照页面/场景"])
    checks = [
        ("运单底表", "应收、成本、毛利与财务源一致", "与源系统抽样对账", "月底归档与运单底表"),
        ("表外", "同运单毛利扣/提成扣费用性质未混", "试算前汇总=导入明细和", "导入表外成本+列表"),
        ("人员映射", "计薪月在映射生效+提成计算窗口内", "无「未关联规则」误杀或需解释", "人员映射+规则包"),
        ("规则包", "分公司必填; 生效/停用与账期", "计薪月命中已发布包", "规则包子母表"),
        ("规则", "优先级/互斥组/条件无重叠未处理", "保存前无冲突告警或已确认", "编辑规则+场景页"),
        ("基数", "方案引用的系统/导入组件齐全", "缺失策略与预检一致", "基数维护"),
        ("客服分片", "有/无业务员命中不同规则编码", "与底表 service_mode/逻辑一致", "CS_WITH/NO_SALES + 归档底表"),
        ("负毛利/异常", "每种策略有人工样例+自动用例", "试算与扩展视图一致", "统一试算结果"),
        ("试算后", "需财务调单行有单号与责任人", "闭环", "统一试算"),
        ("归档", "已发放后批次冻结;不可回写", "审计可追溯", "月底归档"),
    ]
    for c in checks:
        ws6.append(c)
    style_header(ws6)
    for r in range(2, ws6.max_row + 1):
        for c in range(1, 5):
            ws6.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws6.column_dimensions["A"].width = 16
    ws6.column_dimensions["B"].width = 36
    ws6.column_dimensions["C"].width = 32
    ws6.column_dimensions["D"].width = 32
    ws6.freeze_panes = "A2"

    wb.save(OUT)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
