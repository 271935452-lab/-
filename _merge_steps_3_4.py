# -*- coding: utf-8 -*-
"""Merge Excel scenario sheet columns 步骤3 + 步骤4 into one step; renumber 5–8 -> 4–7."""
import re
import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

SRC = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算.xlsx"
BAK = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_backup_steps.xlsx"
OUT = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并.xlsx"


def dec_step_header(val):
    if not isinstance(val, str):
        return val
    m = re.match(r"^步骤(\d+)(.*)$", val)
    if not m:
        return val
    n = int(m.group(1))
    tail = m.group(2)
    if n >= 5:
        return "步骤" + str(n - 1) + tail
    return val


def combine_cells(a, b):
    if a is None and b is None:
        return None
    if a is None:
        return b
    if b is None:
        return a
    sa = str(a).strip()
    sb = str(b).strip()
    if not sa:
        return sb
    if not sb:
        return sa
    return sa + "；" + sb


def fix_narrative_text(s):
    if not isinstance(s, str):
        return s
    s = s.replace("步骤7/8", "步骤6/7")
    s = s.replace("Σ步骤8", "Σ步骤7")
    s = s.replace("步骤8 =", "步骤7 =")
    return s


def transform_row(lst):
    """lst is 0-based values for columns A.. (length >= 17 for wide rows)."""
    if len(lst) < 13:
        return lst
    is_header = lst[10] == "步骤2·映射∩账期" and lst[11] == "步骤3·提成区间" and lst[12] == "步骤4·状态门禁"
    tail_extra = lst[17:] if len(lst) > 17 else []
    if is_header:
        block = [
            "步骤3·提成区间∩状态门禁",
            dec_step_header(lst[13]),
            dec_step_header(lst[14]),
            dec_step_header(lst[15]),
            dec_step_header(lst[16]),
        ]
    else:
        merged_lm = combine_cells(lst[11], lst[12])
        block = [merged_lm] + lst[13:17]
    new_row = lst[:11] + block + tail_extra
    return new_row


def adjust_merge_ranges(range_strings):
    out = []
    for r_s in range_strings:
        min_col, min_row, max_col, max_row = range_boundaries(r_s)
        if max_col >= 17:
            max_col = 16
        out.append(
            f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        )
    return out


def process_sheet(ws):
    merges = [str(r) for r in ws.merged_cells.ranges]
    for r_s in merges:
        ws.unmerge_cells(r_s)

    max_row = ws.max_row
    max_col = ws.max_column
    mc = max(max_col, 17)

    for r in range(1, max_row + 1):
        lst = []
        for c in range(1, mc + 1):
            cell = ws.cell(r, c)
            lst.append(cell.value)
        new_lst = transform_row(lst) if len(lst) >= 13 else lst
        # narrative fixes on every string cell in row (both raw and transformed columns)
        new_lst = [
            fix_narrative_text(v) if isinstance(v, str) else v for v in new_lst
        ]
        # write back
        for c, val in enumerate(new_lst, start=1):
            ws.cell(r, c).value = val
        for c in range(len(new_lst) + 1, mc + 1):
            ws.cell(r, c).value = None

    for new_rng in adjust_merge_ranges(merges):
        ws.merge_cells(new_rng)


def main():
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=False)
    pat = re.compile(r"^S\d+$")
    for name in wb.sheetnames:
        if pat.match(name):
            process_sheet(wb[name])
    try:
        wb.save(SRC)
        print("Saved:", SRC)
    except PermissionError:
        wb.save(OUT)
        print("原文件被占用，已另存为:", OUT)
    wb.close()
    print("Backup:", BAK)


if __name__ == "__main__":
    main()
