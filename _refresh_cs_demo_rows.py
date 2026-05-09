# -*- coding: utf-8 -*-
"""Refresh existing 顺序包客服 rows with fixed sid + CS demo revenue (业务员应收为 0 时用演示应收52800)."""
import re
import shutil
from openpyxl import load_workbook

SRC = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_step34合并.xlsx"
BAK = r"d:\Cursor\头程项目\人员映射与规则中心_计算场景全量梳理_含底表测算_before_cs_refresh.xlsx"
CS_RATE = 0.0075


def sid_from_sheet(ws):
    t = str(ws.cell(1, 1).value or "")
    m = re.search(r"场景(\d{3})", t)
    if m:
        return "S" + m.group(1)
    m = re.search(r"S(\d{3})", t)
    if m:
        return "S" + m.group(1)
    return "S000"


def find_hdr_row(ws):
    for r in range(1, 35):
        v = ws.cell(r, 10).value
        if v and isinstance(v, str) and "步骤1" in v:
            return r
    return None


def demo_rev_waybill(ws, sales_row, sid):
    try:
        raw = float(ws.cell(sales_row, 5).value)
        if raw == 0:
            raw = None
    except (TypeError, ValueError):
        raw = None
    rev = raw if raw else 52800.0
    wb = ws.cell(sales_row, 3).value
    if wb and isinstance(wb, str) and wb.strip() not in ("", "—"):
        waybill = wb.strip()
    else:
        waybill = f"WB-DIRECT-{sid}-202603001"
    return rev, waybill


def find_total_row(ws, hdr_row):
    for r in range(hdr_row + 1, hdr_row + 20):
        v = ws.cell(r, 1).value
        if v and isinstance(v, str) and "合计" in v:
            return r
        v = ws.cell(r, 10).value
        if v and isinstance(v, str) and "Σ步骤" in v:
            return r
    return None


def read_sales_final(ws, r):
    for col in (17, 16):
        v = ws.cell(r, col).value
        try:
            if v is not None and str(v).strip() != "":
                return float(v)
        except (TypeError, ValueError):
            continue
    return 0.0


def patch_sales_prefix(ws, r, cs_amt, pkg):
    prefix_rule = f"前置段1客服计提 {cs_amt:.2f} 已完成（见上行·{pkg}）；"
    cur = str(ws.cell(r, 14).value or "")
    cur = re.sub(
        r"^【顺序执行·段2】前置段1客服计提 [\d.]+\s*已完成（见上行·RP_S\d+_SEQ）；",
        "",
        cur,
    )
    cur = re.sub(
        r"^【顺序执行·段2】段1客服计提已完成（[\d.]+）→",
        "",
        cur,
    )
    cur = re.sub(
        r"^前置段1客服计提 [\d.]+\s*已完成（见上行·RP_S\d+_SEQ）；",
        "",
        cur.strip(),
    )
    ws.cell(r, 13).value = f"【顺序执行·段2】规则包 {pkg}"
    if not cur.startswith("前置段1客服计提"):
        cur = prefix_rule + cur
    ws.cell(r, 14).value = cur


def patch_total_j(ws, total_row, cs_amt, sales_final):
    ws.cell(total_row, 10).value = (
        f"Σ步骤8（业务员本人合计）= {sales_final:.2f} CNY。"
        f"顺序包段1客服应发 {cs_amt:.2f} 记入上行客服演示账号，不与业务员本人合计相加。"
    )


def unmerge_bad_step_merges(ws, hdr):
    """解除误合并的业务员步骤区（单行 J:Q 且非合计行），以便逐步写入步骤列。"""
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row != rng.max_row:
            continue
        if rng.min_col != 10 or rng.max_col != 17:
            continue
        r = rng.min_row
        if r <= hdr + 1:
            continue
        a = ws.cell(r, 1).value
        if a and isinstance(a, str) and "合计" in a:
            continue
        ws.unmerge_cells(str(rng))


def main():
    shutil.copy2(SRC, BAK)
    wb = load_workbook(SRC, data_only=False)
    idx = wb[wb.sheetnames[0]]
    ids = []
    for r in range(2, idx.max_row + 1):
        role = idx.cell(r, 5).value
        sid = idx.cell(r, 7).value
        if not role or not sid:
            continue
        rs = str(role)
        if "直客业务员" not in rs or "主管" in rs:
            continue
        ids.append(str(sid))
    seen = set()
    for sid_sheet in ids:
        if sid_sheet in seen:
            continue
        seen.add(sid_sheet)
        ws = wb[sid_sheet]
        hdr = find_hdr_row(ws)
        if not hdr:
            continue
        if ws.cell(hdr + 1, 4).value != "客服":
            continue
        unmerge_bad_step_merges(ws, hdr)
        rsid = sid_from_sheet(ws)
        pkg = f"RP_{rsid}_SEQ"
        nick = f"顺序包客服_{rsid}"
        name_b = "（演示）顺序执行·段1"
        rev, waybill = demo_rev_waybill(ws, hdr + 2, rsid)
        cs_amt = round(rev * CS_RATE, 2)
        r = hdr + 1
        ws.cell(r, 1).value = nick
        ws.cell(r, 2).value = name_b
        ws.cell(r, 3).value = waybill
        ws.cell(r, 5).value = rev
        ws.cell(r, 10).value = f"命中映射「{nick}+{name_b}」（示意）"
        ws.cell(r, 13).value = f"【顺序执行·段1】规则包 {pkg}"
        ws.cell(r, 14).value = "命中 CS_WITH_SALES_RATE（客服计提先行）"
        ws.cell(r, 15).value = f"基数=本运单归因应收 {rev:.2f}（净应收汇总示意）"
        ws.cell(r, 16).value = cs_amt
        ws.cell(r, 17).value = cs_amt

        total_row = find_total_row(ws, hdr)
        sales_sum = 0.0
        sr_end = (total_row - 1) if total_row else hdr + 2
        for rr in range(hdr + 2, sr_end + 1):
            if ws.cell(rr, 4).value == "客服":
                continue
            patch_sales_prefix(ws, rr, cs_amt, pkg)
            sales_sum += read_sales_final(ws, rr)
        disp = sales_sum if sales_sum != 0 else read_sales_final(ws, hdr + 2)
        if total_row:
            patch_total_j(ws, total_row, cs_amt, disp)

    wb.save(SRC)
    wb.close()
    print("Saved", SRC)


if __name__ == "__main__":
    main()
